from datetime import datetime
import io
import os
import shutil
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import List, Optional
from database import get_db, DB_DIR
import models
import schemas

router = APIRouter()

PHOTO_DIR = os.path.join(DB_DIR, "..", "backend", "static", "photos")
os.makedirs(PHOTO_DIR, exist_ok=True)

@router.get("/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    total_vehicles = db.query(func.count(func.distinct(models.PDIKayit.sasi_no))).scalar() or 0
    total_errors = db.query(func.count(models.PDIKayit.id)).scalar() or 0
    
    # Simple breakdown by type
    type_counts = db.query(models.PDIKayit.arac_tipi, func.count(func.distinct(models.PDIKayit.sasi_no)))\
        .group_by(models.PDIKayit.arac_tipi).all()
    
    breakdown = {t: c for t, c in type_counts if t}
    
    return {
        "total_vehicles": total_vehicles,
        "total_errors": total_errors,
        "breakdown": breakdown
    }

@router.get("/kayitlar", response_model=List[schemas.PDIKayitDetail])
def get_records(
    limit: int = 500,
    offset: int = 0,
    arac_tipi: Optional[str] = None,
    sasi_no: Optional[str] = None,
    alt_grup: Optional[str] = None,
    ay: Optional[int] = None,
    yil: Optional[int] = None,
    hata_nerede: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.PDIKayit)
    if arac_tipi:
        query = query.filter(models.PDIKayit.arac_tipi == arac_tipi)
    if sasi_no:
        query = query.filter(models.PDIKayit.sasi_no.contains(sasi_no))
    if alt_grup:
        query = query.filter(models.PDIKayit.alt_grup == alt_grup)
    if hata_nerede:
        query = query.filter(models.PDIKayit.hata_nerede == hata_nerede)
    if ay and yil:
        m_str = f"{ay:02d}"
        y_str = str(yil)
        query = query.filter(or_(
            (func.substr(models.PDIKayit.tarih_saat, 4, 2) == m_str) & (func.substr(models.PDIKayit.tarih_saat, 7, 4) == y_str),
            (func.substr(models.PDIKayit.tarih_saat, 6, 2) == m_str) & (func.substr(models.PDIKayit.tarih_saat, 1, 4) == y_str)
        ))
    elif yil:
        y_str = str(yil)
        query = query.filter(or_(
            func.substr(models.PDIKayit.tarih_saat, 7, 4) == y_str,
            func.substr(models.PDIKayit.tarih_saat, 1, 4) == y_str
        ))

    return query.order_by(models.PDIKayit.id.desc()).offset(offset).limit(limit).all()

@router.get("/kayitlar/export")
def export_records_excel(
    arac_tipi: Optional[str] = None,
    sasi_no: Optional[str] = None,
    alt_grup: Optional[str] = None,
    ay: Optional[int] = None,
    yil: Optional[int] = None,
    hata_nerede: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    query = db.query(models.PDIKayit)
    if arac_tipi:
        query = query.filter(models.PDIKayit.arac_tipi == arac_tipi)
    if sasi_no:
        query = query.filter(models.PDIKayit.sasi_no.contains(sasi_no))
    if alt_grup:
        query = query.filter(models.PDIKayit.alt_grup == alt_grup)
    if hata_nerede:
        query = query.filter(models.PDIKayit.hata_nerede == hata_nerede)
    if ay and yil:
        m_str = f"{ay:02d}"
        y_str = str(yil)
        query = query.filter(or_(
            (func.substr(models.PDIKayit.tarih_saat, 4, 2) == m_str) & (func.substr(models.PDIKayit.tarih_saat, 7, 4) == y_str),
            (func.substr(models.PDIKayit.tarih_saat, 6, 2) == m_str) & (func.substr(models.PDIKayit.tarih_saat, 1, 4) == y_str)
        ))
    elif yil:
        y_str = str(yil)
        query = query.filter(or_(
            func.substr(models.PDIKayit.tarih_saat, 7, 4) == y_str,
            func.substr(models.PDIKayit.tarih_saat, 1, 4) == y_str
        ))
    records = query.order_by(models.PDIKayit.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PDI Kayıtları"

    headers = ["ID", "BB No", "Şasi No", "Araç", "Tarih", "Alt Grup", "Hata", "Top Hata", "Hata Nerede Giderildi", "Ekleyen"]
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(records, 2):
        ws.cell(row=row, column=1, value=r.id)
        ws.cell(row=row, column=2, value=r.bb_no or "")
        ws.cell(row=row, column=3, value=r.sasi_no or "")
        ws.cell(row=row, column=4, value=r.arac_tipi or "")
        ws.cell(row=row, column=5, value=r.tarih_saat or "")
        ws.cell(row=row, column=6, value=r.alt_grup or "")
        ws.cell(row=row, column=7, value=r.hata_konumu or "")
        ws.cell(row=row, column=8, value=r.top_hata or "")
        ws.cell(row=row, column=9, value=r.hata_nerede or "")
        ws.cell(row=row, column=10, value=r.kullanici or "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pdi_kayitlari.xlsx"}
    )

@router.put("/kayit/{record_id}", response_model=schemas.PDIKayitDetail)
def update_record(record_id: int, obj_in: schemas.PDIKayitUpdate, db: Session = Depends(get_db)):
    db_obj = db.query(models.PDIKayit).filter(models.PDIKayit.id == record_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    update_data = obj_in.model_dump(exclude_unset=True)
    for field in update_data:
        setattr(db_obj, field, update_data[field])
    
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj

@router.delete("/kayit/{record_id}")
def delete_record(record_id: int, db: Session = Depends(get_db)):
    db_obj = db.query(models.PDIKayit).filter(models.PDIKayit.id == record_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    db.delete(db_obj)
    db.commit()
    return {"message": "Kayıt silindi"}

@router.post("/kayit/{record_id}/photo")
def update_record_photo(record_id: int, photo: UploadFile = File(...), db: Session = Depends(get_db)):
    db_obj = db.query(models.PDIKayit).filter(models.PDIKayit.id == record_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    
    ext = os.path.splitext(photo.filename)[1] or ".jpg"
    filename = f"edit_{record_id}_{datetime.now().strftime('%Y%H%M%S')}{ext}"
    filepath = os.path.join(PHOTO_DIR, filename)
    
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(photo.file, buffer)
    
    db_obj.fotograf_yolu = filepath
    db.commit()
    db.refresh(db_obj)
    
    return {"photo_path": filepath, "message": "Fotoğraf güncellendi"}

@router.get("/top-hatalar", response_model=List[schemas.TopHataSchema])
def get_top_hatalar(db: Session = Depends(get_db)):
    return db.query(models.TopHata).filter(models.TopHata.aktif == 1).all()

@router.post("/manual-data")
def update_manual_data(data: schemas.ManualDataUpdate, db: Session = Depends(get_db)):
    db_obj = db.query(models.ReportManualData).filter(
        models.ReportManualData.report_type == data.report_type,
        models.ReportManualData.context_key == data.context_key,
        models.ReportManualData.data_key == data.data_key
    ).first()
    
    if db_obj:
        db_obj.data_value = data.data_value
    else:
        db_obj = models.ReportManualData(**data.model_dump())
        db.add(db_obj)
    
    db.commit()

    # Auto-calculate rate if vehicle_count or error_count is updated
    if data.data_key in ["vehicle_count", "error_count"]:
        v_obj = db.query(models.ReportManualData).filter(
            models.ReportManualData.report_type == data.report_type,
            models.ReportManualData.context_key == data.context_key,
            models.ReportManualData.data_key == "vehicle_count"
        ).first()
        e_obj = db.query(models.ReportManualData).filter(
            models.ReportManualData.report_type == data.report_type,
            models.ReportManualData.context_key == data.context_key,
            models.ReportManualData.data_key == "error_count"
        ).first()

        if v_obj and e_obj and float(v_obj.data_value or 0) > 0:
            rate = float(e_obj.data_value) / float(v_obj.data_value)
            
            r_obj = db.query(models.ReportManualData).filter(
                models.ReportManualData.report_type == data.report_type,
                models.ReportManualData.context_key == data.context_key,
                models.ReportManualData.data_key == "error_rate"
            ).first()
            
            if r_obj:
                r_obj.data_value = str(round(rate, 4))
            else:
                new_rate = models.ReportManualData(
                    report_type=data.report_type,
                    context_key=data.context_key,
                    data_key="error_rate",
                    data_value=str(round(rate, 4))
                )
                db.add(new_rate)
            db.commit()

    return {"message": "Veri güncellendi"}

# --- Lookup Management ---

LOOKUP_MODELS = {
    "arac-tipi": models.AracTipiLookUp,
    "alt-grup": models.AltGrupLookUp,
    "hata-konumu": models.HataKonumuLookUp,
    "hata-nerede": models.HataNeredeLookUp
}

@router.get("/lookups/{type}", response_model=List[schemas.LookUpSchema])
def get_lookups(type: str, db: Session = Depends(get_db)):
    if type not in LOOKUP_MODELS:
        raise HTTPException(status_code=404, detail="Gecersiz lookup tipi")
    return db.query(LOOKUP_MODELS[type]).all()

@router.post("/lookups/{type}", response_model=schemas.LookUpSchema)
def add_lookup(type: str, obj_in: schemas.LookUpCreate, db: Session = Depends(get_db)):
    if type not in LOOKUP_MODELS:
        raise HTTPException(status_code=404, detail="Gecersiz lookup tipi")
    db_obj = LOOKUP_MODELS[type](name=obj_in.name)
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj

@router.delete("/lookups/{type}/{id}")
def delete_lookup(type: str, id: int, db: Session = Depends(get_db)):
    if type not in LOOKUP_MODELS:
        raise HTTPException(status_code=404, detail="Gecersiz lookup tipi")
    db_obj = db.query(LOOKUP_MODELS[type]).filter(LOOKUP_MODELS[type].id == id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Kayit bulunamadi")
    db.delete(db_obj)
    db.commit()
    return {"message": "Kayit silindi"}

# --- Manufacturing Progress Tracking ---

@router.get("/imalat-takip", response_model=List[schemas.ImalatHataTakipSchema])
def get_imalat_takip(db: Session = Depends(get_db)):
    return db.query(models.ImalatHataTakip).order_by(models.ImalatHataTakip.id.desc()).all()

@router.post("/imalat-takip", response_model=schemas.ImalatHataTakipSchema)
def add_imalat_takip(obj_in: schemas.ImalatHataTakipCreate, db: Session = Depends(get_db)):
    db_obj = models.ImalatHataTakip(
        **obj_in.model_dump(),
        olusturma_tarihi=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj

@router.put("/imalat-takip/{item_id}", response_model=schemas.ImalatHataTakipSchema)
def toggle_imalat_takip(item_id: int, db: Session = Depends(get_db)):
    db_obj = db.query(models.ImalatHataTakip).filter(models.ImalatHataTakip.id == item_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Öğe bulunamadı")
    db_obj.cozuldu = 1 if db_obj.cozuldu == 0 else 0
    db.commit()
    db.refresh(db_obj)
    return db_obj

# --- Manuel Cozulenler ---

@router.get("/manuel-cozulenler", response_model=List[schemas.ManuelCozulenSchema])
def get_manuel_cozulenler(db: Session = Depends(get_db)):
    return db.query(models.ManuelCozulen).order_by(models.ManuelCozulen.id.asc()).all()

@router.post("/manuel-cozulenler", response_model=schemas.ManuelCozulenSchema)
def add_manuel_cozulen(obj_in: schemas.ManuelCozulenCreate, db: Session = Depends(get_db)):
    db_obj = models.ManuelCozulen(hata=obj_in.hata, donem=obj_in.donem)
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj

@router.delete("/manuel-cozulenler/{item_id}")
def delete_manuel_cozulen(item_id: int, db: Session = Depends(get_db)):
    db_obj = db.query(models.ManuelCozulen).filter(models.ManuelCozulen.id == item_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    db.delete(db_obj)
    db.commit()
    return {"message": "Kayıt silindi"}

# --- PDI Detayları (per-record tespit details) ---

@router.get("/kayit/{record_id}/detaylar", response_model=List[schemas.PDIDetaySchema])
def get_record_detaylar(record_id: int, db: Session = Depends(get_db)):
    return db.query(models.PDIDetay).filter(models.PDIDetay.pdi_id == record_id).all()

@router.post("/kayit/{record_id}/detay", response_model=schemas.PDIDetaySchema)
def add_record_detay(record_id: int, obj_in: schemas.PDIDetayCreate, db: Session = Depends(get_db)):
    db_obj = models.PDIDetay(pdi_id=record_id, aciklama=obj_in.aciklama)
    db.add(db_obj)
    db.commit()
    db.refresh(db_obj)
    return db_obj

@router.delete("/kayit/detay/{detay_id}")
def delete_record_detay(detay_id: int, db: Session = Depends(get_db)):
    db_obj = db.query(models.PDIDetay).filter(models.PDIDetay.id == detay_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Detay bulunamadı")
    db.delete(db_obj)
    db.commit()
    return {"message": "Detay silindi"}

@router.post("/kayit/detay/{detay_id}/photo")
def upload_detay_photo(detay_id: int, photo: UploadFile = File(...), db: Session = Depends(get_db)):
    db_obj = db.query(models.PDIDetay).filter(models.PDIDetay.id == detay_id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Detay bulunamadı")
    ext = os.path.splitext(photo.filename)[1] or ".jpg"
    filename = f"detay_{detay_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
    filepath = os.path.join(PHOTO_DIR, filename)
    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(photo.file, buffer)
    db_obj.fotograf_yolu = filename
    db.commit()
    db.refresh(db_obj)
    return {"photo_path": filename, "message": "Fotoğraf yüklendi"}

# --- Manual data get for all months (for edit modal) ---
@router.get("/manual-data/{report_type}")
def get_manual_data(report_type: str, db: Session = Depends(get_db)):
    items = db.query(models.ReportManualData).filter(
        models.ReportManualData.report_type == report_type
    ).all()
    result = {}
    for item in items:
        key = f"{item.context_key}_{item.data_key}"
        result[key] = item.data_value
    return result
@router.post("/kayitlar/import")
async def import_records_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas not installed")

    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents), dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    def str_norm(s):
        if not s: return ""
        s = str(s).lower()
        mapping = {'ç':'c','ğ':'g','ı':'i','ö':'o','ş':'s','ü':'u','â':'a','î':'i','û':'u'}
        for k, v in mapping.items():
            s = s.replace(k, v)
        return "".join(c for c in s if 'a' <= c <= 'z' or '0' <= c <= '9').upper()

    cols_map = {
        "bb_no": ["BB No", "BB NO", "BBNO"],
        "sasi_no": ["Şasi No", "ŞASİ NO", "SASİ NO", "Sasi No", "SASI NO", "SASI", "CHASSIS"],
        "arac_tipi": ["Araç Tipi", "ARAÇ TİPİ", "ARAC TIPI", "TİP", "ARAC", "TIP", "MODEL"],
        "is_emri_no": ["İş Emri No", "İŞ EMRİ NO", "IS EMRI NO", "İŞ EMRİ", "IS EMRI", "IS EMRE", "ISEMRI"],
        "tarih_saat": ["PDI Tarihi", "PDI TARİHİ", "PDI Yapılış Tarihi", "Tarih", "TARİH", "YAPILIŞ TARİHİ", "PDI DATE", "DATE"],
        "tespitler": ["Tespitler", "TESPİTLER", "TESPITLER", "HATA", "TESPİT", "FINDINGS", "DESCRIPTION"],
        "hata_konumu": ["Hata Konumu", "HATA KONUMU", "KONUM", "HATA YERİ", "LOCATION"],
        "alt_grup": ["Alt Grup", "ALT GRUP", "GRUP", "SUBGROUP"]
    }

    type_map = {
        "TOU": "Tourismo", "TRV": "Travego", "CON": "Conecto",
        "TOURISMO": "Tourismo", "TRAVEGO": "Travego", "CONNECTO": "Conecto",
        "TOURİSMO": "Tourismo", "CONECTO": "Conecto"
    }

    found_indices = {}
    df_cols_norm = [str_norm(c) for c in df.columns]
    for key, variations in cols_map.items():
        norm_vars = [str_norm(v) for v in variations]
        for idx, norm_col_name in enumerate(df_cols_norm):
            if norm_col_name in norm_vars:
                found_indices[key] = idx
                break

    critical_missing = []
    if "sasi_no" not in found_indices: critical_missing.append("Şasi No")
    if "arac_tipi" not in found_indices: critical_missing.append("Araç Tipi")
    if "tarih_saat" not in found_indices: critical_missing.append("Tarih")
    
    if critical_missing:
        raise HTTPException(status_code=400, detail=f"Excel'de şu kritik sütunlar bulunamadı: {', '.join(critical_missing)}")

    def get_val(row_vals, key, default=""):
        idx = found_indices.get(key)
        if idx is not None and idx < len(row_vals):
            v = row_vals[idx]
            return str(v).strip() if pd.notna(v) and str(v).lower() != 'nan' else default
        return default

    success_count = 0
    duplicate_count = 0
    
    for idx, row in df.iterrows():
        row_vals = row.values
        sasi_no = get_val(row_vals, "sasi_no")
        if not sasi_no: continue

        raw_type = get_val(row_vals, "arac_tipi")
        arac_tipi = type_map.get(raw_type.upper(), raw_type)
        if arac_tipi not in ["Tourismo", "Travego", "Conecto"]:
            continue

        raw_date = get_val(row_vals, "tarih_saat")
        # Normalize date to DD-MM-YYYY
        new_date = raw_date.replace("/", "-").replace(".", "-")
        parts = new_date.split("-")
        try:
            if len(parts) == 3:
                if len(parts[0]) == 4: # YYYY-MM-DD
                    new_date = f"{parts[2].zfill(2)}-{parts[1].zfill(2)}-{parts[0]}"
                else: # D-M-YYYY
                    new_date = f"{parts[0].zfill(2)}-{parts[1].zfill(2)}-{parts[2]}"
        except: pass

        tespitler = get_val(row_vals, "tespitler")
        
        # Check duplicate
        existing = db.query(models.PDIKayit).filter(
            models.PDIKayit.sasi_no == sasi_no,
            models.PDIKayit.tarih_saat == new_date,
            models.PDIKayit.tespitler == tespitler
        ).first()

        if existing:
            duplicate_count += 1
            continue

        new_record = models.PDIKayit(
            bb_no=get_val(row_vals, "bb_no"),
            sasi_no=sasi_no,
            arac_tipi=arac_tipi,
            is_emri_no=get_val(row_vals, "is_emri_no"),
            alt_grup=get_val(row_vals, "alt_grup"),
            tespitler=tespitler,
            hata_konumu=get_val(row_vals, "hata_konumu"),
            tarih_saat=new_date,
            kullanici="Excel Import"
        )
        db.add(new_record)
        success_count += 1

    db.commit()
    return {"message": f"{success_count} kayıt eklendi. {duplicate_count} mükerrer kayıt atlandı."}
