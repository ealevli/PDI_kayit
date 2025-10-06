# PDI Kayıt Sistemi – Streamlit
# Gereken paketler: streamlit, sqlalchemy, psycopg2-binary, pandas, openpyxl

import os
from io import BytesIO
from datetime import date, timedelta

import pandas as pd
import streamlit as st
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from urllib.parse import quote_plus

# ------------------------------------------------------
# Sabitler
# ------------------------------------------------------
ALT_GRUP = ["Boya", "Süsleme", "Mekanik", "Elektrik"]
HATA_KONUM = [
    "Aydınlatma","Ayırma Duvarı","Ayna","Boya","CAM","Çıta","Defroster","Etiket",
    "Kapak","Kapı","Kaplama","Kelepçe","Klima","Koltuk","Körük","Lamba","Montaj",
    "Motor","Mutfak","Paket Raf","Silecek","Siperlik","Stepne","Şoför yatma yeri",
    "Tapa","Telefon","Torpido","Yazı"
]
ARAC_TIPI = ["Tourismo", "Connecto", "Travego"]

st.set_page_config(page_title="PDI Kayıt Sistemi", layout="wide")

# ------------------------------------------------------
# Secrets'ten DB URL oluşturma (db_url veya [db] bloğu)
# ------------------------------------------------------
def build_db_url_and_args():
    # Parçalı secrets (önerilen)
    if "db" in st.secrets:
        s = st.secrets["db"]
        user = s.get("user", "postgres")
        pwd  = quote_plus(s.get("password", ""))
        host = s["host"]
        port = s.get("port", "6543")        # Supabase pooler
        name = s.get("name", "postgres")
        ssl  = s.get("sslmode", "require")
        url = f"postgresql+psycopg2://{user}:{pwd}@{host}:{port}/{name}?sslmode={ssl}"
        return url, {}
    # Tek satır DSN (db_url) ya da ortam değişkeni
    url = st.secrets.get("db_url", os.getenv("DB_URL", ""))
    return url, {}

DB_URL, CONNECT_ARGS = build_db_url_and_args()
if not DB_URL:
    st.error("Veritabanı bağlantısı bulunamadı. secrets.toml içinde db_url ya da [db] girin.")
    st.stop()

# ------------------------------------------------------
# Güvenli bağlantı kur (önce 6543, olmazsa 5432)
# ------------------------------------------------------
def try_connect(url, connect_args):
    eng = create_engine(url, pool_pre_ping=True, connect_args=connect_args)
    with eng.connect() as c:
        c.execute(text("SELECT 1"))
    return eng

try:
    engine = try_connect(DB_URL, CONNECT_ARGS)
except OperationalError:
    if ":6543/" in DB_URL:
        alt = DB_URL.replace(":6543/", ":5432/")
        try:
            engine = try_connect(alt, CONNECT_ARGS)
        except OperationalError:
            st.error("Veritabanına bağlanılamadı. Host/port/SSL ve kullanıcı bilgilerini kontrol edin "
                     "(Supabase için pooler: port 6543, user: postgres.<project-ref>, sslmode=require).")
            st.stop()
    else:
        st.error("Veritabanına bağlanılamadı. Host/port/SSL ve kullanıcı bilgilerini kontrol edin.")
        st.stop()

# ------------------------------------------------------
# İlk kurulum (tablolar & admin kullanıcı)
# ------------------------------------------------------
def init_db():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users(
                username TEXT PRIMARY KEY,
                password TEXT,
                role INT,
                aciklama TEXT
            );
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS pdi_kayitlari(
                id SERIAL PRIMARY KEY,
                bb_no TEXT,
                sasi_no TEXT,
                arac_tipi TEXT,
                is_emri_no TEXT,
                alt_grup TEXT,
                tespitler TEXT,
                hata_konumu TEXT,
                fotograf_yolu TEXT,
                tarih_saat TEXT,
                kullanici TEXT
            );
        """))
        conn.execute(text("""
            INSERT INTO users(username,password,role,aciklama)
            VALUES ('admin','admin123',1,'Sistem Yöneticisi')
            ON CONFLICT (username) DO NOTHING;
        """))

init_db()

# ------------------------------------------------------
# Giriş
# ------------------------------------------------------
def login_form():
    st.title("PDI Kayıt Sistemi – Giriş")
    u = st.text_input("Kullanıcı adı", key="login_u")
    p = st.text_input("Şifre", type="password", key="login_p")
    if st.button("Giriş"):
        with engine.begin() as conn:
            row = conn.execute(
                text("SELECT role FROM users WHERE username=:u AND password=:p"),
                {"u": u, "p": p}
            ).fetchone()
        if row:
            st.session_state["user"] = u
            st.session_state["role"] = "admin" if row[0] == 1 else "viewer"
            st.rerun()
        else:
            st.error("Kullanıcı adı veya şifre hatalı.")

if "user" not in st.session_state:
    login_form()
    st.stop()

user = st.session_state["user"]
role = st.session_state.get("role", "viewer")

with st.sidebar:
    st.markdown(f"**Giriş yapan:** {user} ({role})")
    if st.button("Çıkış"):
        for k in ("user", "role"):
            st.session_state.pop(k, None)
        st.rerun()

st.title("PDI Kayıtları")

# ------------------------------------------------------
# Filtreler
# ------------------------------------------------------
c1, c2, c3, c4, c5, c6 = st.columns([1.2, 1.5, 1.2, 1.5, 1.5, 1.2])
f_sasi = c1.text_input("Şasi No (içeren)")
d1, d2 = c2.date_input(
    "Tarih Aralığı",
    (date.today() - timedelta(days=7), date.today())
)
f_alt = c3.selectbox("Alt Grup", ["Tümü"] + ALT_GRUP)
f_hata = c4.selectbox("Hata Konumu", ["Tümü"] + HATA_KONUM)
f_kul  = c5.text_input("Kullanıcı (boş=hepsi)")
f_arac = c6.selectbox("Araç Tipi", ["Tümü"] + ARAC_TIPI)

# ------------------------------------------------------
# Veriyi çek
# ------------------------------------------------------
where = ["1=1"]
params = {
    "d1": d1.strftime("%Y%m%d"),
    "d2": d2.strftime("%Y%m%d")
}
if f_sasi:
    where.append("sasi_no ILIKE :sasi")
    params["sasi"] = f"%{f_sasi}%"
if f_alt != "Tümü":
    where.append("alt_grup = :alt")
    params["alt"] = f_alt
if f_hata != "Tümü":
    where.append("hata_konumu ILIKE :hk")
    params["hk"] = f"%{f_hata}%"
if f_kul:
    where.append("kullanici = :u")
    params["u"] = f_kul
if f_arac != "Tümü":
    where.append("arac_tipi = :arac")
    params["arac"] = f_arac

# tarih_saat 'dd-MM-YYYY HH:MM:SS' -> YYYYMMDD
where.append("(substr(tarih_saat,7,4)||substr(tarih_saat,4,2)||substr(tarih_saat,1,2)) BETWEEN :d1 AND :d2")

SQL_LIST = f"""
SELECT id,
       bb_no AS "BB No",
       sasi_no AS "Şasi No",
       arac_tipi AS "Araç Tipi",
       is_emri_no AS "İş Emri No",
       tarih_saat AS "PDI Yapılış Tarihi",
       tespitler AS "Tespitler",
       hata_konumu AS "Hata Konumu",
       alt_grup AS "Alt Grup",
       kullanici AS "Kullanıcı",
       fotograf_yolu
FROM pdi_kayitlari
WHERE {' AND '.join(where)}
ORDER BY id DESC;
"""

with engine.begin() as conn:
    df = pd.read_sql(text(SQL_LIST), conn, params=params)

# ------------------------------------------------------
# Tablo
# ------------------------------------------------------
st.dataframe(df.drop(columns=["fotograf_yolu"]), use_container_width=True, height=420)

# ------------------------------------------------------
# Excel indir
# ------------------------------------------------------
def df_to_xlsx_bytes(frame: pd.DataFrame) -> bytes:
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "PDI"
    cols = ["BB No","Şasi No","Araç Tipi","İş Emri No","PDI Yapılış Tarihi",
            "Tespitler","Hata Konumu","Alt Grup"]
    ws.append(cols)
    for row in frame[cols].itertuples(index=False):
        ws.append(list(row))
    header_fill = PatternFill("solid", fgColor="3F4C5C")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

st.download_button(
    "Excel indir",
    data=df_to_xlsx_bytes(df),
    file_name="pdi_kayitlari.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.markdown("---")

# ------------------------------------------------------
# Yeni Kayıt Ekle (isteğe bağlı: herkes ekleyebilir; istersen role=='admin' yap)
# ------------------------------------------------------
with st.expander("➕ Yeni Kayıt Ekle"):
    col1, col2, col3 = st.columns(3)
    bb_no = col1.text_input("BB No")
    sasi_no = col2.text_input("Şasi No")
    arac = col3.selectbox("Araç Tipi", ARAC_TIPI)

    col4, col5, col6 = st.columns(3)
    is_emri = col4.text_input("İş Emri No")
    pdi_tarih = col5.date_input("PDI Yapılış Tarihi", value=date.today())
    alt_grup = col6.selectbox("Alt Grup", ALT_GRUP)

    tespitler = st.text_area("Tespitler", height=120)
    hata_konumu_sel = st.multiselect("Hata Konumu", HATA_KONUM)
    foto_urls = st.text_input("Fotoğraf URL'leri (virgülle)")

    if st.button("Kaydı Ekle"):
        if not sasi_no:
            st.warning("Şasi No gerekli.")
        else:
            with engine.begin() as conn:
                conn.execute(text("""
                    INSERT INTO pdi_kayitlari
                    (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler,
                     hata_konumu, fotograf_yolu, tarih_saat, kullanici)
                    VALUES (:bb, :sasi, :arac, :isn, :alt, :t, :hk, :fp, :ts, :kul)
                """), {
                    "bb": bb_no.strip(),
                    "sasi": sasi_no.strip(),
                    "arac": arac,
                    "isn": is_emri.strip(),
                    "alt": alt_grup,
                    "t": tespitler.strip(),
                    "hk": ", ".join(hata_konumu_sel),
                    "fp": foto_urls.strip(),
                    "ts": pdi_tarih.strftime("%d-%m-%Y") + " 00:00:00",
                    "kul": user
                })
            st.success("Kayıt eklendi.")
            st.rerun()

# ------------------------------------------------------
# Düzenleme (yalnızca admin)
# ------------------------------------------------------
st.subheader("Kayıt Düzenle")
if role != "admin":
    st.info("Kayıt düzenleme yalnızca admin için açıktır.")
else:
    if df.empty:
        st.warning("Düzenlenecek kayıt yok.")
    else:
        rid = st.selectbox("Kayıt seç (ID)", df["id"])
        rec = df[df["id"] == rid].iloc[0]

        c1, c2, c3 = st.columns(3)
        e_bb   = c1.text_input("BB No", rec["BB No"] or "")
        e_sasi = c2.text_input("Şasi No", rec["Şasi No"] or "")
        # Araç tipi
        idx_arac = ARAC_TIPI.index(rec["Araç Tipi"]) if rec["Araç Tipi"] in ARAC_TIPI else 0
        e_arac = c3.selectbox("Araç Tipi", ARAC_TIPI, index=idx_arac)

        c4, c5, c6 = st.columns(3)
        e_isemri = c4.text_input("İş Emri No", rec["İş Emri No"] or "")
        # Tarih (dd-MM-YYYY HH:MM:SS -> date)
        ds = (rec["PDI Yapılış Tarihi"] or "01-01-2000").split(" ")[0]
        d_default = pd.to_datetime(ds, format="%d-%m-%Y", errors="coerce").date() if ds else date.today()
        e_tarih = c5.date_input("PDI Yapılış Tarihi", value=d_default)
        idx_alt = ALT_GRUP.index(rec["Alt Grup"]) if rec["Alt Grup"] in ALT_GRUP else 0
        e_alt = c6.selectbox("Alt Grup", ALT_GRUP, index=idx_alt)

        e_tespit = st.text_area("Tespitler", rec["Tespitler"] or "", height=120)
        mevcut_hk = [h.strip() for h in (rec["Hata Konumu"] or "").split(",") if h.strip()]
        e_hk = st.multiselect("Hata Konumu", HATA_KONUM, default=[h for h in mevcut_hk if h in HATA_KONUM])

        col_btn1, col_btn2 = st.columns([1,1])
        if col_btn1.button("Güncelle"):
            ts = e_tarih.strftime("%d-%m-%Y") + " 00:00:00"
            with engine.begin() as conn:
                conn.execute(text("""
                    UPDATE pdi_kayitlari SET
                      bb_no=:bb, sasi_no=:sasi, arac_tipi=:arac, is_emri_no=:isn,
                      alt_grup=:alt, tespitler=:t, hata_konumu=:hk, tarih_saat=:ts
                    WHERE id=:id
                """), {
                    "bb": e_bb.strip(), "sasi": e_sasi.strip(), "arac": e_arac,
                    "isn": e_isemri.strip(), "alt": e_alt, "t": e_tespit.strip(),
                    "hk": ", ".join(e_hk), "ts": ts, "id": int(rid)
                })
            st.success("Kayıt güncellendi.")
            st.rerun()

        if col_btn2.button("Sil"):
            with engine.begin() as conn:
                conn.execute(text("DELETE FROM pdi_kayitlari WHERE id=:id"), {"id": int(rid)})
            st.success("Kayıt silindi.")
            st.rerun()
