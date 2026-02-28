import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import pandas as pd
from datetime import datetime
import os
import base64
import calendar
from collections import Counter
from PIL import Image as PilImage, ImageTk

# ------------------------------------------------------
# SABİTLER & AYARLAR
# ------------------------------------------------------
ALT_GRUP = ["Boya", "Süsleme", "Mekanik", "Elektrik"]
HATA_KONUM = [
    "Aydınlatma","Ayırma Duvarı","Ayna","Boya","CAM","Çıta","Defroster","Etiket",
    "Kapak","Kapı","Kaplama","Kelepçe","Klima","Koltuk","Körük","Lamba","Montaj",
    "Motor","Mutfak","Paket Raf","Silecek","Siperlik","Stepne","Şoför yatma yeri",
    "Tapa","Telefon","Torpido","Yazı"
]
ARAC_TIPI = ["Tourismo", "Connecto", "Travego"]
DB_NAME = "pdi_veritabani.db" 

# RENK PALETİ
COLOR_SIDEBAR = "#111111"      
COLOR_SIDEBAR_ACT = "#1f1f1f" 
COLOR_BG = "#F0F2F5"           
COLOR_WHITE = "#FFFFFF"        
COLOR_TEXT_MAIN = "#000000"    
COLOR_ACCENT = "#000000"       
COLOR_GREEN = "#28a745"        
COLOR_RED = "#dc3545"          
COLOR_BLUE = "#007bff"
COLOR_CHART_LINE = "#0056b3"

FONT_HEADER = ("Segoe UI", 14, "bold")
FONT_NORMAL = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_CARD_NUM = ("Segoe UI", 32, "bold")

LOGO_DATA = "" 

TURKISH_MONTHS = {
    1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
    7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
}
TURKISH_DAYS = {
    0: "PAZARTESİ", 1: "SALI", 2: "ÇARŞAMBA", 3: "PERŞEMBE", 4: "CUMA", 5: "CUMARTESİ", 6: "PAZAR"
}

# ------------------------------------------------------
# GELİŞMİŞ TAKVİM SINIFI (HIZLI AY SEÇİMİ EKLENDİ)
# ------------------------------------------------------
class AdvancedCalendarDialog(tk.Toplevel):
    def __init__(self, parent, current_date_str=None):
        super().__init__(parent)
        self.title("Tarih Seç")
        self.geometry("320x320")
        self.configure(bg=COLOR_WHITE)
        self.resizable(False, False)
        self.result = None
        
        # Başlangıç Tarihi
        self.now = datetime.now()
        self.year = self.now.year
        self.month = self.now.month
        
        if current_date_str:
            try:
                d = datetime.strptime(current_date_str, "%d-%m-%Y")
                self.year = d.year
                self.month = d.month
            except: pass

        # UI Ana Çerçeve
        self.container = tk.Frame(self, bg=COLOR_WHITE, padx=10, pady=10)
        self.container.pack(fill="both", expand=True)
        
        # Üst Kontroller
        self.header_frame = tk.Frame(self.container, bg=COLOR_WHITE)
        self.header_frame.pack(fill="x", pady=(0, 10))
        
        self.btn_prev = tk.Button(self.header_frame, text="<", command=self.prev_month, relief="flat", bg="#f0f0f0", width=3)
        self.btn_prev.pack(side="left")
        
        self.btn_month_year = tk.Button(self.header_frame, text="", command=self.toggle_view, relief="flat", bg=COLOR_WHITE, font=("Segoe UI", 11, "bold"), cursor="hand2")
        self.btn_month_year.pack(side="left", fill="x", expand=True)
        
        self.btn_next = tk.Button(self.header_frame, text=">", command=self.next_month, relief="flat", bg="#f0f0f0", width=3)
        self.btn_next.pack(side="left")

        # İçerik Alanı (Günler veya Aylar)
        self.content_frame = tk.Frame(self.container, bg=COLOR_WHITE)
        self.content_frame.pack(fill="both", expand=True)
        
        self.view_mode = "days" # veya "months"
        self.draw_days_view()
        
        # Modal Ayarlar
        self.transient(parent)
        self.grab_set()
        self.wait_window()

    def toggle_view(self):
        if self.view_mode == "days":
            self.view_mode = "months"
            self.draw_months_view()
        else:
            self.view_mode = "days"
            self.draw_days_view()

    def draw_days_view(self):
        # Temizlik
        for widget in self.content_frame.winfo_children(): widget.destroy()
        self.btn_prev.config(state="normal")
        self.btn_next.config(state="normal")
        
        # Başlık Güncelle
        tr_month = list(TURKISH_MONTHS.values())[self.month-1]
        self.btn_month_year.config(text=f"{tr_month} {self.year}")
        
        # Gün İsimleri
        days = ["Pt", "Sa", "Ça", "Pe", "Cu", "Ct", "Pa"]
        for i, d in enumerate(days):
            tk.Label(self.content_frame, text=d, font=("Segoe UI", 9, "bold"), bg=COLOR_WHITE, fg="#888").grid(row=0, column=i, sticky="nsew")

        # Takvim Günleri
        cal = calendar.monthcalendar(self.year, self.month)
        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day != 0:
                    btn = tk.Button(self.content_frame, text=str(day), 
                                    command=lambda d=day: self.select_date(d),
                                    relief="flat", bg=COLOR_WHITE, activebackground=COLOR_BLUE, activeforeground="white")
                    btn.grid(row=r+1, column=c, sticky="nsew", padx=1, pady=1)
                    
                    if day == self.now.day and self.month == self.now.month and self.year == self.now.year:
                        btn.config(fg=COLOR_BLUE, font=("Segoe UI", 9, "bold"))

        for i in range(7): self.content_frame.columnconfigure(i, weight=1)
        for i in range(6): self.content_frame.rowconfigure(i+1, weight=1)

    def draw_months_view(self):
        # Temizlik
        for widget in self.content_frame.winfo_children(): widget.destroy()
        
        # Yıl Seçici Modu
        self.btn_month_year.config(text=f"{self.year}")
        self.btn_prev.config(command=lambda: self.change_year(-1))
        self.btn_next.config(command=lambda: self.change_year(1))
        
        # Aylar Grid
        months = list(TURKISH_MONTHS.values())
        for i, m_name in enumerate(months):
            r, c = divmod(i, 3) # 4 satır, 3 sütun
            btn = tk.Button(self.content_frame, text=m_name, 
                            command=lambda m=i+1: self.select_month_from_grid(m),
                            relief="flat", bg="#f9f9f9", activebackground="#e0e0e0")
            btn.grid(row=r, column=c, sticky="nsew", padx=2, pady=2)
            if (i+1) == self.month:
                btn.config(bg=COLOR_BLUE, fg="white")

        for i in range(3): self.content_frame.columnconfigure(i, weight=1)
        for i in range(4): self.content_frame.rowconfigure(i, weight=1)

    def change_year(self, delta):
        self.year += delta
        self.btn_month_year.config(text=f"{self.year}")
        # Görünüm yenilemeye gerek yok, sadece label değişti ama aylar aynı

    def select_month_from_grid(self, month_idx):
        self.month = month_idx
        self.view_mode = "days"
        # Buton komutlarını eski haline getir
        self.btn_prev.config(command=self.prev_month)
        self.btn_next.config(command=self.next_month)
        self.draw_days_view()

    def prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month = 12
            self.year -= 1
        self.draw_days_view()

    def next_month(self):
        self.month += 1
        if self.month == 13:
            self.month = 1
            self.year += 1
        self.draw_days_view()

    def select_date(self, day):
        self.result = f"{day:02d}-{self.month:02d}-{self.year}"
        self.destroy()

# ------------------------------------------------------
# VERİTABANI İŞLEMLERİ
# ------------------------------------------------------
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users(username TEXT PRIMARY KEY, password TEXT, role INT, aciklama TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS pdi_kayitlari(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bb_no TEXT, sasi_no TEXT, arac_tipi TEXT, is_emri_no TEXT, alt_grup TEXT,
        tespitler TEXT, hata_konumu TEXT, fotograf_yolu TEXT, tarih_saat TEXT,
        kullanici TEXT, duzenleyen TEXT, grup_no TEXT, parca_tanimi TEXT,
        hata_tanimi TEXT, musteri_sikayeti TEXT, musteri_beklentisi TEXT,
        musteri_geri_bildirimi TEXT, musteri_teyidi TEXT, musteri_talebi TEXT,
        musteri_adi TEXT, musteri_soyadi TEXT, musteri_iletisim TEXT, musteri_adresi TEXT
    )""")
    # Top Hatalar tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS top_hatalar(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hata_adi TEXT UNIQUE,
        aktif INTEGER DEFAULT 1,
        olusturma_tarihi TEXT
    )""")
    cols = ["grup_no", "parca_tanimi", "hata_tanimi", "musteri_sikayeti", "musteri_beklentisi", "musteri_geri_bildirimi", "musteri_teyidi", "musteri_talebi", "musteri_adi", "musteri_soyadi", "musteri_iletisim", "musteri_adresi", "duzenleyen", "top_hata"]
    for col in cols:
        try: c.execute(f"ALTER TABLE pdi_kayitlari ADD COLUMN {col} TEXT")
        except: pass
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        c.execute("INSERT INTO users VALUES ('admin', 'admin123', 1, 'Sistem Yöneticisi')")
    conn.commit(); conn.close()

# ------------------------------------------------------
# UYGULAMA SINIFI
# ------------------------------------------------------
class PDIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mercedes-Benz Türk | Otobüs Teknik PDI")
        self.root.geometry("1366x768")
        self.root.configure(bg=COLOR_BG)
        self.content_area = None
        
        self.current_user = None; self.current_role = None
        self.configure_styles()
        self.show_login_screen()

    def configure_styles(self):
        style = ttk.Style(); style.theme_use("clam")
        style.configure("Treeview", background=COLOR_WHITE, fieldbackground=COLOR_WHITE, rowheight=35, font=FONT_NORMAL, foreground=COLOR_TEXT_MAIN, borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#E9ECEF", foreground="#495057", relief="flat")
        style.map("Treeview", background=[('selected', '#e8f0fe')], foreground=[('selected', 'black')])
        style.configure("Card.TFrame", background=COLOR_WHITE, relief="flat")

    def get_logo_image(self, size=(40, 40)):
        if not LOGO_DATA: return None
        try:
            img_data = base64.b64decode(LOGO_DATA)
            image = tk.PhotoImage(data=img_data)
            if size[0] < 60: image = image.subsample(2, 2) 
            return image
        except: return None

    def show_login_screen(self):
        self.clear_window(); self.root.configure(bg="#000000")
        card_frame = tk.Frame(self.root, bg=COLOR_WHITE, padx=50, pady=50)
        card_frame.place(relx=0.5, rely=0.5, anchor="center", width=450)
        self.photo_image = self.get_logo_image((80, 80))
        if self.photo_image: tk.Label(card_frame, image=self.photo_image, bg=COLOR_WHITE).pack(pady=(0, 20))
        tk.Label(card_frame, text="Mercedes-Benz", font=("Times New Roman", 24, "bold"), bg=COLOR_WHITE, fg="black").pack()
        tk.Label(card_frame, text="BUS PDI SYSTEM", font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg="#666").pack(pady=(0,30))
        tk.Label(card_frame, text="Kullanıcı Adı", bg=COLOR_WHITE, font=FONT_BOLD, fg="#333").pack(anchor="w")
        self.entry_user = ttk.Entry(card_frame, font=("Segoe UI", 12)); self.entry_user.pack(fill="x", pady=(5, 15), ipady=3)
        tk.Label(card_frame, text="Şifre", bg=COLOR_WHITE, font=FONT_BOLD, fg="#333").pack(anchor="w")
        self.entry_pass = ttk.Entry(card_frame, show="*", font=("Segoe UI", 12)); self.entry_pass.pack(fill="x", pady=(5, 25), ipady=3)
        tk.Button(card_frame, text="GİRİŞ", command=self.login, bg="black", fg="white", font=("Segoe UI", 12, "bold"), relief="flat", height=2).pack(fill="x")
        tk.Label(card_frame, text="Admin: admin / admin123", bg=COLOR_WHITE, fg="#ccc", font=("Segoe UI", 8)).pack(pady=(20,0))

    def login(self):
        u = self.entry_user.get(); p = self.entry_pass.get()
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        c.execute("SELECT role FROM users WHERE username=? AND password=?", (u, p))
        row = c.fetchone(); conn.close()
        if row: self.current_user = u; self.current_role = "admin" if row[0] == 1 else "user"; self.show_main_layout()
        else: messagebox.showerror("Hata", "Giriş başarısız.")

    def logout(self):
        self.current_user = None
        self.current_role = None
        self.show_login_screen()

    def clear_window(self):
        for widget in self.root.winfo_children(): widget.destroy()

    def clear_content(self):
        if self.content_area is not None:
            for w in self.content_area.winfo_children(): w.destroy()

    # ------------------------------------------------------
    # ANA DÜZEN
    # ------------------------------------------------------
    def show_main_layout(self):
        self.clear_window(); self.root.configure(bg=COLOR_BG)
        
        # --- SIDEBAR (SOL) ---
        self.sidebar = tk.Frame(self.root, bg=COLOR_SIDEBAR, width=250)
        self.sidebar.pack(side="left", fill="y"); self.sidebar.pack_propagate(False)
        
        # Logo
        logo_box = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR, height=80); logo_box.pack(fill="x", pady=20, padx=20)
        try:
             # Logo dosyasını kontrol et
             logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Mercedes_Benz_Logo.png")
             if os.path.exists(logo_path):
                 pil_img = Image.open(logo_path).resize((180, 60), Image.Resampling.LANCZOS)
                 self.logo_img = ImageTk.PhotoImage(pil_img)
                 tk.Label(logo_box, image=self.logo_img, bg=COLOR_SIDEBAR).pack(pady=0)
             else:
                 tk.Label(logo_box, text="Mercedes-Benz", fg="white", bg=COLOR_SIDEBAR, font=("Times New Roman", 14, "bold")).pack(anchor="w")
                 tk.Label(logo_box, text="BUS PDI SYSTEM", fg="#888", bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
        except:
             tk.Label(logo_box, text="Mercedes-Benz", fg="white", bg=COLOR_SIDEBAR, font=("Times New Roman", 14, "bold")).pack(anchor="w")
             tk.Label(logo_box, text="BUS PDI SYSTEM", fg="#888", bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
             
        tk.Frame(self.sidebar, bg="#333", height=1).pack(fill="x")

        # Kullanıcı Bilgisi
        user_box = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR, pady=20, padx=20); user_box.pack(fill="x")
        canvas_avatar = tk.Canvas(user_box, width=50, height=50, bg=COLOR_SIDEBAR, highlightthickness=0); canvas_avatar.pack(side="left")
        canvas_avatar.create_oval(2,2,48,48, fill="#f0f0f0", outline=""); canvas_avatar.create_text(25,25, text=self.current_user[0].upper(), font=("Segoe UI", 16, "bold"), fill="#333")
        u_info = tk.Frame(user_box, bg=COLOR_SIDEBAR); u_info.pack(side="left", padx=10)
        tk.Label(u_info, text=self.current_user.capitalize(), fg="white", bg=COLOR_SIDEBAR, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        role_txt = "YÖNETİCİ" if self.current_role == "admin" else "KULLANICI"; role_color = COLOR_GREEN if self.current_role == "admin" else "#17a2b8"
        tk.Label(u_info, text=role_txt, fg=role_color, bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
        tk.Frame(self.sidebar, bg="#333", height=1).pack(fill="x")
        
        # --- MENU BUTTONS ---
        self.create_sidebar_btn("⊞  ÖZET PANEL", self.show_dashboard)
        
        tk.Label(self.sidebar, text="RAPORLAR", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
        self.create_sidebar_btn("📊  TRV & TOU RAPOR", self.show_trv_tou_report)
        self.create_sidebar_btn("📊  CONNECTO RAPOR", self.show_connecto_report)
        self.create_sidebar_btn("📉  TOP 5 ANALİZ", self.show_top5_analysis)
        self.create_sidebar_btn("📉  CONN. TOP 3", self.show_connecto_top3)
        
        tk.Label(self.sidebar, text="KAYIT İŞLEMLERİ", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
        self.create_sidebar_btn("☰  KAYIT LİSTESİ", self.show_list_view)
        self.create_sidebar_btn("⊕  YENİ KAYIT", self.open_add_window)
        
        if self.current_role == "admin": 
            tk.Label(self.sidebar, text="YÖNETİM", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
            self.create_sidebar_btn("👤  KULLANICILAR", self.open_user_management)
            self.create_sidebar_btn("🏷️  TOP HATA YÖNETİMİ", self.open_top_hata_management)
            self.create_sidebar_btn("📥  GEÇMİŞ VERİ YÜKLE", self.open_excel_import)
            
        tk.Frame(self.sidebar, bg=COLOR_SIDEBAR).pack(fill="both", expand=True)
        tk.Button(self.sidebar, text="↪ OTURUMU KAPAT", command=self.logout, bg="#1a1a1a", fg="#aaa", relief="flat", font=("Segoe UI", 9, "bold"), anchor="w", padx=20, pady=15, bd=0).pack(fill="x", side="bottom")

        # --- SAĞ PANEL (HEADER + CONTENT) ---
        right_panel = tk.Frame(self.root, bg=COLOR_BG); right_panel.pack(side="right", fill="both", expand=True)
        
        # Header
        self.header = tk.Frame(right_panel, bg=COLOR_WHITE, height=60, padx=30); self.header.pack(fill="x"); self.header.pack_propagate(False)
        tk.Label(self.header, text="MERKEZ ATÖLYE", font=("Segoe UI", 10, "bold"), fg="#999", bg=COLOR_WHITE).pack(side="left")
        self.lbl_date = tk.Label(self.header, text="", font=("Segoe UI", 10, "bold"), fg="#333", bg=COLOR_WHITE); self.lbl_date.pack(side="right")
        tk.Label(self.header, text="MBT Genel Merkez\nISTANBUL, TR", font=("Segoe UI", 8), fg="#999", bg=COLOR_WHITE, justify="right").pack(side="right", padx=20)
        
        self.update_clock()
        
        # Content Area
        self.content_area = tk.Frame(right_panel, bg=COLOR_BG)
        self.content_area.pack(fill="both", expand=True)

        # Varsayılan Sayfa
        self.show_dashboard()

    def create_sidebar_btn(self, text, command):
        btn = tk.Button(self.sidebar, text=text, command=command, bg=COLOR_SIDEBAR, fg="#ccc", font=("Segoe UI", 10, "bold"), relief="flat", activebackground=COLOR_SIDEBAR_ACT, activeforeground="white", anchor="w", padx=20, pady=12, bd=0)
        btn.pack(fill="x", pady=1); return btn

    def update_clock(self):
        now = datetime.now(); day_name = TURKISH_DAYS[now.weekday()]; month_name = TURKISH_MONTHS[now.month]
        date_str = f"{now.day} {month_name} {now.year} {day_name}"
        if self.lbl_date.winfo_exists(): self.lbl_date.config(text=date_str); self.root.after(60000, self.update_clock)

    # ------------------------------------------------------
    # HOME SUMMARY (YENİ ÖZET EKRANI)
    # ------------------------------------------------------
    def show_dashboard(self):
        self.clear_content()
        
        # Header Area
        header = tk.Frame(self.content_area, bg=COLOR_BG)
        header.pack(fill="x", padx=30, pady=30)
        
        tk.Label(header, text="GENEL ÖZET", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(side="left")
        
        # DB bağlantısı
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        
        # Toplam sayılar
        total = c.execute("SELECT count(*) FROM pdi_kayitlari").fetchone()[0]
        c_tou = c.execute("SELECT count(*) FROM pdi_kayitlari WHERE arac_tipi='Tourismo'").fetchone()[0]
        c_trv = c.execute("SELECT count(*) FROM pdi_kayitlari WHERE arac_tipi='Travego'").fetchone()[0]
        c_con = c.execute("SELECT count(*) FROM pdi_kayitlari WHERE arac_tipi='Connecto'").fetchone()[0]
        conn.close()
        
        # Kartlar
        cards_frame = tk.Frame(self.content_area, bg=COLOR_BG)
        cards_frame.pack(fill="x", padx=30)
        
        self.create_info_card(cards_frame, "TOPLAM ARAÇ", str(total), "🚌", 0)
        self.create_info_card(cards_frame, "TOURISMO", str(c_tou), "", 1)
        self.create_info_card(cards_frame, "TRAVEGO", str(c_trv), "", 2)
        self.create_info_card(cards_frame, "CONNECTO", str(c_con), "", 3)
        
        # Alt kısma belki hoş geldiniz mesajı veya logo
        welcome = tk.Label(self.content_area, text="Detaylı raporlar için soldaki menüden seçim yapınız.", font=("Segoe UI", 12), bg=COLOR_BG, fg="#666")
        welcome.pack(pady=50)

    # ------------------------------------------------------
    # RAPOR SAYFALARI (FULL SCREEN)
    # ------------------------------------------------------
    
    # Yardımcı: Ay seçici başlığı
    def create_report_header(self, title, command_callback):
        header = tk.Frame(self.content_area, bg=COLOR_BG)
        header.pack(fill="x", padx=20, pady=20)
        
        tk.Label(header, text=title, font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(side="left")
        
        ctrl_frame = tk.Frame(header, bg=COLOR_BG)
        ctrl_frame.pack(side="right")
        
        tk.Label(ctrl_frame, text="Rapor Ayı: ", bg=COLOR_BG, font=("Segoe UI", 10, "bold")).pack(side="left")
        
        # Son 12 ay + gelecek ay seçenekleri
        months = []
        now = datetime.now()
        for i in range(13):
            d = now - timedelta(days=30 * (12-i)) # yaklaşık
            # daha düzgün tarih listesi
            m_idx = (now.month - (12-i) - 1) % 12 + 1
            y_offset = (now.month - (12-i) - 1) // 12
            y = now.year + y_offset
            months.append(f"{TURKISH_MONTHS[m_idx]} {y}")
            
        # Gelecek ay da olsun (test için)
        next_m = (now.month % 12) + 1
        next_y = now.year + (1 if now.month == 12 else 0)
        months.append(f"{TURKISH_MONTHS[next_m]} {next_y}")
        
        # Tersten sırala (en yeni en üstte)
        months.reverse()
        
        self.selected_month = tk.StringVar(value=f"{TURKISH_MONTHS[now.month]} {now.year}")
        cb = ttk.Combobox(ctrl_frame, textvariable=self.selected_month, values=months, state="readonly", width=15)
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda e: command_callback())
        
        return header

    def show_trv_tou_report(self):
        try:
            self.clear_content()
            self.create_report_header("TRV & TOU RAPORU", self.show_trv_tou_report)
            
            month, year = self.get_selected_month_year()
            # Veri hesaplama (Hata kaynağı olabilir)
            data = self.calculate_monthly_data(["Tourismo", "Travego"], month, year)
            
            container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
            container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
            
            # Template metin
            template_text = f"{year} {TURKISH_MONTHS[month]} ayında PDI yapılan [TRV & TOU] araç sayısı {data['current_month_vehicles']} adettir. Son 12 ay ortalama hata oranı araç başı {data['avg_12_month']:.2f} adet olup, {TURKISH_MONTHS[month]} ayı özelinde {data['current_month_rate']:.2f} adet olarak gerçekleşmiştir."
            tk.Label(container, text=template_text, font=("Segoe UI", 11), bg=COLOR_WHITE, fg="#333", wraplength=1100, justify="left").pack(anchor="w", pady=(0, 20))
            
            # İçerik Düzeni (GRID Layout)
            content_grid = tk.Frame(container, bg=COLOR_WHITE)
            content_grid.pack(fill="both", expand=True)
            content_grid.columnconfigure(0, weight=3) # Bar grafik geniş
            content_grid.columnconfigure(1, weight=1) # Donut dar
            content_grid.rowconfigure(0, weight=1)
            
            # SOL: Bar Grafik
            left_frame = tk.Frame(content_grid, bg=COLOR_WHITE, highlightthickness=1, highlightbackground="#eee")
            left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
            
            bar_canvas = tk.Canvas(left_frame, bg=COLOR_WHITE, highlightthickness=0)
            h_scroll = ttk.Scrollbar(left_frame, orient="horizontal", command=bar_canvas.xview)
            bar_canvas.configure(xscrollcommand=h_scroll.set)
            bar_canvas.pack(fill="both", expand=True, pady=10)
            h_scroll.pack(fill="x")
            
            self.draw_combo_bar_line_chart(bar_canvas, data['monthly_stats'])
            
            # SAĞ: Donut Grafikleri
            right_frame = tk.Frame(content_grid, bg=COLOR_WHITE)
            right_frame.grid(row=0, column=1, sticky="nesw")
            
            # Dikey yerleşim
            self.draw_donut_section_vertical(right_frame, data)
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor yüklenemedi:\n{str(e)}")
            print(f"Rapor Hatası: {e}")

    def show_connecto_report(self):
        try:
            self.clear_content()
            self.create_report_header("CONNECTO RAPORU", self.show_connecto_report)
            
            month, year = self.get_selected_month_year()
            data = self.calculate_monthly_data(["Connecto"], month, year)
            
            container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
            container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
            
            template_text = f"{year} {TURKISH_MONTHS[month]} ayında PDI yapılan [CONNECTO] araç sayısı {data['current_month_vehicles']} adettir. Son 12 ay ortalama hata oranı araç başı {data['avg_12_month']:.2f} adet olup, {TURKISH_MONTHS[month]} ayı özelinde {data['current_month_rate']:.2f} adet olarak gerçekleşmiştir."
            tk.Label(container, text=template_text, font=("Segoe UI", 11), bg=COLOR_WHITE, fg="#333", wraplength=1100, justify="left").pack(anchor="w", pady=(0, 20))
            
            content_grid = tk.Frame(container, bg=COLOR_WHITE)
            content_grid.pack(fill="both", expand=True)
            content_grid.columnconfigure(0, weight=3)
            content_grid.columnconfigure(1, weight=1)
            content_grid.rowconfigure(0, weight=1)
            
            left_frame = tk.Frame(content_grid, bg=COLOR_WHITE, highlightthickness=1, highlightbackground="#eee")
            left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
            
            bar_canvas = tk.Canvas(left_frame, bg=COLOR_WHITE, highlightthickness=0)
            h_scroll = ttk.Scrollbar(left_frame, orient="horizontal", command=bar_canvas.xview)
            bar_canvas.configure(xscrollcommand=h_scroll.set)
            bar_canvas.pack(fill="both", expand=True, pady=10)
            h_scroll.pack(fill="x")
            
            self.draw_combo_bar_line_chart(bar_canvas, data['monthly_stats'])
            
            right_frame = tk.Frame(content_grid, bg=COLOR_WHITE)
            right_frame.grid(row=0, column=1, sticky="nesw")
            
            self.draw_donut_section_vertical(right_frame, data)
        except Exception as e:
            messagebox.showerror("Hata", f"Connecto Rapor Hatası:\n{str(e)}")
            print(f"Connecto Hata: {e}")

    def show_top5_analysis(self):
        self.clear_content()
        self.create_report_header("TOP 5 HATA ANALİZİ", self.show_top5_analysis)
        
        month, year = self.get_selected_month_year()
        top5_data = self.get_top5_hata_analysis(["Tourismo", "Travego"], month, year)
        
        container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        tk.Label(container, text="Aşağıdaki listeden detay görmek istediğiniz hataya tıklayınız.", font=("Segoe UI", 10, "italic"), bg=COLOR_WHITE, fg="#777").pack(anchor="w", pady=(0, 10))
        
        # Tablo Başlıkları
        hdr_frame = tk.Frame(container, bg="#eee", height=35)
        hdr_frame.pack(fill="x")
        cols = [("TOP HATA TANIMI", 0.4), ("TRAVEGO %", 0.15), ("TOURISMO %", 0.15), ("GENEL ORTALAMA %", 0.15), ("GÜNCEL DURUM", 0.15)]
        for txt, relw in cols:
            tk.Label(hdr_frame, text=txt, font=("Segoe UI", 9, "bold"), bg="#eee", fg="#333").pack(side="left", fill="both", expand=True) # expand=True aslında oranları bozar, frame ile yapmak lazım ama basit tutalım
        
        # Basit layout için pack side=left yerine grid kullanalım içerde
        for w in hdr_frame.winfo_children(): w.destroy()
        
        grid_frame = tk.Frame(container, bg=COLOR_WHITE)
        grid_frame.pack(fill="both", expand=True)
        
        # Headers
        for i, (txt, _) in enumerate(cols):
            tk.Label(grid_frame, text=txt, font=("Segoe UI", 9, "bold"), bg="#eee", fg="#333", pady=8, borderwidth=1, relief="solid").grid(row=0, column=i, sticky="nsew")
        
        grid_frame.columnconfigure(0, weight=4)
        for i in range(1, 5): grid_frame.columnconfigure(i, weight=2)
        
        if not top5_data:
            tk.Label(grid_frame, text="Veri bulunamadı.", bg=COLOR_WHITE).grid(row=1, column=0, columnspan=5, pady=20)
            return

        for idx, item in enumerate(top5_data):
            row_idx = idx + 1
            bg_col = "#fff" if idx % 2 == 0 else "#f9f9f9"
            
            # Satır container (tıklanabilir olması için label kullanacağız)
            # Ama grid içinde tek tek label koymak lazım
            
            def create_cell(text, col_idx, fg="black", font=FONT_NORMAL):
                lbl = tk.Label(grid_frame, text=text, bg=bg_col, fg=fg, font=font, pady=8, borderwidth=0, relief="flat", cursor="hand2")
                lbl.grid(row=row_idx, column=col_idx, sticky="nsew", padx=1, pady=1)
                lbl.bind("<Button-1>", lambda e, h=item['hata_adi']: self.show_top_hata_detail(h))
                return lbl
                
            create_cell(item['hata_adi'], 0, font=("Segoe UI", 10, "bold"))
            create_cell(f"%{item['trv_rate']:.2f}", 1)
            create_cell(f"%{item['tou_rate']:.2f}", 2)
            create_cell(f"%{item['avg_rate']:.2f}", 3)
            
            # Durum ikonu/rengi
            status_color = "black"
            status_text = "Stabil"
            diff = item['diff']
            if diff > 10: status_color, status_text = "#d32f2f", "▲ Kritik Artış"
            elif diff > 4: status_color, status_text = "#f57c00", "▲ Artış"
            elif diff < 0: status_color, status_text = "#388e3c", "▼ Düşüş"
            
            create_cell(status_text, 4, fg=status_color, font=("Segoe UI", 9, "bold"))

    def show_connecto_top3(self):
        self.clear_content()
        self.create_report_header("CONNECTO TOP 3 ANALİZİ", self.show_connecto_top3)
        
        month, year = self.get_selected_month_year()
        data = self.get_connecto_top3(month, year)
        
        container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        title = f"{year} yılı {TURKISH_MONTHS[month]} ayı için Connecto Top 3 Hataları"
        tk.Label(container, text=title, font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE).pack(pady=(0, 20))
        
        # 3x3 Tablo (Header + 3 satır)
        table_frame = tk.Frame(container, bg=COLOR_WHITE, highlightbackground="#ccc", highlightthickness=1)
        table_frame.pack(fill="x", padx=50)
        
        headers = ["TOP HATA", "YTD Toplam Adet", "YTD Toplam Oran"]
        for i, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=FONT_BOLD, bg="#eee", pady=10).grid(row=0, column=i, sticky="nsew")
        
        table_frame.columnconfigure(0, weight=3)
        table_frame.columnconfigure(1, weight=1)
        table_frame.columnconfigure(2, weight=1)
        
        if not data:
             tk.Label(table_frame, text="Veri Yok", bg=COLOR_WHITE).grid(row=1, column=0, columnspan=3, pady=20)
        else:
            for idx, item in enumerate(data):
                r = idx + 1
                tk.Label(table_frame, text=item['hata'], font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=0, sticky="nsew")
                tk.Label(table_frame, text=str(item['count']), font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=1, sticky="nsew")
                tk.Label(table_frame, text=f"%{item['rate']:.2f}", font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=2, sticky="nsew")
    
    def get_selected_month_year(self):
        try:
            sel = self.selected_month.get()
            parts = sel.split(" ")
            month_name = parts[0]
            year = int(parts[1])
            month = [k for k, v in TURKISH_MONTHS.items() if v == month_name][0]
            return month, year
        except:
            now = datetime.now()
            return now.month, now.year
    
    # HELPER: Aylık veri hesaplama
    def calculate_monthly_data(self, arac_tipleri, month, year):
        conn = sqlite3.connect(DB_NAME)
        placeholders = ",".join(["?" for _ in arac_tipleri])
        
        # Son 12 ay için veri al
        monthly_stats = []
        for i in range(12):
            m = month - i
            y = year
            while m <= 0: m += 12; y -= 1
            
            month_start = f"01-{m:02d}-{y}"
            month_end = f"{calendar.monthrange(y, m)[1]:02d}-{m:02d}-{y}"
            
            query = f"""SELECT COUNT(DISTINCT sasi_no) as arac, 
                        COUNT(*) as hata_sayisi
                        FROM pdi_kayitlari 
                        WHERE arac_tipi IN ({placeholders})
                        AND substr(tarih_saat,4,2) = ? AND substr(tarih_saat,7,4) = ?"""
            
            result = conn.execute(query, arac_tipleri + [f"{m:02d}", str(y)]).fetchone()
            arac_sayisi = result[0] if result[0] else 0
            hata_sayisi = result[1] if result[1] else 0
            hata_orani = hata_sayisi / arac_sayisi if arac_sayisi > 0 else 0
            
            monthly_stats.insert(0, {
                'ay': TURKISH_MONTHS[m][:3],
                'ay_full': TURKISH_MONTHS[m],
                'yil': y,
                'arac_sayisi': arac_sayisi,
                'hata_sayisi': hata_sayisi,
                'hata_orani': hata_orani
            })
        
        # Hesaplamalar
        current = monthly_stats[-1] if monthly_stats else {'arac_sayisi': 0, 'hata_sayisi': 0, 'hata_orani': 0}
        total_vehicles = sum(m['arac_sayisi'] for m in monthly_stats)
        total_errors = sum(m['hata_sayisi'] for m in monthly_stats)
        avg_12 = total_errors / total_vehicles if total_vehicles > 0 else 0
        
        # Geçen yıl ortalaması
        last_year_query = f"""SELECT COUNT(DISTINCT sasi_no), COUNT(*) FROM pdi_kayitlari 
                             WHERE arac_tipi IN ({placeholders}) AND substr(tarih_saat,7,4) = ?"""
        ly_result = conn.execute(last_year_query, arac_tipleri + [str(year-1)]).fetchone()
        ly_vehicles = ly_result[0] if ly_result[0] else 0
        ly_errors = ly_result[1] if ly_result[1] else 0
        avg_last_year = ly_errors / ly_vehicles if ly_vehicles > 0 else 0
        
        # Mevcut yıl
        cy_query = f"""SELECT COUNT(DISTINCT sasi_no), COUNT(*) FROM pdi_kayitlari 
                       WHERE arac_tipi IN ({placeholders}) AND substr(tarih_saat,7,4) = ?"""
        cy_result = conn.execute(cy_query, arac_tipleri + [str(year)]).fetchone()
        cy_vehicles = cy_result[0] if cy_result[0] else 0
        cy_errors = cy_result[1] if cy_result[1] else 0
        avg_current_year = cy_errors / cy_vehicles if cy_vehicles > 0 else 0
        
        conn.close()
        
        return {
            'monthly_stats': monthly_stats,
            'current_month_vehicles': current['arac_sayisi'],
            'current_month_rate': current['hata_orani'],
            'avg_12_month': avg_12,
            'avg_last_year': avg_last_year,
            'avg_current_year': avg_current_year,
            'last_year_vehicles': ly_vehicles,
            'last_year_errors': ly_errors,
            'current_year_vehicles': cy_vehicles,
            'current_year_errors': cy_errors,
            'total_12_vehicles': total_vehicles,
            'total_12_errors': total_errors
        }
    
    # HELPER: Top 5 Hata analizi
    def get_top5_hata_analysis(self, arac_tipleri, month, year):
        conn = sqlite3.connect(DB_NAME)
        top_hatalar = conn.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()
        
        results = []
        for (hata_adi,) in top_hatalar[:5]:
            # Bu ay TRV
            trv_query = """SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi='Travego' 
                          AND top_hata=? AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?"""
            trv_count = conn.execute(trv_query, (hata_adi, f"{month:02d}", str(year))).fetchone()[0]
            trv_total = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi='Travego' 
                                       AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?""", 
                                    (f"{month:02d}", str(year))).fetchone()[0]
            trv_oran = trv_count / trv_total if trv_total > 0 else 0
            
            # Bu ay TOU
            tou_query = """SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi='Tourismo' 
                          AND top_hata=? AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?"""
            tou_count = conn.execute(tou_query, (hata_adi, f"{month:02d}", str(year))).fetchone()[0]
            tou_total = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi='Tourismo' 
                                       AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?""", 
                                    (f"{month:02d}", str(year))).fetchone()[0]
            tou_oran = tou_count / tou_total if tou_total > 0 else 0
            
            # Önceki ay karşılaştırma
            prev_m = month - 1
            prev_y = year
            if prev_m <= 0: prev_m = 12; prev_y -= 1
            
            prev_trv = conn.execute(trv_query, (hata_adi, f"{prev_m:02d}", str(prev_y))).fetchone()[0]
            prev_trv_total = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi='Travego' 
                                            AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?""", 
                                         (f"{prev_m:02d}", str(prev_y))).fetchone()[0]
            prev_trv_oran = prev_trv / prev_trv_total if prev_trv_total > 0 else 0
            
            change = trv_oran - prev_trv_oran
            if change < 0:
                durum = "Azalış"
            elif change >= 0.04:
                durum = "Dikkat"
            else:
                durum = "İzle"
            
            results.append({
                'hata_adi': hata_adi,
                'trv_oran': trv_oran,
                'tou_oran': tou_oran,
                'durum': durum
            })
        
        conn.close()
        return results
    
    # HELPER: Connecto Top 3
    def get_connecto_top3(self, month, year):
        conn = sqlite3.connect(DB_NAME)
        
        # YTD toplam Connecto hata sayısı
        ytd_total = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari 
                                   WHERE arac_tipi='Connecto' AND substr(tarih_saat,7,4)=?""", 
                                (str(year),)).fetchone()[0]
        
        # Top 3 hata
        top_hatalar = conn.execute("""SELECT top_hata, COUNT(*) as cnt FROM pdi_kayitlari 
                                     WHERE arac_tipi='Connecto' AND substr(tarih_saat,7,4)=? 
                                     AND top_hata IS NOT NULL AND top_hata != ''
                                     GROUP BY top_hata ORDER BY cnt DESC LIMIT 3""", 
                                  (str(year),)).fetchall()
        
        results = []
        for hata, cnt in top_hatalar:
            oran = cnt / ytd_total if ytd_total > 0 else 0
            results.append({'hata_adi': hata, 'ytd_sayi': cnt, 'oran': oran})
        
        conn.close()
        return results
    
    # HELPER: Top Hata Detay Modalı
    def draw_donut_section_vertical(self, parent_frame, data):
        # 3 tane dikey donut (Geçen Yıl, Son 12 Ay, Mevcut Yıl)
        
        titles = [
            f"{data['prev_year']} Yılı AVG", 
            "Son 12 Ay AVG", 
            f"{data['current_year']} Yılı AVG"
        ]
        values = [data['prev_year_rate'], data['avg_12_month'], data['current_year_rate']]
        counts = [f"Araç: {data['prev_year_cnt']} | Hata: {data['prev_year_err']}",
                  f"Araç: {data['last_12_cnt']} | Hata: {data['last_12_err']}",
                  f"Araç: {data['curr_year_cnt']} | Hata: {data['curr_year_err']}"]
        
        # Grid ile yerleştir (tek sütun, 3 satır)
        parent_frame.columnconfigure(0, weight=1)
        parent_frame.rowconfigure(0, weight=1)
        parent_frame.rowconfigure(1, weight=1)
        parent_frame.rowconfigure(2, weight=1)
        
        for i in range(3):
            f = tk.Frame(parent_frame, bg=COLOR_WHITE)
            f.grid(row=i, column=0, sticky="nsew", pady=10)
            
            # Canvas
            canvas = tk.Canvas(f, bg=COLOR_WHITE, highlightthickness=0, height=180) # Yükseklik biraz daha büyük
            canvas.pack(fill="both", expand=True)
            
            # Çizim
            val = values[i]
            # Donut için açı hesabı yok aslında burada sadece 'rate' çiziyoruz.
            # Ama basit olması için sadece daire ve içinde yazı göstereceğiz.
            # Veya full circle olabilir, çünkü 'hata oranı' bir yüzde değil (0.5, 1.2 vs olabilir).
            # O yüzden "gauge" gibi bir şey de olabilir ama donut basit.
            
            w = 180
            h = 160 # Biraz daha kareye yakın
            x, y = w/2, h/2
            r = 60
            
            # Arka plan dairesi
            canvas.create_oval(x-r, y-r, x+r, y+r, outline="#ddd", width=15)
            
            # Değer var mı? Varsa kısmi daire.
            # Hata oranı genelde 0-2 arası. 2'yi %100 kabul edelim görsel için.
            extent = min((val / 2.5) * 360, 360) 
            # 2.5 hata oranı = tam tur. 
            
            color = "#5c8a8a" if i==0 else "#4682b4" if i==1 else "#1a5276"
            
            if val > 0:
                canvas.create_arc(x-r, y-r, x+r, y+r, start=90, extent=-extent, outline=color, width=15, style="arc")
            
            # İç yazı
            title_parts = titles[i].split(" AVG")
            canvas.create_text(x, y-25, text=title_parts[0], font=("Segoe UI", 9, "bold"), fill="#666")
            canvas.create_text(x, y-10, text="AVG", font=("Segoe UI", 8), fill="#999")
            canvas.create_text(x, y+10, text=f"{val:.2f}", font=("Segoe UI", 16, "bold"), fill="#333")
            
            canvas.create_text(x, y+90, text=counts[i], font=("Segoe UI", 8), fill="#777")
            
            # Canvas boyutunu güncelle (Event ile daha iyi olur ama statik şimdilik)
            # pack layout olduğu için ortalamak zor olabilir, width fixledik.
    def show_top_hata_detail(self, hata_adi, arac_tipleri):
        month, year = self.get_selected_month_year()
        
        win = tk.Toplevel(self.root)
        win.title(f"Detay: {hata_adi}")
        win.geometry("1000x700")
        win.configure(bg=COLOR_WHITE)
        
        tk.Label(win, text=f"{hata_adi} - Son 12 Ay Analizi", font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE).pack(pady=15)
        
        # Grafik
        canvas = tk.Canvas(win, bg=COLOR_WHITE, highlightthickness=0, height=250)
        canvas.pack(fill="x", padx=20)
        
        # Veri hesapla
        conn = sqlite3.connect(DB_NAME)
        monthly_data = {'TRV': [], 'TOU': []}
        
        for i in range(12):
            m = month - i
            y = year
            while m <= 0: m += 12; y -= 1
            
            for tip, label in [("Travego", "TRV"), ("Tourismo", "TOU")]:
                total = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi=? 
                                       AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?""",
                                    (tip, f"{m:02d}", str(y))).fetchone()[0]
                hata = conn.execute("""SELECT COUNT(*) FROM pdi_kayitlari WHERE arac_tipi=? 
                                      AND top_hata=? AND substr(tarih_saat,4,2)=? AND substr(tarih_saat,7,4)=?""",
                                   (tip, hata_adi, f"{m:02d}", str(y))).fetchone()[0]
                oran = hata / total if total > 0 else 0
                monthly_data[label].insert(0, {'ay': TURKISH_MONTHS[m][:3], 'arac': total, 'hata': hata, 'oran': oran})
        
        conn.close()
        
        # Çizgi grafik çiz
        self.draw_dual_line_chart(canvas, monthly_data)
        
        # Tablo
        table_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20)
        table_frame.pack(fill="both", expand=True, pady=20)
        
        self.draw_detail_table(table_frame, monthly_data)
    
    # GRAFİK: Bar + Line kombinasyonu
    def draw_combo_bar_line_chart(self, canvas, monthly_stats):
        if not monthly_stats:
            canvas.create_text(400, 140, text="Veri Bulunamadı", font=FONT_NORMAL, fill="#999")
            return
        
        h = 280; pad_x = 60; pad_y = 50
        num_months = len(monthly_stats)
        bar_width = 30; gap = 60
        chart_width = pad_x + num_months * (bar_width * 2 + gap) + pad_x
        canvas.config(scrollregion=(0, 0, chart_width, h))
        
        max_count = max(max(m['arac_sayisi'], m['hata_sayisi']) for m in monthly_stats)
        max_count = max_count if max_count > 0 else 1
        max_rate = max(m['hata_orani'] for m in monthly_stats)
        max_rate = max_rate if max_rate > 0 else 1
        
        graph_h = h - 2 * pad_y
        rate_points = []
        
        for i, m in enumerate(monthly_stats):
            x = pad_x + i * (bar_width * 2 + gap)
            
            # Araç sayısı bar (açık mavi)
            arac_h = (m['arac_sayisi'] / max_count) * graph_h
            canvas.create_rectangle(x, h - pad_y - arac_h, x + bar_width, h - pad_y, fill="#8FADC5", outline="")
            canvas.create_text(x + bar_width/2, h - pad_y - arac_h - 10, text=str(m['arac_sayisi']), font=("Segoe UI", 8), fill="#666")
            
            # Hata sayısı bar (koyu mavi)
            hata_h = (m['hata_sayisi'] / max_count) * graph_h
            canvas.create_rectangle(x + bar_width + 5, h - pad_y - hata_h, x + bar_width * 2 + 5, h - pad_y, fill="#3D5A73", outline="")
            canvas.create_text(x + bar_width * 1.5 + 5, h - pad_y - hata_h - 10, text=str(m['hata_sayisi']), font=("Segoe UI", 8, "bold"), fill="#333")
            
            # Ay etiketi
            canvas.create_text(x + bar_width, h - pad_y + 15, text=m['ay'], font=("Segoe UI", 9), fill="#666")
            
            # Rate çizgisi için nokta
            rate_y = h - pad_y - (m['hata_orani'] / max_rate) * graph_h * 0.8
            rate_points.append((x + bar_width, rate_y))
            canvas.create_oval(x + bar_width - 4, rate_y - 4, x + bar_width + 4, rate_y + 4, fill="#dc3545", outline="white", width=2)
            canvas.create_text(x + bar_width, rate_y - 15, text=f"{m['hata_orani']:.2f}", font=("Segoe UI", 9, "bold"), fill="#dc3545")
        
        # Rate çizgisi
        if len(rate_points) > 1:
            canvas.create_line(rate_points, fill="#dc3545", width=2)
        
        # Legend
        canvas.create_rectangle(chart_width - 200, 10, chart_width - 180, 25, fill="#8FADC5", outline="")
        canvas.create_text(chart_width - 170, 17, text="Araç Sayısı", font=("Segoe UI", 8), anchor="w")
        canvas.create_rectangle(chart_width - 200, 30, chart_width - 180, 45, fill="#3D5A73", outline="")
        canvas.create_text(chart_width - 170, 37, text="Hata Sayısı", font=("Segoe UI", 8), anchor="w")
        canvas.create_line(chart_width - 200, 55, chart_width - 180, 55, fill="#dc3545", width=2)
        canvas.create_text(chart_width - 170, 55, text="Hata Oranı", font=("Segoe UI", 8), anchor="w")
    
    # GRAFİK: Donut bölümü
    def draw_donut_section(self, parent, data):
        parent.pack(fill="x", pady=10)
        
        donuts = [
            (f"{datetime.now().year - 1} AVG", data['avg_last_year'], data['last_year_vehicles'], data['last_year_errors'], "#888888"),
            ("12 Ay AVG", data['avg_12_month'], data['total_12_vehicles'], data['total_12_errors'], "#4A90A4"),
            (f"{datetime.now().year} AVG", data['avg_current_year'], data['current_year_vehicles'], data['current_year_errors'], "#2E5A6B")
        ]
        
        for i, (label, rate, vehicles, errors, color) in enumerate(donuts):
            frame = tk.Frame(parent, bg=COLOR_WHITE, padx=20)
            frame.pack(side="left", expand=True)
            
            canvas = tk.Canvas(frame, width=150, height=150, bg=COLOR_WHITE, highlightthickness=0)
            canvas.pack()
            
            # Donut
            canvas.create_arc(10, 10, 140, 140, start=90, extent=-270, fill=color, outline="")
            canvas.create_arc(10, 10, 140, 140, start=90, extent=-90, fill="#e0e0e0", outline="")
            canvas.create_oval(40, 40, 110, 110, fill=COLOR_WHITE, outline="")
            
            # Merkez metin
            canvas.create_text(75, 60, text=label, font=("Segoe UI", 9, "bold"), fill="#666")
            canvas.create_text(75, 75, text="Hata Oranı:", font=("Segoe UI", 8), fill="#999")
            canvas.create_text(75, 95, text=f"{rate:.2f}", font=("Segoe UI", 16, "bold"), fill=color)
            
            # Alt etiketler
            tk.Label(frame, text=f"Araç: {vehicles} | Hata: {errors}", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#666").pack()
    
    # GRAFİK: Dual line chart
    def draw_dual_line_chart(self, canvas, monthly_data):
        h = 250; w = 960; pad_x = 60; pad_y = 40
        canvas.config(width=w, height=h)
        
        max_rate = 0
        for label in ['TRV', 'TOU']:
            if monthly_data[label]:
                max_rate = max(max_rate, max(m['oran'] for m in monthly_data[label]))
        max_rate = max_rate if max_rate > 0 else 1
        
        graph_h = h - 2 * pad_y
        num_months = len(monthly_data['TRV'])
        step_x = (w - 2 * pad_x) / max(num_months - 1, 1)
        
        colors = {'TRV': '#0066cc', 'TOU': '#ff8c00'}
        
        for label in ['TRV', 'TOU']:
            points = []
            for i, m in enumerate(monthly_data[label]):
                x = pad_x + i * step_x
                y = h - pad_y - (m['oran'] / max_rate) * graph_h
                points.append((x, y))
                canvas.create_oval(x-4, y-4, x+4, y+4, fill=colors[label], outline="white", width=2)
                canvas.create_text(x, y-15, text=f"{m['oran']:.0%}", font=("Segoe UI", 8), fill=colors[label])
                if label == 'TRV':
                    canvas.create_text(x, h - pad_y + 15, text=m['ay'], font=("Segoe UI", 8), fill="#666")
            
            if len(points) > 1:
                canvas.create_line(points, fill=colors[label], width=2)
        
        # Legend
        canvas.create_line(w - 150, 20, w - 130, 20, fill=colors['TRV'], width=2)
        canvas.create_text(w - 120, 20, text="TRV Hata Oranı", font=("Segoe UI", 8), anchor="w")
        canvas.create_line(w - 150, 35, w - 130, 35, fill=colors['TOU'], width=2)
        canvas.create_text(w - 120, 35, text="TOU Hata Oranı", font=("Segoe UI", 8), anchor="w")
    
    # TABLO: Detay tablosu
    def draw_detail_table(self, parent, monthly_data):
        # Header
        header_frame = tk.Frame(parent, bg="#4a4a4a")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="Son 12 Ay", font=FONT_BOLD, bg="#4a4a4a", fg="white", width=12).pack(side="left", padx=2, pady=5)
        for m in monthly_data['TRV']:
            tk.Label(header_frame, text=m['ay'], font=FONT_BOLD, bg="#4a4a4a", fg="white", width=7).pack(side="left", padx=2, pady=5)
        
        # TRV rows
        for metric in ['arac', 'hata', 'oran']:
            row_frame = tk.Frame(parent, bg="#e8f4f8" if metric != 'oran' else "#d4e8ed")
            row_frame.pack(fill="x")
            label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
            tk.Label(row_frame, text=f"TRV {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
            for m in monthly_data['TRV']:
                val = m[metric] if metric != 'oran' else f"{m['oran']:.0%}"
                tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)
        
        # TOU rows
        for metric in ['arac', 'hata', 'oran']:
            row_frame = tk.Frame(parent, bg="#fff5e6" if metric != 'oran' else "#ffe8cc")
            row_frame.pack(fill="x")
            label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
            tk.Label(row_frame, text=f"TOU {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
            for m in monthly_data['TOU']:
                val = m[metric] if metric != 'oran' else f"{m['oran']:.0%}"
                tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)

    def export_dashboard_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Ozet_Rapor.xlsx")
        if not file_path: return
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet oluştur ve genişlikleri ayarla
                writer.book.create_sheet('Özet Tablolar')
                ws = writer.book['Özet Tablolar']
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15

                # 1. Alt Grup
                if not self.df_alt.empty:
                    ws.cell(row=1, column=2, value="ALT GRUP ANALİZİ").font = ws.cell(row=1, column=2).font.copy(bold=True)
                    self.df_alt.to_excel(writer, sheet_name='Özet Tablolar', startrow=2, startcol=1, index=False)
                
                # 2. Top 5
                current_row = len(self.df_alt) + 5 if not self.df_alt.empty else 1
                if self.top_5_errors:
                    ws.cell(row=current_row, column=2, value="EN ÇOK GÖRÜLEN HATALAR (TOP 5)").font = ws.cell(row=current_row, column=2).font.copy(bold=True)
                    df_err = pd.DataFrame(self.top_5_errors, columns=["Hata Konumu", "Adet"])
                    df_err.to_excel(writer, sheet_name='Özet Tablolar', startrow=current_row+1, startcol=1, index=False)
                    current_row += len(df_err) + 4
                
                # 3. Aylık Trend
                if self.monthly_data:
                    ws.cell(row=current_row, column=2, value="AYLIK HATA ORANI TRENDİ").font = ws.cell(row=current_row, column=2).font.copy(bold=True)
                    df_trend = pd.DataFrame(self.monthly_data, columns=["Ay", "Araç Sayısı", "Hata Sayısı", "Oran"])
                    df_trend.to_excel(writer, sheet_name='Özet Tablolar', startrow=current_row+1, startcol=1, index=False)
            
            messagebox.showinfo("Başarılı", "Özet tablolar Excel'e aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulamadı: {str(e)}")

    def create_info_card(self, parent, title, value, icon, col):
        card = tk.Frame(parent, bg=COLOR_WHITE, padx=20, pady=20); card.grid(row=0, column=col, sticky="ew", padx=10)
        header = tk.Frame(card, bg=COLOR_WHITE); header.pack(fill="x")
        tk.Label(header, text=title, font=("Segoe UI", 9, "bold"), fg="#999", bg=COLOR_WHITE).pack(side="left")
        if icon: tk.Label(header, text=icon, font=("Segoe UI", 14), fg="#ccc", bg=COLOR_WHITE).pack(side="right")
        tk.Label(card, text=value, font=FONT_CARD_NUM, bg=COLOR_WHITE, fg="black").pack(anchor="w", pady=(10,0))

    def draw_simple_bar_chart(self, canvas, df):
        if df.empty: return
        w = 400; h = 250; max_val = df['sayi'].max() if not df.empty else 1; max_val = max_val if max_val > 0 else 1
        bar_width = 40; gap = 30; start_x = 50; base_y = h - 30
        canvas.create_line(40, 20, 40, base_y, fill="#eee", width=1); canvas.create_line(40, base_y, w, base_y, fill="#eee", width=1)
        for i, row in df.iterrows():
            val = row['sayi']; label = row['alt_grup']; bar_h = (val / max_val) * (h - 60)
            x0 = start_x + (i * (bar_width + gap)); y0 = base_y - bar_h; x1 = x0 + bar_width; y1 = base_y
            canvas.create_rectangle(x0, y0, x1, y1, fill="#111", outline="")
            canvas.create_text((x0+x1)/2, y1+15, text=label, font=("Segoe UI", 8), fill="#666")
            canvas.create_text((x0+x1)/2, y0-10, text=str(val), font=("Segoe UI", 9, "bold"), fill="#000")

    def draw_dynamic_line_chart(self, canvas, top_5_data):
        if not top_5_data: canvas.create_text(200, 125, text="Yeterli Veri Yok", fill="#999"); return 
        w = 400; h = 250; max_val = top_5_data[0][1]; max_val = max_val if max_val > 0 else 1
        base_y = h - 40; start_x = 50; step_x = (w - 100) / max(len(top_5_data)-1, 1)
        canvas.create_line(40, 20, 40, base_y, fill="#eee", width=1); canvas.create_line(40, base_y, w, base_y, fill="#eee", width=1)
        points = []
        for i, (label, val) in enumerate(top_5_data):
            x = start_x + (i * step_x); y = base_y - ((val / max_val) * (h - 80)); points.append((x, y))
            canvas.create_oval(x-4, y-4, x+4, y+4, fill="black", outline="white", width=2)
            canvas.create_text(x, base_y + 15, text=label[:10], font=("Segoe UI", 8), fill="#666")
            canvas.create_text(x, y - 15, text=str(val), font=("Segoe UI", 9, "bold"), fill="#000")
        if len(points) > 1: canvas.create_line(points, fill="#333", width=2, smooth=False)

    def draw_monthly_rate_chart(self, canvas, data):
        canvas.delete("all")
        if not data:
            canvas.create_text(400, 125, text="Görüntülenecek Veri Yok", fill="#999", font=FONT_NORMAL)
            return
        h = 250; pad_x = 60; pad_y = 40; point_spacing = 150 
        num_points = len(data); required_width = pad_x + (max(num_points - 1, 1) * point_spacing) + pad_x
        visible_width = 900; final_width = max(required_width, visible_width)
        canvas.config(scrollregion=(0, 0, final_width, h))
        max_rate = max([x[3] for x in data]) if data else 1; max_rate = max_rate if max_rate > 0 else 1
        graph_h = h - 2*pad_y
        canvas.create_line(pad_x, h-pad_y, final_width - pad_x, h-pad_y, fill="#ddd", width=2) 
        canvas.create_line(pad_x, pad_y, pad_x, h-pad_y, fill="#ddd", width=2)
        points = []
        for i, row in enumerate(data):
            month_label = row[0]; rate = row[3]; x = pad_x + (i * point_spacing); y = (h - pad_y) - ((rate / max_rate) * graph_h); points.append((x, y))
            canvas.create_line(x, h-pad_y, x, y, fill="#f0f0f0", dash=(2, 4))
            canvas.create_oval(x-6, y-6, x+6, y+6, fill=COLOR_CHART_LINE, outline="white", width=2)
            canvas.create_text(x, h-pad_y+20, text=month_label, font=("Segoe UI", 9), fill="#666")
            canvas.create_text(x, y-20, text=f"{rate:.2f}", font=("Segoe UI", 10, "bold"), fill=COLOR_CHART_LINE)
        if len(points) > 1: canvas.create_line(points, fill=COLOR_CHART_LINE, width=3, smooth=False)

    # ------------------------------------------------------
    # LİSTE EKRANI
    # ------------------------------------------------------
    def show_list_view(self):
        self.clear_content()
        filter_frame = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=15, pady=15); filter_frame.pack(fill="x", pady=(0, 10))
        tk.Label(filter_frame, text="Şasi No:", bg=COLOR_WHITE).grid(row=0, column=0, sticky="w", padx=5); self.f_sasi = ttk.Entry(filter_frame); self.f_sasi.grid(row=1, column=0, padx=5)
        tk.Label(filter_frame, text="Alt Grup:", bg=COLOR_WHITE).grid(row=0, column=1, sticky="w", padx=5); self.f_alt = ttk.Combobox(filter_frame, values=["Tümü"] + ALT_GRUP, state="readonly"); self.f_alt.current(0); self.f_alt.grid(row=1, column=1, padx=5)
        tk.Button(filter_frame, text="FİLTRELE", command=self.load_list_data, bg="black", fg="white", relief="flat").grid(row=1, column=2, padx=20)
        tk.Button(filter_frame, text="EXCEL İNDİR", command=self.export_excel_with_images, bg="#007bff", fg="white", relief="flat").grid(row=1, column=3)
        table_frame = tk.Frame(self.content_area, bg=COLOR_WHITE); table_frame.pack(fill="both", expand=True)
        cols = ("ID", "BB No", "Şasi No", "Araç", "Tarih", "Alt Grup", "Hata", "Ekleyen")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=vsb.set); vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        for c in cols: self.tree.heading(c, text=c); self.tree.column(c, width=100)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        bottom_container = tk.Frame(self.content_area, bg=COLOR_BG, pady=10); bottom_container.pack(fill="x")
        self.preview_frame = tk.Frame(bottom_container, bg=COLOR_WHITE, height=150, bd=1, relief="solid"); self.preview_frame.pack(side="left", fill="x", expand=True, padx=(0, 20)); self.preview_frame.pack_propagate(False)
        self.lbl_preview_img = tk.Label(self.preview_frame, text="Kayıt seçiniz...", bg=COLOR_WHITE, fg="#999"); self.lbl_preview_img.pack(side="left", padx=10, fill="y")
        self.btn_open_big = tk.Button(self.preview_frame, text="BÜYÜK GÖRÜNTÜLE", state="disabled", bg="#eee", relief="flat"); self.btn_open_big.pack(side="right", padx=10, pady=10); self.current_preview_path = None
        action_bar = tk.Frame(bottom_container, bg=COLOR_BG); action_bar.pack(side="right", fill="y")
        tk.Button(action_bar, text="DÜZENLE", command=self.open_edit_window, bg="#ffc107", relief="flat", width=15, height=2).pack(side="top", pady=2)
        if self.current_role == "admin": tk.Button(action_bar, text="SİL", command=self.delete_record, bg="#dc3545", fg="white", relief="flat", width=15, height=2).pack(side="top", pady=2)
        self.load_list_data()

    def on_tree_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        rec_id = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); path = conn.execute("SELECT fotograf_yolu FROM pdi_kayitlari WHERE id=?", (rec_id,)).fetchone()[0]; conn.close(); self.current_preview_path = path
        if path and os.path.exists(path):
            try:
                img = PilImage.open(path); base_height = 130; w_percent = (base_height / float(img.size[1])); w_size = int((float(img.size[0]) * float(w_percent)))
                img = img.resize((w_size, base_height), PilImage.Resampling.LANCZOS); photo = ImageTk.PhotoImage(img); self.lbl_preview_img.config(image=photo, text=""); self.lbl_preview_img.image = photo
                self.btn_open_big.config(state="normal", bg=COLOR_BLUE, fg="white", command=lambda: os.startfile(path))
            except: self.lbl_preview_img.config(image="", text="⚠️ Resim Önizlenemiyor"); self.btn_open_big.config(state="disabled", bg="#eee")
        else: self.lbl_preview_img.config(image="", text="📷 Dosya Yok"); self.btn_open_big.config(state="disabled", bg="#eee")

    def load_list_data(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        sasi = self.f_sasi.get(); alt = self.f_alt.get()
        q = "SELECT id, bb_no, sasi_no, arac_tipi, tarih_saat, alt_grup, hata_konumu, kullanici FROM pdi_kayitlari WHERE 1=1"
        p = []
        if sasi: q += " AND sasi_no LIKE ?"; p.append(f"%{sasi}%")
        if alt != "Tümü": q += " AND alt_grup = ?"; p.append(alt)
        q += " ORDER BY id DESC"
        conn = sqlite3.connect(DB_NAME); rows = conn.execute(q, p).fetchall(); conn.close()
        for r in rows:
            date_d = r[4].split(" ")[0] if r[4] else ""
            self.tree.insert("", "end", values=(r[0], r[1], r[2], r[3], date_d, r[5], r[6], r[7]))

    def clear_content(self):
        for widget in self.content_area.winfo_children(): widget.destroy()

    def export_excel_with_images(self):
        conn = sqlite3.connect(DB_NAME); query = "SELECT id, bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, tarih_saat, kullanici, duzenleyen, fotograf_yolu FROM pdi_kayitlari"; df = pd.read_sql_query(query, conn); conn.close()
        if df.empty: messagebox.showwarning("Uyarı", "Veri yok."); return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not file_path: return
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.drop(columns=['fotograf_yolu']).to_excel(writer, sheet_name='Kayıtlar', index=False)
                wb = writer.book; ws_img = wb.create_sheet("FOTOĞRAFLAR"); ws_img.column_dimensions['A'].width = 10; ws_img.column_dimensions['B'].width = 50
                ws_img.cell(row=1, column=1, value="ID"); ws_img.cell(row=1, column=2, value="FOTOĞRAF"); row_idx = 2
                from openpyxl.drawing.image import Image as XLImage
                for index, row in df.iterrows():
                    rid = row['id']; path = row['fotograf_yolu']; ws_img.cell(row=row_idx, column=1, value=rid); ws_img.row_dimensions[row_idx].height = 100
                    if path and os.path.exists(path):
                        try:
                            img = PilImage.open(path); img.thumbnail((300, 130)); xl_img = XLImage(path); xl_img.height = 130; xl_img.width = 130 * (img.width / img.height); ws_img.add_image(xl_img, f"B{row_idx}")
                        except Exception as e: ws_img.cell(row=row_idx, column=2, value=f"Hata: {str(e)}")
                    else: ws_img.cell(row=row_idx, column=2, value="Resim Yok")
                    row_idx += 1
            messagebox.showinfo("Başarılı", "Excel başarıyla oluşturuldu.")
        except Exception as e: messagebox.showerror("Hata", f"Excel oluşturulurken hata: {e}")

    def delete_record(self):
        sel = self.tree.selection(); 
        if not sel: return
        if messagebox.askyesno("Sil", "Kayıt silinsin mi?"): rid = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); conn.execute("DELETE FROM pdi_kayitlari WHERE id=?", (rid,)); conn.commit(); conn.close(); self.load_list_data()

    def open_add_window(self): self.popup_window("Yeni Kayıt Ekle")
    def open_edit_window(self):
        sel = self.tree.selection(); 
        if not sel: return
        rid = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); row = conn.execute("SELECT * FROM pdi_kayitlari WHERE id=?", (rid,)).fetchone(); conn.close(); self.popup_window("Kayıt Düzenle", row)

    def open_user_management(self):
        win = tk.Toplevel(self.root); win.title("Kullanıcı Yönetimi"); win.geometry("800x500"); win.configure(bg=COLOR_BG)
        tk.Label(win, text="KULLANICI YÖNETİMİ", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=10)
        main_frame = tk.Frame(win, bg=COLOR_BG); main_frame.pack(fill="both", expand=True, padx=20, pady=10)
        left_frame = tk.Frame(main_frame, bg=COLOR_WHITE, padx=10, pady=10); left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        tk.Label(left_frame, text="Kayıtlı Kullanıcılar", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w", pady=(0,5))
        lb_users = tk.Listbox(left_frame, font=("Segoe UI", 11), borderwidth=0, bg="#f9f9f9", selectbackground="#ddd", selectforeground="black"); lb_users.pack(fill="both", expand=True)
        right_frame = tk.Frame(main_frame, bg=COLOR_WHITE, padx=10, pady=10); right_frame.pack(side="right", fill="both", expand=True)
        tk.Label(right_frame, text="Yeni Kullanıcı Ekle / Düzenle", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w", pady=(0,15))
        tk.Label(right_frame, text="Kullanıcı Adı:", bg=COLOR_WHITE).pack(anchor="w"); e_u = ttk.Entry(right_frame); e_u.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Şifre:", bg=COLOR_WHITE).pack(anchor="w"); e_p = ttk.Entry(right_frame); e_p.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Rol:", bg=COLOR_WHITE).pack(anchor="w"); cb_role = ttk.Combobox(right_frame, values=["Admin", "Kullanıcı"], state="readonly"); cb_role.current(1); cb_role.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Açıklama:", bg=COLOR_WHITE).pack(anchor="w"); e_desc = ttk.Entry(right_frame); e_desc.pack(fill="x", pady=(0,20))
        def load_users():
            lb_users.delete(0, "end"); conn = sqlite3.connect(DB_NAME); rows = conn.execute("SELECT username, role FROM users").fetchall(); conn.close()
            for r in rows: lb_users.insert("end", r[0] + (" (Yönetici)" if r[1] == 1 else " (Kullanıcı)"))
        
        def on_user_select(event):
            selection = lb_users.curselection()
            if not selection: return
            text = lb_users.get(selection[0])
            username = text.split(" (")[0]
            conn = sqlite3.connect(DB_NAME)
            user_data = conn.execute("SELECT username, password, role, aciklama FROM users WHERE username=?", (username,)).fetchone()
            conn.close()
            if user_data:
                e_u.delete(0, "end"); e_p.delete(0, "end"); e_desc.delete(0, "end")
                e_u.insert(0, user_data[0]); e_p.insert(0, user_data[1])
                cb_role.set("Admin" if user_data[2] == 1 else "Kullanıcı")
                if user_data[3]: e_desc.insert(0, user_data[3])

        lb_users.bind("<<ListboxSelect>>", on_user_select)

        def save_user():
            u = e_u.get().strip(); p = e_p.get().strip(); role = 1 if cb_role.get() == "Admin" else 0; desc = e_desc.get().strip()
            if not u or not p: messagebox.showwarning("Hata", "Kullanıcı adı ve şifre zorunludur."); return
            conn = sqlite3.connect(DB_NAME)
            try: conn.execute("INSERT OR REPLACE INTO users (username, password, role, aciklama) VALUES (?,?,?,?)", (u, p, role, desc)); conn.commit(); messagebox.showinfo("Başarılı", "Kullanıcı kaydedildi."); e_u.delete(0, "end"); e_p.delete(0, "end"); e_desc.delete(0, "end"); load_users()
            except Exception as e: messagebox.showerror("Hata", str(e));
            finally: conn.close()
        def delete_user():
            sel = lb_users.curselection(); 
            if not sel: return
            username = lb_users.get(sel[0]).split(" (")[0]
            if username == "admin": messagebox.showerror("Hata", "Ana 'admin' silinemez."); return
            if messagebox.askyesno("Sil", f"{username} silinsin mi?"): conn = sqlite3.connect(DB_NAME); conn.execute("DELETE FROM users WHERE username=?", (username,)); conn.commit(); conn.close(); load_users(); messagebox.showinfo("Silindi", "Kullanıcı silindi.")
        tk.Button(right_frame, text="KAYDET", command=save_user, bg="black", fg="white", relief="flat", height=2).pack(fill="x", pady=(0,5))
        tk.Button(right_frame, text="SEÇİLİ KULLANICIYI SİL", command=delete_user, bg=COLOR_RED, fg="white", relief="flat").pack(fill="x"); load_users()

    def popup_window(self, title, data=None):
        top = tk.Toplevel(self.root); top.title(title); top.geometry("950x750"); top.configure(bg=COLOR_BG)
        header_frame = tk.Frame(top, bg="black", pady=10, padx=20); header_frame.pack(fill="x")
        tk.Label(header_frame, text=title.upper(), font=("Segoe UI", 14, "bold"), bg="black", fg="white").pack(side="left")
        main_frame = tk.Frame(top, bg=COLOR_WHITE, padx=20, pady=20); main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        info_lbl = tk.Label(main_frame, text="ARAÇ BİLGİLERİ", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888"); info_lbl.pack(anchor="w", pady=(0,10))
        grid_frame = tk.Frame(main_frame, bg=COLOR_WHITE); grid_frame.pack(fill="x", pady=(0, 20))
        
        # --- TARİH SEÇİCİ İÇİN ÖZEL ALAN ---
        def open_calendar(entry_widget):
            current_val = entry_widget.get()
            dialog = AdvancedCalendarDialog(top, current_val)
            if dialog.result:
                entry_widget.delete(0, "end")
                entry_widget.insert(0, dialog.result)

        def add_field(lbl, r, c, val=None, combo=None, is_date=False):
            tk.Label(grid_frame, text=lbl, bg=COLOR_WHITE, font=FONT_NORMAL, fg="#333").grid(row=r, column=c, sticky="w", padx=10)
            
            if is_date:
                f = tk.Frame(grid_frame, bg=COLOR_WHITE)
                f.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
                w = ttk.Entry(f, font=FONT_NORMAL)
                w.pack(side="left", fill="x", expand=True)
                if val: w.insert(0, val)
                btn_cal = tk.Button(f, text="📅", command=lambda: open_calendar(w), 
                                    relief="flat", bg="#eee", width=3)
                btn_cal.pack(side="right", padx=(5,0))
            elif combo: 
                w = ttk.Combobox(grid_frame, values=combo, state="readonly", font=FONT_NORMAL)
                if not val: w.current(0) 
                else: w.set(val)
                w.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
            else: 
                w = ttk.Entry(grid_frame, font=FONT_NORMAL)
                if val: w.insert(0, val)
                w.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
            return w
            
        grid_frame.columnconfigure(0, weight=1); grid_frame.columnconfigure(1, weight=1); grid_frame.columnconfigure(2, weight=1)
        
        e_bb = add_field("BB No", 0, 0, data[1] if data else "")
        e_sasi = add_field("Şasi No", 0, 1, data[2] if data else "")
        e_arac = add_field("Araç Tipi", 0, 2, data[3] if data else "", ARAC_TIPI)
        e_isemri = add_field("İş Emri No", 2, 0, data[4] if data else "")
        
        today_str = datetime.now().strftime("%d-%m-%Y")
        e_tarih = add_field("Tarih (G-A-Y)", 2, 1, data[9].split(" ")[0] if data and data[9] else today_str, is_date=True)
        
        e_alt = add_field("Alt Grup", 2, 2, data[5] if data else "", ALT_GRUP)
        
        split_frame = tk.Frame(main_frame, bg=COLOR_WHITE); split_frame.pack(fill="both", expand=True)
        left_col = tk.Frame(split_frame, bg=COLOR_WHITE, width=300); left_col.pack(side="left", fill="both", expand=False, padx=(0, 20))
        
        # HATA KONUMU - Artık serbest metin alanı
        tk.Label(left_col, text="HATA KONUMU", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w")
        tk.Label(left_col, text="(Virgülle ayırarak yazın)", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#aaa").pack(anchor="w")
        e_hata_konumu = tk.Text(left_col, height=4, font=FONT_NORMAL, bg="#f9f9f9", bd=1, relief="solid")
        e_hata_konumu.pack(fill="x", pady=5)
        if data and data[7]: e_hata_konumu.insert("1.0", data[7])
        
        # TOP HATA - Sadece düzenleme modunda görünür (kayıt sonrası eklenir)
        e_top_hata = None
        if data:
            tk.Label(left_col, text="TOP HATA", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w", pady=(15, 0))
            tk.Label(left_col, text="(Kayıt sonrası seçilir)", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#aaa").pack(anchor="w")
            
            # Top Hata listesini veritabanından al
            conn = sqlite3.connect(DB_NAME)
            top_hatalar = conn.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()
            conn.close()
            top_hata_list = [""] + [h[0] for h in top_hatalar]
            
            e_top_hata = ttk.Combobox(left_col, values=top_hata_list, state="readonly", font=FONT_NORMAL)
            e_top_hata.pack(fill="x", pady=5)
            
            # Mevcut top_hata değerini bul ve seç
            try:
                current_top_hata = conn.execute("SELECT top_hata FROM pdi_kayitlari WHERE id=?", (data[0],)).fetchone()
                if current_top_hata and current_top_hata[0]:
                    conn = sqlite3.connect(DB_NAME)
                    current_top_hata = conn.execute("SELECT top_hata FROM pdi_kayitlari WHERE id=?", (data[0],)).fetchone()
                    conn.close()
                    if current_top_hata and current_top_hata[0] in top_hata_list:
                        e_top_hata.set(current_top_hata[0])
            except: pass
        
        right_col = tk.Frame(split_frame, bg=COLOR_WHITE); right_col.pack(side="right", fill="both", expand=True)
        tk.Label(right_col, text="TESPİT DETAYLARI", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w")
        txt_tespit = tk.Text(right_col, height=6, font=FONT_NORMAL, bg="#f9f9f9", bd=1, relief="solid"); txt_tespit.pack(fill="x", pady=5)
        if data and data[6]: txt_tespit.insert("1.0", data[6])
        
        # FOTOĞRAF KANITI - Sadece düzenleme modunda görünür
        self.selected_photo_path = ""
        if data:
            tk.Label(right_col, text="FOTOĞRAF KANITI", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w", pady=(10,0))
            photo_frame = tk.Frame(right_col, bg="#f9f9f9", bd=1, relief="solid", pady=10, padx=10); photo_frame.pack(fill="x", pady=5)
            self.selected_photo_path = data[8] if data[8] else ""
            lbl_path = tk.Label(photo_frame, text=self.selected_photo_path if self.selected_photo_path else "Dosya seçilmedi...", bg="#f9f9f9", fg="#555", anchor="w"); lbl_path.pack(side="left", fill="x", expand=True)
            def sel_photo():
                f = filedialog.askopenfilename(filetypes=[("Resimler", "*.jpg *.png *.jpeg")]); 
                if f: self.selected_photo_path = f; lbl_path.config(text=os.path.basename(f))
            tk.Button(photo_frame, text="GÖZAT...", command=sel_photo, bg="#ddd", relief="flat").pack(side="right")
        
        def save():
            bb = e_bb.get(); sasi = e_sasi.get()
            hata_str = e_hata_konumu.get("1.0", "end-1c").strip()
            top_hata_val = e_top_hata.get() if e_top_hata else ""
            
            if not sasi: messagebox.showwarning("Eksik", "Şasi No giriniz."); return
            conn = sqlite3.connect(DB_NAME); c = conn.cursor()
            if data: 
                c.execute("""UPDATE pdi_kayitlari SET bb_no=?, sasi_no=?, arac_tipi=?, is_emri_no=?, alt_grup=?, tespitler=?, hata_konumu=?, fotograf_yolu=?, tarih_saat=?, duzenleyen=?, top_hata=? WHERE id=?""", 
                         (bb, sasi, e_arac.get(), e_isemri.get(), e_alt.get(), txt_tespit.get("1.0", "end-1c"), hata_str, self.selected_photo_path, e_tarih.get(), self.current_user, top_hata_val, data[0]))
            else: 
                c.execute("""INSERT INTO pdi_kayitlari (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, fotograf_yolu, tarih_saat, kullanici) VALUES (?,?,?,?,?,?,?,?,?,?)""", 
                         (bb, sasi, e_arac.get(), e_isemri.get(), e_alt.get(), txt_tespit.get("1.0", "end-1c"), hata_str, "", e_tarih.get(), self.current_user))
            conn.commit(); conn.close(); top.destroy(); self.show_dashboard(); messagebox.showinfo("Başarılı", "Kayıt işlemi tamamlandı.")
        tk.Button(main_frame, text="KAYDET VE KAPAT", command=save, bg="black", fg="white", font=("Segoe UI", 12, "bold"), height=2, relief="flat").pack(fill="x", pady=20)

    # ------------------------------------------------------
    # TOP HATA YÖNETİMİ (ADMIN)
    # ------------------------------------------------------
    def open_top_hata_management(self):
        win = tk.Toplevel(self.root)
        win.title("Top Hata Yönetimi")
        win.geometry("600x500")
        win.configure(bg=COLOR_BG)
        
        tk.Label(win, text="TOP HATA YÖNETİMİ", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=15)
        
        main_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Liste
        tk.Label(main_frame, text="Kayıtlı Top Hatalar", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w")
        
        list_frame = tk.Frame(main_frame, bg="#f9f9f9", bd=1, relief="solid")
        list_frame.pack(fill="both", expand=True, pady=10)
        
        lb_hatalar = tk.Listbox(list_frame, font=FONT_NORMAL, height=12, bd=0, bg="#f9f9f9", selectbackground="#ddd", selectforeground="black")
        lb_hatalar.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        sb = tk.Scrollbar(list_frame, orient="vertical", command=lb_hatalar.yview)
        sb.pack(side="right", fill="y")
        lb_hatalar.config(yscrollcommand=sb.set)
        
        def load_hatalar():
            lb_hatalar.delete(0, "end")
            conn = sqlite3.connect(DB_NAME)
            rows = conn.execute("SELECT hata_adi, aktif FROM top_hatalar ORDER BY hata_adi").fetchall()
            conn.close()
            for r in rows:
                status = " ✓" if r[1] == 1 else " (pasif)"
                lb_hatalar.insert("end", r[0] + status)
        
        # Ekleme alanı
        add_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        add_frame.pack(fill="x", pady=10)
        
        tk.Label(add_frame, text="Yeni Top Hata:", bg=COLOR_WHITE, font=FONT_NORMAL).pack(side="left")
        e_new = ttk.Entry(add_frame, font=FONT_NORMAL, width=40)
        e_new.pack(side="left", padx=10)
        
        def add_hata():
            hata = e_new.get().strip()
            if not hata:
                messagebox.showwarning("Uyarı", "Hata adı boş olamaz.")
                return
            try:
                conn = sqlite3.connect(DB_NAME)
                conn.execute("INSERT INTO top_hatalar (hata_adi, aktif, olusturma_tarihi) VALUES (?, 1, ?)", 
                           (hata, datetime.now().strftime("%d-%m-%Y")))
                conn.commit()
                conn.close()
                e_new.delete(0, "end")
                load_hatalar()
                messagebox.showinfo("Başarılı", f"'{hata}' eklendi.")
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu hata zaten mevcut.")
        
        tk.Button(add_frame, text="EKLE", command=add_hata, bg=COLOR_GREEN, fg="white", relief="flat", padx=15).pack(side="left")
        
        # Silme butonu
        def delete_hata():
            sel = lb_hatalar.curselection()
            if not sel:
                messagebox.showwarning("Uyarı", "Silmek için bir hata seçin.")
                return
            hata = lb_hatalar.get(sel[0]).replace(" ✓", "").replace(" (pasif)", "")
            if messagebox.askyesno("Sil", f"'{hata}' silinsin mi?"):
                conn = sqlite3.connect(DB_NAME)
                conn.execute("DELETE FROM top_hatalar WHERE hata_adi=?", (hata,))
                conn.commit()
                conn.close()
                load_hatalar()
                messagebox.showinfo("Silindi", "Hata silindi.")
        
        btn_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="SEÇİLİ HATAYI SİL", command=delete_hata, bg=COLOR_RED, fg="white", relief="flat", padx=15).pack(side="right")
        
        load_hatalar()

    # ------------------------------------------------------
    # EXCEL IMPORT (ADMIN)
    # ------------------------------------------------------
    def open_excel_import(self):
        win = tk.Toplevel(self.root)
        win.title("Geçmiş Veri Yükle")
        win.geometry("700x550")
        win.configure(bg=COLOR_BG)
        
        tk.Label(win, text="GEÇMİŞ VERİ YÜKLE (EXCEL)", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=15)
        
        main_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Açıklama
        info_text = """Excel dosyanız aşağıdaki sütunları içermelidir (Otomatik tanınır):
        
Sıralama Örneği:
BB No | Şasi No | Araç Tipi | İş Emri No | PDI Tarihi | Tespitler | Hata Konumu | Alt Grup

⚠️ "Kullanıcı" sütunu gerekmez (otomatik atanır).
⚠️ Araç Tipi Kısaltmaları Desteklenir:
   - TOU -> Tourismo
   - TRV -> Travego
   - CON -> Connecto
   - CONNECTO -> Connecto
"""
        
        tk.Label(main_frame, text=info_text, font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555", justify="left").pack(anchor="w", pady=(0, 20))
        
        # Dosya seçici
        file_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        file_frame.pack(fill="x", pady=10)
        
        self.import_file_path = tk.StringVar()
        tk.Label(file_frame, text="Excel Dosyası:", bg=COLOR_WHITE, font=FONT_BOLD).pack(side="left")
        tk.Entry(file_frame, textvariable=self.import_file_path, font=FONT_NORMAL, width=50, state="readonly").pack(side="left", padx=10)
        
        def browse_file():
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
            if f:
                self.import_file_path.set(f)
        
        tk.Button(file_frame, text="GÖZAT", command=browse_file, bg="#ddd", relief="flat", padx=10).pack(side="left")
        
        # Sonuç alanı
        result_frame = tk.Frame(main_frame, bg="#f9f9f9", bd=1, relief="solid")
        result_frame.pack(fill="both", expand=True, pady=10)
        
        result_text = tk.Text(result_frame, font=("Consolas", 9), bg="#f9f9f9", height=10, bd=0)
        result_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        def import_data():
            file_path = self.import_file_path.get()
            if not file_path:
                messagebox.showwarning("Uyarı", "Lütfen bir Excel dosyası seçin.")
                return
            
            try:
                # Tüm sütunları string olarak oku
                df = pd.read_excel(file_path, dtype=str)
                result_text.delete("1.0", "end")
                
                # Sütun isimlerini normalize et (küçük harf, trim)
                df.columns = [str(c).strip() for c in df.columns]
                
                # Gerekli sütunları bulmaya çalış (Index tabanlı yaklaşım fallback olarak)
                # Beklenen sıra: BB, Sasi, Tip, IsEmri, Tarih, Tespit, HataKon, Alt
                
                success_count = 0
                error_count = 0
                
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                
                # Kısaltma sözlüğü
                type_map = {
                    "TOU": "Tourismo", "TRV": "Travego", "CON": "Connecto",
                    "TOURISMO": "Tourismo", "TRAVEGO": "Travego", "CONNECTO": "Connecto",
                    "TOURİSMO": "Tourismo"
                }

                for idx, row in df.iterrows():
                    try:
                        # İndeks bazlı okuma (En güvenilir yöntem kullanıcı standart formatı için)
                        # 0: BB, 1: Sasi, 2: Tip, 3: IsEmri, 4: Tarih, 5: Tespit, 6: HataKon, 7: Alt
                        # Eğer sütun sayısı az ise hata vermemesi için kontrol
                        vals = [str(v).strip() if pd.notna(v) and str(v).lower() != 'nan' else "" for v in row.values]
                        
                        bb_no = vals[0] if len(vals) > 0 else ""
                        sasi_no = vals[1] if len(vals) > 1 else ""
                        raw_type = vals[2] if len(vals) > 2 else ""
                        is_emri_no = vals[3] if len(vals) > 3 else ""
                        raw_date = vals[4] if len(vals) > 4 else ""
                        tespitler = vals[5] if len(vals) > 5 else ""
                        hata_konumu = vals[6] if len(vals) > 6 else ""
                        alt_grup = vals[7] if len(vals) > 7 else ""
                        
                        # Top Hata geçmiş veride olmayabilir, boş geç
                        top_hata = "" 
                        
                        # Araç Tipi Dönüştürme
                        arac_tipi = type_map.get(raw_type.upper(), raw_type)
                        
                        # Tarih formatlama (Excel bazen 2024-01-05 00:00:00 string veriyor)
                        tarih = raw_date.split(" ")[0] # Sadece tarihi al
                        # Eğer tarih boşsa bugünü atama, boş bırak veya uyar
                        if not tarih:
                             result_text.insert("end", f"⚠️ Satır {idx+2}: Tarih boş, atlanıyor.\n")
                             error_count += 1
                             continue

                        # Zorunlu alan kontrolü
                        if not sasi_no:
                            continue # Boş satır
                        
                        if arac_tipi not in ARAC_TIPI:
                            result_text.insert("end", f"❌ Satır {idx+2}: Tanımsız araç tipi '{raw_type}' -> '{arac_tipi}'\n")
                            error_count += 1
                            continue
                        
                        kullanici = self.current_user if hasattr(self, 'current_user') else "admin"

                        c.execute("""INSERT INTO pdi_kayitlari 
                                    (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, top_hata, tarih_saat, kullanici) 
                                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                                 (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, top_hata, tarih, kullanici))
                        success_count += 1
                        
                    except Exception as e:
                        result_text.insert("end", f"❌ Satır {idx+2}: Hata: {str(e)}\n")
                        error_count += 1
                
                conn.commit()
                conn.close()
                
                result_text.insert("end", f"\n{'='*50}\n")
                result_text.insert("end", f"✅ Başarılı: {success_count} kayıt\n")
                result_text.insert("end", f"❌ Hatalı: {error_count} kayıt\n")
                
                messagebox.showinfo("Tamamlandı", f"{success_count} kayıt başarıyla yüklendi.\n{error_count} kayıt hatalı.")
                
            except Exception as e:
                messagebox.showerror("Hata", f"Excel okunamadı: {str(e)}")
        
        tk.Button(main_frame, text="VERİLERİ YÜKLE", command=import_data, bg="black", fg="white", font=("Segoe UI", 12, "bold"), height=2, relief="flat").pack(fill="x", pady=10)


# ------------------------------------------------------
# MAIN (GÜNCELLENDİ - VPN/DOSYA KONTROLÜ EKLENDİ)
# ------------------------------------------------------
if __name__ == "__main__":
    db_folder = os.path.dirname(os.path.abspath(DB_NAME))
    if not os.access(db_folder, os.W_OK):
         root = tk.Tk()
         root.withdraw()
         messagebox.showerror("Bağlantı Hatası", 
                              "Veritabanı konumuna erişilemiyor!\n\n"
                              "Lütfen şunları kontrol edin:\n"
                              "1. VPN veya Şirket Ağına bağlı mısınız?\n"
                              "2. Ortak klasöre erişim izniniz var mı?")
         root.destroy()
    else:
        init_db()  # Veritabanı tablolarını oluştur
        root = tk.Tk()
        try: from ctypes import windll; windll.shcore.SetProcessDpiAwareness(1)
        except: pass
        app = PDIApp(root)
        root.mainloop()