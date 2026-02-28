import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import os
import base64
import calendar
import shutil
from collections import Counter
from PIL import Image as PilImage, ImageTk
import threading

# ------------------------------------------------------
# SABİTLER & AYARLAR
# ------------------------------------------------------
ALT_GRUP = ["Boya", "Süsleme", "Mekanik", "Elektrik"]
HATA_KONUM = [
    "Aydınlatma","Ayırma Duvarı","Ayna","Boya","CAM","Çıta","Defroster","Etiket",
    "Kapak","Kapı","Kaplama","Kelepçe","Klima","Koltuk","Körük","Lamba","Montaj",
    "Motor","Mutfak","Paket Raf","Silecek","Siperlik","Stepne","Şoför yatma yeri",
    "Tapa","Telefon","Torpido","Yazı"
]
ARAC_TIPI = ["Tourismo", "Connecto", "Travego"]

# ŞİRKET ORTAK ALAN AYARLARI
APP_DIR = r"F:\M_AB\10. KİŞİSEL KLASÖRLER\Deha\PDI-App"
DB_NAME = os.path.join(APP_DIR, "pdi_veritabani.db")
PHOTO_DIR = os.path.join(APP_DIR, "photos")

# Klasörleri kontrol et
if not os.path.exists(APP_DIR):
    try: os.makedirs(APP_DIR)
    except: pass
if not os.path.exists(PHOTO_DIR):
    try: os.makedirs(PHOTO_DIR)
    except: pass

# RENK PALETİ
COLOR_SIDEBAR = "#111111"      
COLOR_SIDEBAR_ACT = "#1f1f1f" 
COLOR_BG = "#F0F2F5"           
COLOR_WHITE = "#FFFFFF"        
COLOR_TEXT_MAIN = "#000000"    
COLOR_ACCENT = "#000000"       
COLOR_GREEN = "#28a745"        
COLOR_RED = "#dc3545"          
COLOR_BLUE = "#007bff"
COLOR_CHART_LINE = "#0056b3"

FONT_HEADER = ("Segoe UI", 14, "bold")
FONT_NORMAL = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_CARD_NUM = ("Segoe UI", 32, "bold")

LOGO_DATA = "" 

TURKISH_MONTHS = {
    1: "OCAK", 2: "ŞUBAT", 3: "MART", 4: "NİSAN", 5: "MAYIS", 6: "HAZİRAN",
    7: "TEMMUZ", 8: "AĞUSTOS", 9: "EYLÜL", 10: "EKİM", 11: "KASIM", 12: "ARALIK"
}
TURKISH_DAYS = {
    0: "PAZARTESİ", 1: "SALI", 2: "ÇARŞAMBA", 3: "PERŞEMBE", 4: "CUMA", 5: "CUMARTESİ", 6: "PAZAR"
}

# ------------------------------------------------------
# GELİŞMİŞ TAKVİM SINIFI (HIZLI AY SEÇİMİ EKLENDİ)
# ------------------------------------------------------
class AdvancedCalendarDialog(tk.Toplevel):
    def __init__(self, parent, current_date_str=None):
        super().__init__(parent)
        self.title("Tarih Seç")
        self.geometry("320x320")
        self.configure(bg=COLOR_WHITE)
        self.resizable(False, False)
        self.result = None
        
        # Başlangıç Tarihi
        self.now = datetime.now()
        self.year = self.now.year
        self.month = self.now.month
        
        if current_date_str:
            try:
                d = datetime.strptime(current_date_str, "%d-%m-%Y")
                self.year = d.year
                self.month = d.month
            except: pass

        # UI Ana Çerçeve
        self.container = tk.Frame(self, bg=COLOR_WHITE, padx=10, pady=10)
        self.container.pack(fill="both", expand=True)
        
        # Üst Kontroller
        self.header_frame = tk.Frame(self.container, bg=COLOR_WHITE)
        self.header_frame.pack(fill="x", pady=(0, 10))
        
        self.btn_prev = tk.Button(self.header_frame, text="<", command=self.prev_month, relief="flat", bg="#f0f0f0", width=3)
        self.btn_prev.pack(side="left")
        
        self.btn_month_year = tk.Button(self.header_frame, text="", command=self.toggle_view, relief="flat", bg=COLOR_WHITE, font=("Segoe UI", 11, "bold"), cursor="hand2")
        self.btn_month_year.pack(side="left", fill="x", expand=True)
        
        self.btn_next = tk.Button(self.header_frame, text=">", command=self.next_month, relief="flat", bg="#f0f0f0", width=3)
        self.btn_next.pack(side="left")

        # İçerik Alanı (Günler veya Aylar)
        self.content_frame = tk.Frame(self.container, bg=COLOR_WHITE)
        self.content_frame.pack(fill="both", expand=True)
        
        self.view_mode = "days" # veya "months"
        self.draw_days_view()
        
        # Modal Ayarlar
        self.transient(parent)
        self.grab_set()
        self.wait_window()

    def toggle_view(self):
        if self.view_mode == "days":
            self.view_mode = "months"
            self.draw_months_view()
        else:
            self.view_mode = "days"
            self.draw_days_view()

    def draw_days_view(self):
        # Temizlik
        for widget in self.content_frame.winfo_children(): widget.destroy()
        self.btn_prev.config(state="normal")
        self.btn_next.config(state="normal")
        
        # Başlık Güncelle
        tr_month = list(TURKISH_MONTHS.values())[self.month-1]
        self.btn_month_year.config(text=f"{tr_month} {self.year}")
        
        # Gün İsimleri
        days = ["Pt", "Sa", "Ça", "Pe", "Cu", "Ct", "Pa"]
        for i, d in enumerate(days):
            tk.Label(self.content_frame, text=d, font=("Segoe UI", 9, "bold"), bg=COLOR_WHITE, fg="#888").grid(row=0, column=i, sticky="nsew")

        # Takvim Günleri
        cal = calendar.monthcalendar(self.year, self.month)
        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day != 0:
                    btn = tk.Button(self.content_frame, text=str(day), 
                                    command=lambda d=day: self.select_date(d),
                                    relief="flat", bg=COLOR_WHITE, activebackground=COLOR_BLUE, activeforeground="white")
                    btn.grid(row=r+1, column=c, sticky="nsew", padx=1, pady=1)
                    
                    if day == self.now.day and self.month == self.now.month and self.year == self.now.year:
                        btn.config(fg=COLOR_BLUE, font=("Segoe UI", 9, "bold"))

        for i in range(7): self.content_frame.columnconfigure(i, weight=1)
        for i in range(6): self.content_frame.rowconfigure(i+1, weight=1)

    def draw_months_view(self):
        # Temizlik
        for widget in self.content_frame.winfo_children(): widget.destroy()
        
        # Yıl Seçici Modu
        self.btn_month_year.config(text=f"{self.year}")
        self.btn_prev.config(command=lambda: self.change_year(-1))
        self.btn_next.config(command=lambda: self.change_year(1))
        
        # Aylar Grid
        months = list(TURKISH_MONTHS.values())
        for i, m_name in enumerate(months):
            r, c = divmod(i, 3) # 4 satır, 3 sütun
            btn = tk.Button(self.content_frame, text=m_name, 
                            command=lambda m=i+1: self.select_month_from_grid(m),
                            relief="flat", bg="#f9f9f9", activebackground="#e0e0e0")
            btn.grid(row=r, column=c, sticky="nsew", padx=2, pady=2)
            if (i+1) == self.month:
                btn.config(bg=COLOR_BLUE, fg="white")

        for i in range(3): self.content_frame.columnconfigure(i, weight=1)
        for i in range(4): self.content_frame.rowconfigure(i, weight=1)

    def change_year(self, delta):
        self.year += delta
        self.btn_month_year.config(text=f"{self.year}")
        # Görünüm yenilemeye gerek yok, sadece label değişti ama aylar aynı

    def select_month_from_grid(self, month_idx):
        self.month = month_idx
        self.view_mode = "days"
        # Buton komutlarını eski haline getir
        self.btn_prev.config(command=self.prev_month)
        self.btn_next.config(command=self.next_month)
        self.draw_days_view()

    def prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month = 12
            self.year -= 1
        self.draw_days_view()

    def next_month(self):
        self.month += 1
        if self.month == 13:
            self.month = 1
            self.year += 1
        self.draw_days_view()

    def select_date(self, day):
        self.result = f"{day:02d}-{self.month:02d}-{self.year}"
        self.destroy()

# ------------------------------------------------------
# VERİTABANI İŞLEMLERİ
# ------------------------------------------------------
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users(username TEXT PRIMARY KEY, password TEXT, role INT, aciklama TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS pdi_kayitlari(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bb_no TEXT, sasi_no TEXT, arac_tipi TEXT, is_emri_no TEXT, alt_grup TEXT,
        tespitler TEXT, hata_konumu TEXT, fotograf_yolu TEXT, tarih_saat TEXT,
        kullanici TEXT, duzenleyen TEXT, grup_no TEXT, parca_tanimi TEXT,
        hata_tanimi TEXT, musteri_sikayeti TEXT, musteri_beklentisi TEXT,
        musteri_geri_bildirimi TEXT, musteri_teyidi TEXT, musteri_talebi TEXT,
        musteri_adi TEXT, musteri_soyadi TEXT, musteri_iletisim TEXT, musteri_adresi TEXT
    )""")
    # Top Hatalar tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS top_hatalar(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hata_adi TEXT UNIQUE,
        aktif INTEGER DEFAULT 1,
        olusturma_tarihi TEXT
    )""")
    # İmalata Gönderilenler tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS imalat_kayitlari(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        arac_no TEXT,
        tarih TEXT,
        adet INTEGER,
        top_hata TEXT,
        hata_metni TEXT,
        durum TEXT DEFAULT 'Bekleniyor',
        kullanici TEXT,
        olusturma_tarihi TEXT
    )""")
    # Manuel Rapor Verileri tablosu
    c.execute("""CREATE TABLE IF NOT EXISTS report_manual_data(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_type TEXT,
        context_key TEXT,
        data_key TEXT,
        data_value TEXT,
        UNIQUE(report_type, context_key, data_key)
    )""")
    cols = ["grup_no", "parca_tanimi", "hata_tanimi", "musteri_sikayeti", "musteri_beklentisi", "musteri_geri_bildirimi", "musteri_teyidi", "musteri_talebi", "musteri_adi", "musteri_soyadi", "musteri_iletisim", "musteri_adresi", "duzenleyen", "top_hata"]
    for col in cols:
        try: c.execute(f"ALTER TABLE pdi_kayitlari ADD COLUMN {col} TEXT")
        except: pass
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        c.execute("INSERT INTO users VALUES ('admin', 'admin123', 1, 'Sistem Yöneticisi')")
    # Tarih Normalizasyonu (TÜM FORMATLAR -> DD-MM-YYYY)
    all_rows = c.execute("SELECT id, tarih_saat FROM pdi_kayitlari").fetchall()
    for row_id, raw_dt in all_rows:
        if not raw_dt: continue
        new_dt = raw_dt.strip().replace("/", "-").replace(".", "-")
        parts = new_dt.split("-")
        try:
            if len(parts) == 3:
                # YYYY-MM-DD -> DD-MM-YYYY
                if len(parts[0]) == 4:
                    new_dt = f"{parts[2].zfill(2)}-{parts[1].zfill(2)}-{parts[0]}"
                # D-M-YYYY -> DD-MM-YYYY
                else:
                    new_dt = f"{parts[0].zfill(2)}-{parts[1].zfill(2)}-{parts[2]}"
                if new_dt != raw_dt:
                    c.execute("UPDATE pdi_kayitlari SET tarih_saat=? WHERE id=?", (new_dt, row_id))
        except: pass
    
    # ===== PERFORMANS OPTİMİZASYONU: INDEX'LER =====
    # Sık kullanılan sorgular için index oluştur
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_tarih_saat ON pdi_kayitlari(tarih_saat)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_arac_tipi ON pdi_kayitlari(arac_tipi)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_top_hata ON pdi_kayitlari(top_hata)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_sasi_no ON pdi_kayitlari(sasi_no)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_imalat_tarih ON imalat_kayitlari(tarih)")
    except: pass
    # ================================================
    
    conn.commit(); conn.close()

# ------------------------------------------------------
# UYGULAMA SINIFI
# ------------------------------------------------------
class PDIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Mercedes-Benz Türk | Otobüs Teknik PDI")
        self.root.geometry("1366x768")
        self.root.configure(bg=COLOR_BG)
        self.content_area = None
        
        self.current_user = None; self.current_role = None
        self.selected_month = None
        
        # ===== PERFORMANS OPTİMİZASYONU =====
        # DB Connection Cache (tek connection kullan)
        self._db_connection = None
        
        # Data Cache (aynı sorguları tekrar çalıştırma)
        self._cache = {}
        self._cache_timeout = 30  # saniye
        
        # Widget referansları (gereksiz yeniden oluşturmayı önle)
        self._widget_refs = {}
        # =====================================
        
        self.configure_styles()
        self.show_login_screen()
    
    def get_db_connection(self):
        """Cached database connection döndürür"""
        if self._db_connection is None:
            self._db_connection = sqlite3.connect(DB_NAME, check_same_thread=False)
        return self._db_connection
    
    def close_db_connection(self):
        """Uygulama kapanırken connection'ı kapat"""
        if self._db_connection:
            self._db_connection.close()
            self._db_connection = None
    
    def get_cached_data(self, cache_key):
        """Cache'den veri al (timeout kontrolü ile)"""
        if cache_key in self._cache:
            data, timestamp = self._cache[cache_key]
            if (datetime.now() - timestamp).seconds < self._cache_timeout:
                return data
        return None
    
    def set_cached_data(self, cache_key, data):
        """Cache'e veri kaydet"""
        self._cache[cache_key] = (data, datetime.now())
    
    def clear_cache(self):
        """Tüm cache'i temizle"""
        self._cache.clear()

    def configure_styles(self):
        style = ttk.Style(); style.theme_use("clam")
        style.configure("Treeview", background=COLOR_WHITE, fieldbackground=COLOR_WHITE, rowheight=35, font=FONT_NORMAL, foreground=COLOR_TEXT_MAIN, borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#E9ECEF", foreground="#495057", relief="flat")
        style.map("Treeview", background=[('selected', '#e8f0fe')], foreground=[('selected', 'black')])
        style.configure("Card.TFrame", background=COLOR_WHITE, relief="flat")

    def get_logo_image(self, size=(40, 40)):
        if not LOGO_DATA: return None
        try:
            img_data = base64.b64decode(LOGO_DATA)
            image = tk.PhotoImage(data=img_data)
            if size[0] < 60: image = image.subsample(2, 2) 
            return image
        except: return None

    def show_login_screen(self):
        self.clear_window(); self.root.configure(bg="#000000")
        card_frame = tk.Frame(self.root, bg=COLOR_WHITE, padx=50, pady=50)
        card_frame.place(relx=0.5, rely=0.5, anchor="center", width=450)
        self.photo_image = self.get_logo_image((80, 80))
        if self.photo_image: tk.Label(card_frame, image=self.photo_image, bg=COLOR_WHITE).pack(pady=(0, 20))
        tk.Label(card_frame, text="Mercedes-Benz", font=("Times New Roman", 24, "bold"), bg=COLOR_WHITE, fg="black").pack()
        tk.Label(card_frame, text="BUS PDI SYSTEM", font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg="#666").pack(pady=(0,30))
        tk.Label(card_frame, text="Kullanıcı Adı", bg=COLOR_WHITE, font=FONT_BOLD, fg="#333").pack(anchor="w")
        self.entry_user = ttk.Entry(card_frame, font=("Segoe UI", 12)); self.entry_user.pack(fill="x", pady=(5, 15), ipady=3)
        tk.Label(card_frame, text="Şifre", bg=COLOR_WHITE, font=FONT_BOLD, fg="#333").pack(anchor="w")
        self.entry_pass = ttk.Entry(card_frame, show="*", font=("Segoe UI", 12)); self.entry_pass.pack(fill="x", pady=(5, 25), ipady=3)
        tk.Button(card_frame, text="GİRİŞ", command=self.login, bg="black", fg="white", font=("Segoe UI", 12, "bold"), relief="flat", height=2).pack(fill="x")
        tk.Label(card_frame, text="Admin: admin / admin123", bg=COLOR_WHITE, fg="#ccc", font=("Segoe UI", 8)).pack(pady=(20,0))

    def login(self):
        u = self.entry_user.get(); p = self.entry_pass.get()
        conn = self.get_db_connection(); c = conn.cursor()
        c.execute("SELECT role FROM users WHERE username=? AND password=?", (u, p))
        row = c.fetchone()
        if row: self.current_user = u; self.current_role = "admin" if row[0] == 1 else "user"; self.show_main_layout()
        else: messagebox.showerror("Hata", "Giriş başarısız.")

    def logout(self):
        self.current_user = None
        self.current_role = None
        self.show_login_screen()

    def clear_window(self):
        """Tüm widget'ları temizle ve bellek sızıntılarını önle"""
        # Önce event binding'leri kaldır
        try:
            self.root.unbind_all("<MouseWheel>")
        except: pass
        # Widget'ları sil
        for widget in self.root.winfo_children():
            try:
                # Canvas ise çizimlerini temizle
                if isinstance(widget, tk.Canvas):
                    widget.delete("all")
                widget.destroy()
            except: pass

    def clear_content(self):
        """İçerik alanını temizle - optimize edilmiş versiyon"""
        if self.content_area is not None:
            # Varsa loading indicator'ı durdur
            if hasattr(self, '_loading_active'):
                self._loading_active = False
            for w in self.content_area.winfo_children():
                try:
                    # Canvas ise çizimlerini temizle
                    if isinstance(w, tk.Canvas):
                        w.delete("all")
                    # Recursive olarak child widget'ları temizle
                    for child in w.winfo_children():
                        if isinstance(child, tk.Canvas):
                            child.delete("all")
                        child.destroy()
                    w.destroy()
                except: pass

    def run_async(self, func, callback, *args, **kwargs):
        """İşlemi arka planda çalıştırır ve loading gösterir"""
        self.clear_content()
        loading_frame = tk.Frame(self.content_area, bg=COLOR_BG)
        loading_frame.pack(expand=True, fill="both")
        
        lbl = tk.Label(loading_frame, text="⌛ Yükleniyor...", font=("Segoe UI", 16, "bold"), bg=COLOR_BG, fg="#333")
        lbl.pack(expand=True)
        
        self._loading_active = True
        def animate_loading():
            if not hasattr(self, '_loading_active') or not self._loading_active: return
            dots = lbl.cget("text").count(".")
            new_text = "⌛ Yükleniyor" + ("." * ((dots % 3) + 1))
            lbl.config(text=new_text)
            self.root.after(500, animate_loading)
        
        animate_loading()

        def wrapper():
            try:
                result = func(*args, **kwargs)
                self.root.after(0, lambda: self._on_async_complete(callback, result))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Hata", f"İşlem sırasında hata oluştu:\n{e}"))
                self.root.after(0, self.clear_content)
            finally:
                self._loading_active = False

        threading.Thread(target=wrapper, daemon=True).start()

    def _on_async_complete(self, callback, result):
        self._loading_active = False
        self.clear_content()
        callback(result)

    # ------------------------------------------------------
    # ANA DÜZEN
    # ------------------------------------------------------
    def show_main_layout(self):
        self.clear_window(); self.root.configure(bg=COLOR_BG)
        
        # --- SIDEBAR (SOL) ---
        self.sidebar_container = tk.Frame(self.root, bg=COLOR_SIDEBAR, width=250)
        self.sidebar_container.pack(side="left", fill="y"); self.sidebar_container.pack_propagate(False)
        
        # Canvas for scrollable menu
        self.sidebar_canvas = tk.Canvas(self.sidebar_container, bg=COLOR_SIDEBAR, highlightthickness=0, width=250)
        self.sidebar_canvas.pack(side="left", fill="both", expand=True)
        
        # Scrollable Frame inside Canvas
        self.sidebar = tk.Frame(self.sidebar_canvas, bg=COLOR_SIDEBAR, width=250)
        self.sidebar_canvas.create_window((0, 0), window=self.sidebar, anchor="nw", width=250)
        
        # Scroll Logic
        def on_sidebar_configure(event):
            self.sidebar_canvas.configure(scrollregion=self.sidebar_canvas.bbox("all"))
        self.sidebar.bind("<Configure>", on_sidebar_configure)
        
        def on_sidebar_mousewheel(event):
            self.sidebar_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        # Bind mousewheel when hovering sidebar
        self.sidebar_canvas.bind("<Enter>", lambda e: self.sidebar_canvas.bind_all("<MouseWheel>", on_sidebar_mousewheel))
        self.sidebar_canvas.bind("<Leave>", lambda e: self.sidebar_canvas.unbind_all("<MouseWheel>"))

        # Logo
        logo_box = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR, height=80); logo_box.pack(fill="x", pady=20, padx=20)
        try:
             logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Mercedes_Benz_Logo.png")
             if os.path.exists(logo_path):
                 pil_img = Image.open(logo_path).resize((180, 60), Image.Resampling.LANCZOS)
                 self.logo_img = ImageTk.PhotoImage(pil_img)
                 tk.Label(logo_box, image=self.logo_img, bg=COLOR_SIDEBAR).pack(pady=0)
             else:
                 tk.Label(logo_box, text="Mercedes-Benz", fg="white", bg=COLOR_SIDEBAR, font=("Times New Roman", 14, "bold")).pack(anchor="w")
                 tk.Label(logo_box, text="BUS PDI SYSTEM", fg="#888", bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
        except:
             tk.Label(logo_box, text="Mercedes-Benz", fg="white", bg=COLOR_SIDEBAR, font=("Times New Roman", 14, "bold")).pack(anchor="w")
             tk.Label(logo_box, text="BUS PDI SYSTEM", fg="#888", bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
             
        tk.Frame(self.sidebar, bg="#333", height=1).pack(fill="x")

        # Kullanıcı Bilgisi
        user_box = tk.Frame(self.sidebar, bg=COLOR_SIDEBAR, pady=20, padx=20); user_box.pack(fill="x")
        canvas_avatar = tk.Canvas(user_box, width=50, height=50, bg=COLOR_SIDEBAR, highlightthickness=0); canvas_avatar.pack(side="left")
        canvas_avatar.create_oval(2,2,48,48, fill="#f0f0f0", outline=""); canvas_avatar.create_text(25,25, text=self.current_user[0].upper(), font=("Segoe UI", 16, "bold"), fill="#333")
        u_info = tk.Frame(user_box, bg=COLOR_SIDEBAR); u_info.pack(side="left", padx=10)
        tk.Label(u_info, text=self.current_user.capitalize(), fg="white", bg=COLOR_SIDEBAR, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        role_txt = "YÖNETİCİ" if self.current_role == "admin" else "KULLANICI"; role_color = COLOR_GREEN if self.current_role == "admin" else "#17a2b8"
        tk.Label(u_info, text=role_txt, fg=role_color, bg=COLOR_SIDEBAR, font=("Segoe UI", 8)).pack(anchor="w")
        tk.Frame(self.sidebar, bg="#333", height=1).pack(fill="x")
        
        # --- MENU BUTTONS ---
        self.create_sidebar_btn("⊞  ÖZET PANEL", self.show_dashboard)
        
        tk.Label(self.sidebar, text="RAPORLAR", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
        self.create_sidebar_btn("📊  TRV & TOU RAPOR", self.show_trv_tou_report)
        self.create_sidebar_btn("📊  CONNECTO RAPOR", self.show_connecto_report)
        self.create_sidebar_btn("📉  TOP 5 ANALİZ", self.show_top5_analysis)
        self.create_sidebar_btn("📉  CONN. TOP 3", self.show_connecto_top3)
        
        tk.Label(self.sidebar, text="KAYIT İŞLEMLERİ", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
        self.create_sidebar_btn("☰  KAYIT LİSTESİ", self.show_list_view)
        self.create_sidebar_btn("⊕  YENİ KAYIT", self.open_add_window)
        
        if self.current_role == "admin": 
            tk.Label(self.sidebar, text="İMALAT TAKİP", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
            self.create_sidebar_btn("🛠️  İMALATA GÖNDER", self.show_imalat_form)
            self.create_sidebar_btn("📋  İMALAT LİSTESİ", self.show_imalat_list)
            self.create_sidebar_btn("📉  İMALAT RAPORLARI", self.show_imalat_reports)
            
            tk.Label(self.sidebar, text="YÖNETİM", fg="#666", bg=COLOR_SIDEBAR, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=20, pady=(15, 5))
            self.create_sidebar_btn("👤  KULLANICILAR", self.open_user_management)
            self.create_sidebar_btn("🏷️  TOP HATA YÖNETİMİ", self.open_top_hata_management)
            self.create_sidebar_btn("📥  GEÇMİŞ VERİ YÜKLE", self.open_excel_import)
            
        tk.Frame(self.sidebar, bg=COLOR_SIDEBAR, height=50).pack(fill="x")
        tk.Button(self.sidebar, text="↪ OTURUMU KAPAT", command=self.logout, bg="#1a1a1a", fg="#aaa", relief="flat", font=("Segoe UI", 9, "bold"), anchor="w", padx=20, pady=15, bd=0).pack(fill="x")

        # --- SAĞ PANEL (HEADER + CONTENT) ---
        right_panel = tk.Frame(self.root, bg=COLOR_BG); right_panel.pack(side="right", fill="both", expand=True)
        
        # Header
        self.header = tk.Frame(right_panel, bg=COLOR_WHITE, height=60, padx=30); self.header.pack(fill="x"); self.header.pack_propagate(False)
        tk.Label(self.header, text="MERKEZ ATÖLYE", font=("Segoe UI", 10, "bold"), fg="#999", bg=COLOR_WHITE).pack(side="left")
        self.lbl_date = tk.Label(self.header, text="", font=("Segoe UI", 10, "bold"), fg="#333", bg=COLOR_WHITE); self.lbl_date.pack(side="right")
        tk.Label(self.header, text="MBT Genel Merkez\nISTANBUL, TR", font=("Segoe UI", 8), fg="#999", bg=COLOR_WHITE, justify="right").pack(side="right", padx=20)
        
        self.update_clock()
        
        # Content Area
        self.content_area = tk.Frame(right_panel, bg=COLOR_BG)
        self.content_area.pack(fill="both", expand=True)

        # Varsayılan Sayfa
        self.show_dashboard()

    def create_sidebar_btn(self, text, command):
        btn = tk.Button(self.sidebar, text=text, command=command, bg=COLOR_SIDEBAR, fg="#ccc", font=("Segoe UI", 10, "bold"), relief="flat", activebackground=COLOR_SIDEBAR_ACT, activeforeground="white", anchor="w", padx=20, pady=12, bd=0)
        btn.pack(fill="x", pady=1); return btn

    def update_clock(self):
        now = datetime.now(); day_name = TURKISH_DAYS[now.weekday()]; month_name = TURKISH_MONTHS[now.month]
        date_str = f"{now.day} {month_name} {now.year} {day_name}"
        if self.lbl_date.winfo_exists(): self.lbl_date.config(text=date_str); self.root.after(60000, self.update_clock)

    # ------------------------------------------------------
    # HOME SUMMARY (YENİ ÖZET EKRANI)
    # ------------------------------------------------------
    def show_dashboard(self):
        self.run_async(self._load_dashboard_data, self._render_dashboard)

    def _load_dashboard_data(self):
        # Filtreye göre sorgu hazırla
        filter_val = self.dashboard_month_filter.get() if hasattr(self, 'dashboard_month_filter') else "TÜMÜ"
        
        # Cache kontrolü
        cache_key = f"dashboard_{filter_val}"
        cached = self.get_cached_data(cache_key)
        if cached: return cached

        conn = self.get_db_connection()
        query = "SELECT sasi_no, arac_tipi, tarih_saat FROM pdi_kayitlari"
        df = pd.read_sql_query(query, conn)
        df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
        df = df.dropna(subset=['dt'])

        if filter_val != "TÜMÜ":
            try:
                parts = filter_val.split()
                m_name = parts[0]
                year = int(parts[1])
                m_idx = next((k for k, v in TURKISH_MONTHS.items() if v == m_name), 0)
                df = df[(df['dt'].dt.month == m_idx) & (df['dt'].dt.year == year)]
            except: pass

        total_veh = df['sasi_no'].nunique()
        total_err = len(df)
        c_tou = df[df['arac_tipi'] == 'Tourismo']['sasi_no'].nunique()
        c_trv = df[df['arac_tipi'] == 'Travego']['sasi_no'].nunique()
        c_con = df[df['arac_tipi'] == 'Connecto']['sasi_no'].nunique()
        
        result = (total_veh, total_err, c_tou, c_trv, c_con, filter_val)
        self.set_cached_data(cache_key, result)
        return result

    def _render_dashboard(self, data):
        total_veh, total_err, c_tou, c_trv, c_con, filter_val = data
        self.clear_content()
        
        # Header Area
        header = tk.Frame(self.content_area, bg=COLOR_BG)
        header.pack(fill="x", padx=30, pady=30)
        
        tk.Label(header, text="GENEL ÖZET", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(side="left")
        
        # Filtreleme Alanı
        filter_frame = tk.Frame(header, bg=COLOR_BG)
        filter_frame.pack(side="right")
        
        tk.Label(filter_frame, text="Zaman Filtresi: ", bg=COLOR_BG, font=FONT_BOLD).pack(side="left")
        
        if not hasattr(self, 'dashboard_month_filter'):
            self.dashboard_month_filter = tk.StringVar(value="TÜMÜ")
            
        # Ay listesi oluştur
        now = datetime.now()
        months = ["TÜMÜ"]
        m_list = []
        for i in range(14):
            m_idx = (now.month - i - 1) % 12 + 1
            y_offset = (now.month - i - 1) // 12
            y = now.year + y_offset
            m_list.append(f"{TURKISH_MONTHS[m_idx]} {y}")
        months.extend(m_list)
        
        cb = ttk.Combobox(filter_frame, textvariable=self.dashboard_month_filter, values=months, state="readonly", width=15)
        cb.pack(side="left", padx=5)
        cb.set(filter_val)
        cb.bind("<<ComboboxSelected>>", lambda e: self.show_dashboard())
        
        # Refresh Button
        tk.Button(filter_frame, text=" 🔄 YENİLE ", command=self.clear_cache_and_refresh_dashboard, bg="#eee", relief="flat", font=("Segoe UI", 9, "bold")).pack(side="left", padx=5)

        # Kartlar
        cards_frame = tk.Frame(self.content_area, bg=COLOR_BG)
        cards_frame.pack(fill="x", padx=30)
        
        self.create_info_card(cards_frame, "TOPLAM ARAÇ", str(total_veh), "🚌", 0)
        self.create_info_card(cards_frame, "TOPLAM HATA", str(total_err), "📋", 1)
        self.create_info_card(cards_frame, "TOURISMO", str(c_tou), "", 2)
        self.create_info_card(cards_frame, "TRAVEGO", str(c_trv), "", 3)
        self.create_info_card(cards_frame, "CONNECTO", str(c_con), "", 4)
        
        # Filtre bilgisi
        info_txt = "Tüm zamanların verileri gösteriliyor." if filter_val == "TÜMÜ" else f"{filter_val} ayına ait veriler gösteriliyor."
        welcome = tk.Label(self.content_area, text=info_txt, font=("Segoe UI", 12), bg=COLOR_BG, fg="#666")
        welcome.pack(pady=50)

    def clear_cache_and_refresh_dashboard(self):
        self.clear_cache()
        self.show_dashboard()

    # ------------------------------------------------------
    # RAPOR SAYFALARI (FULL SCREEN)
    # ------------------------------------------------------
    
    # Yardımcı: Ay seçici başlığı
    def create_report_header(self, title, command_callback, show_filter=True):
        header = tk.Frame(self.content_area, bg=COLOR_BG)
        header.pack(fill="x", padx=20, pady=20)
        
        tk.Label(header, text=title, font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(side="left")
        
        if not show_filter: return header
        
        ctrl_frame = tk.Frame(header, bg=COLOR_BG)
        ctrl_frame.pack(side="right")
        
        tk.Label(ctrl_frame, text="Rapor Ayı: ", bg=COLOR_BG, font=("Segoe UI", 10, "bold")).pack(side="left")
        
        # Son 12 ay + gelecek ay seçenekleri
        months = []
        now = datetime.now()
        for i in range(13):
            d = now - timedelta(days=30 * (12-i)) # yaklaşık
            # daha düzgün tarih listesi
            m_idx = (now.month - (12-i) - 1) % 12 + 1
            y_offset = (now.month - (12-i) - 1) // 12
            y = now.year + y_offset
            months.append(f"{TURKISH_MONTHS[m_idx]} {y}")
            
        # Gelecek ay da olsun (test için)
        next_m = (now.month % 12) + 1
        next_y = now.year + (1 if now.month == 12 else 0)
        months.append(f"{TURKISH_MONTHS[next_m]} {next_y}")
        
        # Tersten sırala (en yeni en üstte)
        months.reverse()
        months.insert(0, "TÜM ZAMANLAR")
        
        if self.selected_month is None:
            self.selected_month = tk.StringVar(value=f"{TURKISH_MONTHS[now.month]} {now.year}")
        
        cb = ttk.Combobox(ctrl_frame, textvariable=self.selected_month, values=months, state="readonly", width=15)
        cb.pack(side="left")
        cb.bind("<<ComboboxSelected>>", lambda e: self.root.after(10, command_callback))
        
        return header

    def show_trv_tou_report(self):
        try:
            month, year = self.get_selected_month_year()
            if month is None or year is None:
                self.clear_content()
                tk.Label(self.content_area, text="Lütfen rapor dönemini seçiniz.", bg=COLOR_BG).pack(pady=50)
                return

            self.run_async(
                self.calculate_monthly_data,
                self._render_trv_tou_report,
                ["Tourismo", "Travego"], month, year
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Rapor yüklenemedi:\n{str(e)}")

    def _render_trv_tou_report(self, data):
        self.clear_content()
        self.create_report_header("TRV & TOU RAPORU", self.show_trv_tou_report)
        
        month = data['target_month']
        year = data['target_year']
        container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Template metin (Kopyalanabilir yapıldı)
        template_text = f"{year} {TURKISH_MONTHS[month]} ayında PDI yapılan [TRV & TOU] araç sayısı {data['current_month_vehicles']} adettir. Son 12 ay ortalama hata oranı araç başı {data['avg_12_month']:.2f} adet olup, {TURKISH_MONTHS[month]} ayı özelinde {data['current_month_rate']:.2f} adet olarak gerçekleşmiştir."
        
        txt_box = tk.Text(container, font=("Segoe UI", 11), bg=COLOR_WHITE, fg="#333", height=2, relief="flat", highlightthickness=0)
        txt_box.insert("1.0", template_text)
        txt_box.config(state="disabled", cursor="arrow") # Yazılabilir değil ama seçilebilir
        txt_box.pack(fill="x", pady=(0, 20))
        
        # İçerik Düzeni (GRID Layout)
        content_grid = tk.Frame(container, bg=COLOR_WHITE)
        content_grid.pack(fill="both", expand=True)
        content_grid.columnconfigure(0, weight=12) # Bar grafiğe çok daha fazla alan
        content_grid.columnconfigure(1, weight=0)  # Separator
        content_grid.columnconfigure(2, weight=1)  # Donut dar alan
        content_grid.rowconfigure(0, weight=1)
        
        # SOL: Bar Grafik
        left_frame = tk.Frame(content_grid, bg=COLOR_WHITE, highlightthickness=1, highlightbackground="#eee")
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        # Zoom Butonu
        tk.Button(left_frame, text="🔍 Grafiği Tam Ekran Yap (Screenshot İçin)", command=lambda: self.zoom_chart(data['monthly_stats'], context_key="trv_tou"), 
                  bg="#f8f9fa", fg="#333", relief="flat", font=("Segoe UI", 9)).pack(anchor="ne", padx=10, pady=5)
        
        bar_canvas = tk.Canvas(left_frame, bg=COLOR_WHITE, highlightthickness=0)
        h_scroll = ttk.Scrollbar(left_frame, orient="horizontal", command=bar_canvas.xview)
        bar_canvas.configure(xscrollcommand=h_scroll.set)
        bar_canvas.pack(fill="both", expand=True, pady=10)
        h_scroll.pack(fill="x")
        
        self.draw_combo_bar_line_chart(bar_canvas, data['monthly_stats'], context_key="trv_tou")
        
        # ORTA: Ayırıcı Çizgi
        tk.Frame(content_grid, bg="#333", width=2).grid(row=0, column=1, sticky="ns", padx=10)

        # SAĞ: Donut Grafikleri (Genişlik sabitlendi ve sağa yaslandı)
        right_frame = tk.Frame(content_grid, bg=COLOR_WHITE, width=220)
        right_frame.grid(row=0, column=2, sticky="nse")
        right_frame.grid_propagate(False)
        
        # Dikey yerleşim
        self.draw_donut_section_vertical(right_frame, data, context_key="trv_tou")

    def show_connecto_report(self):
        try:
            month, year = self.get_selected_month_year()
            if month is None or year is None:
                self.clear_content()
                tk.Label(self.content_area, text="Lütfen rapor dönemini seçiniz.", bg=COLOR_BG).pack(pady=50)
                return

            self.run_async(
                self.calculate_monthly_data,
                self._render_connecto_report,
                ["Connecto"], month, year
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Connecto Rapor Hatası:\n{str(e)}")

    def _render_connecto_report(self, data):
        self.clear_content()
        self.create_report_header("CONNECTO RAPORU", self.show_connecto_report)
        
        month = data['target_month']
        year = data['target_year']
        container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        template_text = f"{year} {TURKISH_MONTHS[month]} ayında PDI yapılan [CONNECTO] araç sayısı {data['current_month_vehicles']} adettir. Son 12 ay ortalama hata oranı araç başı {data['avg_12_month']:.2f} adet olup, {TURKISH_MONTHS[month]} ayı özelinde {data['current_month_rate']:.2f} adet olarak gerçekleşmiştir."
        
        txt_box = tk.Text(container, font=("Segoe UI", 11), bg=COLOR_WHITE, fg="#333", height=2, relief="flat", highlightthickness=0)
        txt_box.insert("1.0", template_text)
        txt_box.config(state="disabled", cursor="arrow")
        txt_box.pack(fill="x", pady=(0, 20))
        
        content_grid = tk.Frame(container, bg=COLOR_WHITE)
        content_grid.pack(fill="both", expand=True)
        content_grid.columnconfigure(0, weight=12)
        content_grid.columnconfigure(1, weight=0)
        content_grid.columnconfigure(2, weight=1)
        content_grid.rowconfigure(0, weight=1)
        
        left_frame = tk.Frame(content_grid, bg=COLOR_WHITE, highlightthickness=1, highlightbackground="#eee")
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        # Zoom Butonu
        tk.Button(left_frame, text="🔍 Grafiği Tam Ekran Yap (Screenshot İçin)", command=lambda: self.zoom_chart(data['monthly_stats'], context_key="connecto"), 
                  bg="#f8f9fa", fg="#333", relief="flat", font=("Segoe UI", 9)).pack(anchor="ne", padx=10, pady=5)
        
        bar_canvas = tk.Canvas(left_frame, bg=COLOR_WHITE, highlightthickness=0)
        h_scroll = ttk.Scrollbar(left_frame, orient="horizontal", command=bar_canvas.xview)
        bar_canvas.configure(xscrollcommand=h_scroll.set)
        bar_canvas.pack(fill="both", expand=True, pady=10)
        h_scroll.pack(fill="x")
        
        self.draw_combo_bar_line_chart(bar_canvas, data['monthly_stats'], context_key="connecto")
        
        # ORTA: Ayırıcı Çizgi
        tk.Frame(content_grid, bg="#333", width=2).grid(row=0, column=1, sticky="ns", padx=10)
        
        right_frame = tk.Frame(content_grid, bg=COLOR_WHITE, width=220)
        right_frame.grid(row=0, column=2, sticky="nse")
        right_frame.grid_propagate(False)
        
        self.draw_donut_section_vertical(right_frame, data, context_key="connecto")

    def show_top5_analysis(self):
        month, year = self.get_selected_month_year()
        if month is None or year is None:
            self.clear_content()
            tk.Label(self.content_area, text="Lütfen sağ üstten rapor dönemini seçiniz.", bg=COLOR_BG).pack(pady=50)
            return

        self.run_async(
            self.get_top5_hata_analysis,
            lambda result: self._render_top5_analysis(result, month, year),
            ["Tourismo", "Travego"], month, year
        )

    def _render_top5_analysis(self, top5_data, month, year):
        self.clear_content()
        self.create_report_header("TÜM TOP HATALARIN ANALİZİ", self.show_top5_analysis)
        
        container = tk.Frame(self.content_area, bg=COLOR_BG)
        container.pack(fill="both", expand=True, padx=20, pady=10)
        def edit_top5_table():
            win = tk.Toplevel(self.root)
            win.title("Top Hata Analizi Düzenle")
            win.geometry("800x600")
            
            main_f = tk.Frame(win, padx=20, pady=20)
            main_f.pack(fill="both", expand=True)
            
            entries = []
            canvas_edit = tk.Canvas(main_f)
            canvas_edit.pack(side="left", fill="both", expand=True)
            vsb_edit = ttk.Scrollbar(main_f, orient="vertical", command=canvas_edit.yview)
            vsb_edit.pack(side="right", fill="y")
            scroll_f = tk.Frame(canvas_edit)
            canvas_edit.create_window((0,0), window=scroll_f, anchor="nw")
            
            headers = ["Hata Tanımı", "Travego %", "Tourismo %", "Genel %", "Durum"]
            for i, h in enumerate(headers): tk.Label(scroll_f, text=h, font=FONT_BOLD).grid(row=0, column=i, padx=5)
            
            # Show 10 rows for manual entry
            for idx in range(10):
                r = idx + 1
                db_hata = self.get_manual_data("top5_table", f"hata_{idx}", "top5")
                db_trv = self.get_manual_data("top5_table", f"trv_{idx}", "top5")
                db_tou = self.get_manual_data("top5_table", f"tou_{idx}", "top5")
                db_avg = self.get_manual_data("top5_table", f"avg_{idx}", "top5")
                db_status = self.get_manual_data("top5_table", f"status_{idx}", "top5")
                
                # Default if no manual data but we have DB data
                def_h = top5_data[idx]['hata_adi'] if idx < len(top5_data) else ""
                def_trv = f"{top5_data[idx]['trv_rate']:.2f}" if idx < len(top5_data) else ""
                def_tou = f"{top5_data[idx]['tou_rate']:.2f}" if idx < len(top5_data) else ""
                def_avg = f"{top5_data[idx]['avg_rate']:.2f}" if idx < len(top5_data) else ""

                e1 = ttk.Entry(scroll_f, width=25); e1.insert(0, db_hata if db_hata is not None else def_h); e1.grid(row=r, column=0, pady=2)
                e2 = ttk.Entry(scroll_f, width=10); e2.insert(0, db_trv if db_trv is not None else def_trv); e2.grid(row=r, column=1)
                e3 = ttk.Entry(scroll_f, width=10); e3.insert(0, db_tou if db_tou is not None else def_tou); e3.grid(row=r, column=2)
                e4 = ttk.Entry(scroll_f, width=10); e4.insert(0, db_avg if db_avg is not None else def_avg); e4.grid(row=r, column=3)
                e5 = ttk.Entry(scroll_f, width=15); e5.insert(0, db_status if db_status is not None else "Stabil"); e5.grid(row=r, column=4)
                entries.append((e1, e2, e3, e4, e5))
                
            def save_top5():
                for i in range(10):
                    self.save_manual_data("top5_table", f"hata_{i}", "top5", entries[i][0].get())
                    self.save_manual_data("top5_table", f"trv_{i}", "top5", entries[i][1].get())
                    self.save_manual_data("top5_table", f"tou_{i}", "top5", entries[i][2].get())
                    self.save_manual_data("top5_table", f"avg_{i}", "top5", entries[i][3].get())
                    self.save_manual_data("top5_table", f"status_{i}", "top5", entries[i][4].get())
                win.destroy()
                self.show_top5_analysis()
            
            tk.Button(win, text="TABLOYU GÜNCELLE", command=save_top5, bg="black", fg="white", font=FONT_BOLD, pady=10).pack(pady=10)
            scroll_f.update_idletasks(); canvas_edit.config(scrollregion=canvas_edit.bbox("all"))

        # Tablo Başlıkları (Sabit kalmalı)
        hdr_frame = tk.Frame(container, bg="#eee")
        hdr_frame.pack(fill="x")
        
        # Column weights for header
        hdr_frame.columnconfigure(0, weight=4)
        for i in range(1, 5): hdr_frame.columnconfigure(i, weight=2)
        
        cols = [("TOP HATA TANIMI", 0), ("TRAVEGO %", 1), ("TOURISMO %", 2), ("GENEL ORTALAMA %", 3), ("GÜNCEL DURUM", 4)]
        
        for txt, idx in cols:
            f = tk.Frame(hdr_frame, bg="#eee", borderwidth=1, relief="solid")
            f.grid(row=0, column=idx, sticky="nsew")
            tk.Label(f, text=txt, font=("Segoe UI", 9, "bold"), bg="#eee", fg="#333", pady=8).pack(side="left", padx=5, expand=True)
            if idx == 4: # Add edit icon in the last column header
                tk.Button(f, text="📝", command=edit_top5_table, bg="#eee", bd=0, font=("Segoe UI", 10, "bold"), cursor="hand2").pack(side="right", padx=5)
        
        # Kaydırılabilir içerik alanı
        scroll_canvas = tk.Canvas(container, bg=COLOR_WHITE, highlightthickness=0)
        scroll_canvas.pack(side="left", fill="both", expand=True)
        
        v_scroll = ttk.Scrollbar(container, orient="vertical", command=scroll_canvas.yview)
        v_scroll.pack(side="right", fill="y")
        
        scroll_canvas.configure(yscrollcommand=v_scroll.set)
        
        grid_frame = tk.Frame(scroll_canvas, bg=COLOR_WHITE)
        scroll_canvas.create_window((0, 0), window=grid_frame, anchor="nw", tags="frame")
        
        def on_frame_configure(e):
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
            # Grid frame genişliğini canvas genişliğine eşitle
            scroll_canvas.itemconfig("frame", width=scroll_canvas.winfo_width())

        grid_frame.bind("<Configure>", on_frame_configure)
        scroll_canvas.bind("<Configure>", lambda e: scroll_canvas.itemconfig("frame", width=e.width))

        grid_frame.columnconfigure(0, weight=4)
        for i in range(1, 5): grid_frame.columnconfigure(i, weight=2)
        
        for idx in range(10):
            # Başlık-Değer ikilisi için DB kontrolü
            db_hata = self.get_manual_data("top5_table", f"hata_{idx}", "top5")
            db_trv = self.get_manual_data("top5_table", f"trv_{idx}", "top5")
            db_tou = self.get_manual_data("top5_table", f"tou_{idx}", "top5")
            db_avg = self.get_manual_data("top5_table", f"avg_{idx}", "top5")
            db_status = self.get_manual_data("top5_table", f"status_{idx}", "top5")

            if db_hata is None and idx >= len(top5_data): continue
            
            row_idx = idx + 1
            bg_col = "#fff" if idx % 2 == 0 else "#f9f9f9"
            
            def create_cell(text, col_idx, fg="black", font=FONT_NORMAL, link_hata=None):
                lbl = tk.Label(grid_frame, text=text, bg=bg_col, fg=fg, font=font, pady=8, borderwidth=0, relief="flat", cursor="hand2")
                lbl.grid(row=row_idx, column=col_idx, sticky="nsew", padx=1)
                if link_hata:
                    lbl.bind("<Button-1>", lambda e, h=link_hata: self.show_top_hata_detail(h, ["Tourismo", "Travego"]))
                return lbl
            
            item = top5_data[idx] if idx < len(top5_data) else None
            hata_text = db_hata if db_hata is not None else (item['hata_adi'] if item else "")
            if not hata_text and db_hata is None: continue # Skip empty rows

            trv_text = f"%{db_trv}" if db_trv is not None else (f"%{item['trv_rate']:.2f}" if item else "%0.00")
            tou_text = f"%{db_tou}" if db_tou is not None else (f"%{item['tou_rate']:.2f}" if item else "%0.00")
            avg_text = f"%{db_avg}" if db_avg is not None else (f"%{item['avg_rate']:.2f}" if item else "%0.00")
            
            create_cell(hata_text, 0, font=("Segoe UI", 10, "bold"), link_hata=hata_text)
            create_cell(trv_text, 1)
            create_cell(tou_text, 2)
            create_cell(avg_text, 3)
            
            # Durum ikonu/rengi
            status_color = "black"
            status_text = "Stabil"
            if item:
                diff = item['diff']
                if diff > 25: status_color, status_text = "#d32f2f", "▲ Kritik Artış"
                elif diff > 0: status_color, status_text = "#f57c00", "▲ Artış"
                elif diff < 0: status_color, status_text = "#388e3c", "▼ Düşüş"
            
            if db_status is not None: status_text = db_status
            
            create_cell(status_text, 4, fg=status_color, font=("Segoe UI", 9, "bold"))

    def show_connecto_top3(self):
        try:
            self.clear_content()
            self.create_report_header("CONNECTO TOP 3 ANALİZİ", self.show_connecto_top3)
            
            month, year = self.get_selected_month_year()
            self.run_async(
                lambda: self.get_connecto_top3(month, year),
                lambda data_res: self._render_connecto_top3(data_res, month, year)
            )
        except Exception as e:
            messagebox.showerror("Hata", f"Connecto Top 3 Raporu Hatası:\n{str(e)}")

    def _render_connecto_top3(self, data_res, month, year):
        data = data_res['results']
        total_err = data_res['total_errors']
        
        container = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=20, pady=20)
        container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        if month is None:
            title = "TÜM ZAMANLAR Connecto Top 3 Hata Analizi"
            period_label = "TÜM"
        else:
            title = f"{year} yılı {TURKISH_MONTHS.get(month, '')} ayı için Connecto Top 3 Hataları"
            period_label = f"YTD{month}"
        
        tk.Label(container, text=title, font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE).pack(pady=(0, 20))
        
        # 3x3 Tablo (Header + 3 satır)
        table_frame = tk.Frame(container, bg=COLOR_WHITE, highlightbackground="#ccc", highlightthickness=1)
        table_frame.pack(fill="x", padx=50)
        
        headers = ["TOP HATA", f"{period_label} Hata Sayısı\n({period_label} Toplam Hata: {total_err} adet)", f"{period_label} Toplam Oran"]
        for i, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=("Segoe UI", 10, "bold"), bg="#eee", pady=10).grid(row=0, column=i, sticky="nsew")
        
        table_frame.columnconfigure(0, weight=3)
        table_frame.columnconfigure(1, weight=2)
        table_frame.columnconfigure(2, weight=2)
        
        if not data:
             tk.Label(table_frame, text="Veri Yok", bg=COLOR_WHITE).grid(row=1, column=0, columnspan=3, pady=20)
        else:
            for idx, item in enumerate(data):
                r = idx + 1
                h_name = item.get('hata_adi') or item.get('hata', 'N/A')
                h_count = item.get('ytd_sayi') or item.get('count', 0)
                h_rate = item.get('oran') or item.get('rate', 0)
                
                tk.Label(table_frame, text=h_name, font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=0, sticky="nsew")
                tk.Label(table_frame, text=str(h_count), font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=1, sticky="nsew")
                tk.Label(table_frame, text=f"%{(h_rate*100):.2f}", font=FONT_NORMAL, bg=COLOR_WHITE, pady=10).grid(row=r, column=2, sticky="nsew")
    
    def get_selected_month_year(self):
        try:
            sel = self.selected_month.get().strip()
            if sel == "TÜM ZAMANLAR": return None, None
            parts = sel.split()
            if len(parts) < 2: return datetime.now().month, datetime.now().year
            
            month_name = parts[0].upper()
            year = int(parts[1])
            month = next((k for k, v in TURKISH_MONTHS.items() if v == month_name), 1)
            return month, year
        except:
            now = datetime.now()
            return now.month, now.year
    
    # HELPER: Aylık veri hesaplama - OBTİMİZE EDİLMİŞ (PANDAS)
    def calculate_monthly_data(self, arac_tipleri, month, year):
        conn = self.get_db_connection()
        
        # Tüm veriyi bir kerede çek (arac_tipi filtresi ile)
        placeholders = ",".join(["?" for _ in arac_tipleri])
        query = f"SELECT sasi_no, arac_tipi, tarih_saat FROM pdi_kayitlari WHERE arac_tipi IN ({placeholders})"
        df = pd.read_sql_query(query, conn, params=arac_tipleri)
        
        # Tarih parse (DD-MM-YYYY formatında olduğu varsayılıyor)
        df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
        df = df.dropna(subset=['dt'])
        
        monthly_stats = []
        for i in range(12):
            m = month - i
            y = year
            while m <= 0: m += 12; y -= 1
            
            # Pandas ile filtrele
            mask = (df['dt'].dt.month == m) & (df['dt'].dt.year == y)
            m_df = df[mask]
            
            arac_sayisi = m_df['sasi_no'].nunique()
            hata_sayisi = len(m_df)
            hata_orani = hata_sayisi / arac_sayisi if arac_sayisi > 0 else 0
            
            monthly_stats.insert(0, {
                'ay': TURKISH_MONTHS[m][:3],
                'ay_full': TURKISH_MONTHS[m],
                'yil': y,
                'arac_sayisi': arac_sayisi,
                'hata_sayisi': hata_sayisi,
                'hata_orani': hata_orani
            })
        
        # Hesaplamalar
        current = monthly_stats[-1] if monthly_stats else {'arac_sayisi': 0, 'hata_sayisi': 0, 'hata_orani': 0}
        total_vehicles = sum(m['arac_sayisi'] for m in monthly_stats)
        total_errors = sum(m['hata_sayisi'] for m in monthly_stats)
        avg_12 = total_errors / total_vehicles if total_vehicles > 0 else 0
        
        # Geçen yıl ortalaması
        ly_mask = df['dt'].dt.year == (year - 1)
        ly_df = df[ly_mask]
        ly_vehicles = ly_df['sasi_no'].nunique()
        ly_errors = len(ly_df)
        avg_last_year = ly_errors / ly_vehicles if ly_vehicles > 0 else 0
        
        # Mevcut yıl
        cy_mask = df['dt'].dt.year == year
        cy_df = df[cy_mask]
        cy_vehicles = cy_df['sasi_no'].nunique()
        cy_errors = len(cy_df)
        avg_current_year = cy_errors / cy_vehicles if cy_vehicles > 0 else 0
        
        return {
            'monthly_stats': monthly_stats,
            'current_month_vehicles': current['arac_sayisi'],
            'current_month_rate': current['hata_orani'],
            'avg_12_month': avg_12,
            'target_month': month,
            'target_year': year,
            
            'prev_year': year - 1,
            'prev_year_rate': avg_last_year,
            'prev_year_cnt': ly_vehicles,
            'prev_year_err': ly_errors,
            
            'last_12_cnt': total_vehicles,
            'last_12_err': total_errors,
            
            'current_year': year,
            'current_year_rate': avg_current_year,
            'curr_year_cnt': cy_vehicles,
            'curr_year_err': cy_errors
        }
    
    # HELPER: Top Hata analizi (PANDAS OBTİMİZE EDİLMİŞ)
    def get_top5_hata_analysis(self, arac_tipleri, month, year):
        conn = self.get_db_connection()
        top_hatalar_rows = conn.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()
        top_hatalar = [r[0] for r in top_hatalar_rows]
        
        # Tüm PDI kayıtlarını çek
        query = "SELECT sasi_no, arac_tipi, top_hata, tarih_saat FROM pdi_kayitlari"
        df = pd.read_sql_query(query, conn)
        df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
        df = df.dropna(subset=['dt'])

        # Tarih ve Araç Tipi Filtreleri
        if month is not None and year is not None:
            # Mevcut ay
            curr_mask = (df['dt'].dt.month == month) & (df['dt'].dt.year == year)
            curr_df = df[curr_mask]
            
            # Önceki ay
            prev_m = (month - 2) % 12 + 1
            prev_y = year - (1 if month == 1 else 0)
            prev_mask = (df['dt'].dt.month == prev_m) & (df['dt'].dt.year == prev_y)
            prev_df = df[prev_mask]
        else:
            curr_df = df
            prev_df = pd.DataFrame()

        def calculate_rates(data_df):
            if data_df.empty: return {}
            
            # Toplam araç sayıları
            trv_total = data_df[data_df['arac_tipi'] == 'Travego']['sasi_no'].nunique()
            tou_total = data_df[data_df['arac_tipi'] == 'Tourismo']['sasi_no'].nunique()
            all_total = trv_total + tou_total
            
            rates = {}
            for hata in top_hatalar:
                trv_count = len(data_df[(data_df['arac_tipi'] == 'Travego') & (data_df['top_hata'] == hata)])
                tou_count = len(data_df[(data_df['arac_tipi'] == 'Tourismo') & (data_df['top_hata'] == hata)])
                all_count = trv_count + tou_count
                
                trv_rate = (trv_count / trv_total * 100.0) if trv_total > 0 else 0
                tou_rate = (tou_count / tou_total * 100.0) if tou_total > 0 else 0
                avg_rate = (all_count / all_total * 100.0) if all_total > 0 else 0
                
                rates[hata] = {'trv_rate': trv_rate, 'tou_rate': tou_rate, 'avg_rate': avg_rate}
            return rates

        curr_rates = calculate_rates(curr_df)
        prev_rates = calculate_rates(prev_df)

        results = []
        for hata in top_hatalar:
            rates = curr_rates.get(hata, {'trv_rate': 0, 'tou_rate': 0, 'avg_rate': 0})
            prev_avg = prev_rates.get(hata, {}).get('avg_rate', 0)
            
            results.append({
                'hata_adi': hata,
                'trv_rate': rates['trv_rate'],
                'tou_rate': rates['tou_rate'],
                'avg_rate': rates['avg_rate'],
                'diff': rates['avg_rate'] - prev_avg
            })
        
        # Orana göre sırala
        results.sort(key=lambda x: x['avg_rate'], reverse=True)
        return results
    
    # HELPER: Connecto Top 3
    def get_connecto_top3(self, month, year):
        conn = self.get_db_connection()
        
        # OBTİMİZE EDİLMİŞ (PANDAS)
        query = "SELECT top_hata, tarih_saat FROM pdi_kayitlari WHERE arac_tipi='Connecto'"
        df = pd.read_sql_query(query, conn)
        df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
        df = df.dropna(subset=['dt'])

        if month is not None and year is not None:
             # YTD (Seçilen yıla göre ve seçilen aya kadar)
             df = df[(df['dt'].dt.year == year) & (df['dt'].dt.month <= month)]

        total_errors = len(df)
        
        # Top 3 hata
        top_df = df[df['top_hata'].notna() & (df['top_hata'] != '')]
        top_hatalar = top_df.groupby('top_hata').size().sort_values(ascending=False).head(3).reset_index().values.tolist()
        
        results = []
        for hata, cnt in top_hatalar:
            oran = cnt / total_errors if total_errors > 0 else 0
            results.append({'hata_adi': hata, 'ytd_sayi': cnt, 'oran': oran})
        
        return {'results': results, 'total_errors': total_errors}
    
    # HELPER: Top Hata Detay Modalı
    def draw_donut_section_vertical(self, parent_frame, data, context_key="general"):
        # 3 tane dikey donut (Geçen Yıl, Son 12 Ay, Mevcut Yıl)
        
        titles = [
            f"{data['prev_year']} Yılı AVG", 
            "Son 12 Ay AVG", 
            f"{data['current_year']} Yılı AVG"
        ]
        values = [data['prev_year_rate'], data['avg_12_month'], data['current_year_rate']]
        counts = [f"Araç: {data['prev_year_cnt']} | Hata: {data['prev_year_err']}",
                  f"Araç: {data['last_12_cnt']} | Hata: {data['last_12_err']}",
                  f"Araç: {data['curr_year_cnt']} | Hata: {data['curr_year_err']}"]
        
        # Grid ile yerleştir (tek sütun, 3 satır)
        parent_frame.columnconfigure(0, weight=1)
        parent_frame.rowconfigure(0, weight=1)
        parent_frame.rowconfigure(1, weight=1)
        parent_frame.rowconfigure(2, weight=1)
        
        for i in range(3):
            f = tk.Frame(parent_frame, bg=COLOR_WHITE)
            f.grid(row=i, column=0, sticky="nsew", pady=5)
            
            w = 210
            h = 140
            canvas = tk.Canvas(f, bg=COLOR_WHITE, highlightthickness=0, height=h, width=w)
            canvas.pack(fill="both", expand=True)
            
            # Başlık-Değer ikilisi için DB kontrolü (context_key ile ayrıştır)
            db_val = self.get_manual_data("donuts", f"{context_key}_value_{i}", titles[i])
            val = float(db_val) if db_val and db_val.strip() else values[i]
            
            db_det = self.get_manual_data("donuts", f"{context_key}_detail_{i}", titles[i])
            det_text = db_det if db_det and db_det.strip() else counts[i]
            
            x, y = w/2, h/2
            r = 45 # Yarıçap küçültüldü
            
            # Donut çizimi - halka şeklinde
            canvas.create_oval(x-r, y-r, x+r, y+r, outline="#eee", width=12)
            extent = min((val / 4.0) * 360, 359.9)  # Max 4.0 oranına kadar
            color = "#5c8a8a" if i==0 else "#4682b4" if i==1 else "#1a5276"
            
            if val > 0:
                canvas.create_arc(x-r, y-r, x+r, y+r, start=90, extent=-extent, outline=color, width=12, style="arc")
            
            # İç yazı - ortada
            title_parts = titles[i].split(" AVG")
            canvas.create_text(x, y-18, text=title_parts[0], font=("Segoe UI", 10, "bold"), fill="#333")
            canvas.create_text(x, y, text="AVG", font=("Segoe UI", 8), fill="#666")
            canvas.create_text(x, y+20, text=f"{val:.2f}".replace('.', ','), font=("Segoe UI", 18, "bold"), fill=color)
            
            # Altına detay yazısı (Araç ve Hata sayıları)
            tk.Label(f, text=det_text, font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#777").pack(pady=(0, 5))
            
        # Düzenleme Butonu
        def edit_donuts():
            win = tk.Toplevel(self.root)
            win.title("Grafik Verilerini Düzenle")
            win.geometry("450x500")
            
            entries = []
            for i in range(3):
                tk.Label(win, text=titles[i], font=FONT_BOLD).pack(pady=(15, 5))
                fr = tk.Frame(win)
                fr.pack(pady=5)
                
                # Kaydedilmiş değerleri al (context_key ile)
                db_det_edit = self.get_manual_data("donuts", f"{context_key}_detail_{i}", titles[i])
                current_det = db_det_edit if db_det_edit and db_det_edit.strip() else counts[i]
                
                # Detaydan Araç ve Hata sayılarını parse et
                try:
                    parts = current_det.replace("Araç:", "").replace("Hata:", "").split("|")
                    arac_val = int(parts[0].strip())
                    hata_val = int(parts[1].strip())
                except:
                    arac_val = 0
                    hata_val = 0
                
                # Araç girişi
                tk.Label(fr, text="Araç Sayısı:").grid(row=0, column=0, padx=5, sticky="e")
                e_arac = ttk.Entry(fr, width=10)
                e_arac.insert(0, str(arac_val))
                e_arac.grid(row=0, column=1, padx=5)
                
                # Hata girişi
                tk.Label(fr, text="Hata Sayısı:").grid(row=0, column=2, padx=5, sticky="e")
                e_hata = ttk.Entry(fr, width=10)
                e_hata.insert(0, str(hata_val))
                e_hata.grid(row=0, column=3, padx=5)
                
                entries.append((e_arac, e_hata))
            
            def save():
                for i in range(3):
                    try:
                        arac = int(entries[i][0].get())
                        hata = int(entries[i][1].get())
                        # AVG hesapla (hata / araç)
                        avg = (hata / arac) if arac > 0 else 0
                        detay = f"Araç: {arac} | Hata: {hata}"
                        
                        self.save_manual_data("donuts", f"{context_key}_value_{i}", titles[i], f"{avg:.2f}")
                        self.save_manual_data("donuts", f"{context_key}_detail_{i}", titles[i], detay)
                    except ValueError:
                        pass  # Hatalı giriş varsa atla
                
                # Yeniden Çiz
                for w in parent_frame.winfo_children(): w.destroy()
                self.draw_donut_section_vertical(parent_frame, data, context_key)
                win.destroy()
                
            tk.Button(win, text="GÜNCELLE", command=save, bg="black", fg="white", font=FONT_BOLD, pady=10).pack(pady=20)
            
        tk.Button(parent_frame, text="📝 Verileri Düzenle", command=edit_donuts, bg="#eee", relief="flat", font=("Segoe UI", 8)).grid(row=3, column=0, pady=5)
    def show_top_hata_detail(self, hata_adi, arac_tipleri):
        month, year = self.get_selected_month_year()
        
        # Detay penceresini hemen oluştur ama boş/yükleniyor göster
        win = tk.Toplevel(self.root)
        win.title(f"Detay: {hata_adi}")
        win.geometry("1000x700")
        win.configure(bg=COLOR_WHITE)
        
        loading_lbl = tk.Label(win, text="⌛ Veriler Analiz Ediliyor...", font=("Segoe UI", 12), bg=COLOR_WHITE)
        loading_lbl.pack(expand=True)

        def load_data():
            conn = self.get_db_connection()
            query = "SELECT sasi_no, arac_tipi, top_hata, tarih_saat FROM pdi_kayitlari"
            df = pd.read_sql_query(query, conn)
            df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
            df = df.dropna(subset=['dt'])

            monthly_data = {'TRV': [], 'TOU': []}
            for i in range(12):
                m = month - i
                y = year
                while m <= 0: m += 12; y -= 1
                
                for tip, label in [("Travego", "TRV"), ("Tourismo", "TOU")]:
                    m_df = df[(df['arac_tipi'] == tip) & (df['dt'].dt.month == m) & (df['dt'].dt.year == y)]
                    total = m_df['sasi_no'].nunique()
                    hata_cnt = len(m_df[m_df['top_hata'] == hata_adi])
                    oran = hata_cnt / total if total > 0 else 0
                    monthly_data[label].insert(0, {'ay': TURKISH_MONTHS[m][:3], 'arac': total, 'hata': hata_cnt, 'oran': oran})
            return monthly_data

        def render(monthly_data):
            if not win.winfo_exists(): return
            loading_lbl.destroy()
            
            tk.Label(win, text=f"{hata_adi} - Son 12 Ay Analizi", font=("Segoe UI", 14, "bold"), bg=COLOR_WHITE).pack(pady=15)
            
            # Grafik
            canvas = tk.Canvas(win, bg=COLOR_WHITE, highlightthickness=0, height=250)
            canvas.pack(fill="x", padx=20)
            self.draw_dual_line_chart(canvas, monthly_data, context_key=f"trend_{hata_adi}")
            
            # Tablo
            table_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20)
            table_frame.pack(fill="both", expand=True, pady=10)
            self.draw_detail_table(table_frame, monthly_data)
            
            # Legend
            legend_frame = tk.Frame(win, bg=COLOR_WHITE, pady=10)
            legend_frame.pack(fill="x", padx=20)
            
            # Legend TRV/TOU
            def add_legend(color, text):
                f = tk.Frame(legend_frame, bg=COLOR_WHITE)
                f.pack(side="left", padx=10)
                tk.Frame(f, bg=color, width=20, height=14).pack(side="left", padx=5)
                tk.Label(f, text=text, font=("Segoe UI", 9), bg=COLOR_WHITE).pack(side="left")
            
            add_legend("#0066cc", "TRV Hata Oranı")
            add_legend("#ff8c00", "TOU Hata Oranı")

        # threading kullanıyoruz (run_async burada content_area'yı sildiği için uygun değil, win üzerinde bağımsız çalışmalı)
        def wrapper():
            try:
                res = load_data()
                self.root.after(0, lambda: render(res))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Hata", f"Detay yüklenemedi: {e}"))

        threading.Thread(target=wrapper, daemon=True).start()
    
    def zoom_chart(self, monthly_stats, context_key="general"):
        win = tk.Toplevel(self.root)
        win.title("Genişletilmiş Grafik Görünümü")
        win.state('zoomed')  # Windows'ta tam ekran
        win.configure(bg=COLOR_WHITE)
        
        container = tk.Frame(win, bg=COLOR_WHITE, padx=40, pady=30)
        container.pack(fill="both", expand=True)
        
        tk.Label(container, text="Son 12 Ay PDI Analizi (Tam Görünüm)", font=("Segoe UI", 18, "bold"), bg=COLOR_WHITE, fg="#dc3545").pack(pady=(0, 20))
        
        canvas = tk.Canvas(container, bg=COLOR_WHITE, highlightthickness=0)
        canvas.pack(fill="both", expand=True)
        
        # Pencere boyutu değiştiğinde grafiği yeniden çiz
        def redraw(event=None):
            canvas.delete("all")
            self.draw_combo_bar_line_chart(canvas, monthly_stats, context_key=context_key, fullscreen=True)
        
        canvas.bind("<Configure>", redraw)
        win.after(100, redraw)

    # GRAFİK: Bar + Line kombinasyonu
    def draw_combo_bar_line_chart(self, canvas, monthly_stats, context_key="general", fullscreen=False):
        if not monthly_stats:
            canvas.create_text(400, 140, text="Veri Bulunamadı", font=FONT_NORMAL, fill="#999")
            return
        
        # Connecto için yeşil renk paleti, diğerleri için mavi
        if context_key == "connecto":
            color_bar1 = "#90EE90"  # Açık yeşil
            color_bar2 = "#228B22"  # Koyu yeşil (Forest Green)
            color_line = "#006400"  # Çizgi için koyu yeşil
        else:
            color_bar1 = "#8FADC5"  # Açık mavi
            color_bar2 = "#3D5A73"  # Koyu mavi
            color_line = "#dc3545"  # Kırmızı
        
        # Düzenleme butonu
        def edit_combo():
            win = tk.Toplevel(self.root)
            win.title("Grafik Verilerini Düzenle")
            win.geometry("650x600")
            
            main_f = tk.Frame(win, padx=20, pady=20)
            main_f.pack(fill="both", expand=True)
            
            tk.Label(main_f, text="Değerleri düzenleyebilirsiniz. Oran otomatik hesaplanır.", font=("Segoe UI", 10), fg="#666").pack(pady=(0,10))
            
            canvas_edit = tk.Canvas(main_f); canvas_edit.pack(side="left", fill="both", expand=True)
            vsb = ttk.Scrollbar(main_f, orient="vertical", command=canvas_edit.yview); vsb.pack(side="right", fill="y"); scroll_f = tk.Frame(canvas_edit); canvas_edit.create_window((0,0), window=scroll_f, anchor="nw")
            
            # Başlıklar
            tk.Label(scroll_f, text="Ay", font=FONT_BOLD).grid(row=0, column=0, padx=5)
            tk.Label(scroll_f, text="Araç Sayısı", font=FONT_BOLD).grid(row=0, column=1, padx=5)
            tk.Label(scroll_f, text="Hata Sayısı", font=FONT_BOLD).grid(row=0, column=2, padx=5)
            tk.Label(scroll_f, text="Oran (Otomatik)", font=FONT_BOLD, fg="#dc3545").grid(row=0, column=3, padx=5)
            
            entries = []
            oran_labels = []
            
            def update_oran(idx):
                try:
                    a = int(entries[idx][0].get())
                    h = int(entries[idx][1].get())
                    oran = h / a if a > 0 else 0
                    oran_labels[idx].config(text=f"{oran:.2f}")
                except:
                    oran_labels[idx].config(text="0.00")
            
            for i, m in enumerate(monthly_stats):
                r = i + 1
                tk.Label(scroll_f, text=m['ay'], font=("Segoe UI", 10, "bold")).grid(row=r, column=0, padx=5, pady=2)
                
                db_a = self.get_manual_data("combo", f"arac_{i}", context_key)
                e_a = ttk.Entry(scroll_f, width=12); e_a.insert(0, db_a if db_a is not None else m['arac_sayisi']); e_a.grid(row=r, column=1, padx=5)
                
                db_h = self.get_manual_data("combo", f"hata_{i}", context_key)
                e_h = ttk.Entry(scroll_f, width=12); e_h.insert(0, db_h if db_h is not None else m['hata_sayisi']); e_h.grid(row=r, column=2, padx=5)
                
                # Oran hesaplanır (read-only label)
                try:
                    a_val = int(db_a) if db_a and str(db_a).strip() else m['arac_sayisi']
                    h_val = int(db_h) if db_h and str(db_h).strip() else m['hata_sayisi']
                    o_val = h_val / a_val if a_val > 0 else 0
                except:
                    o_val = m['hata_orani']
                    
                lbl_o = tk.Label(scroll_f, text=f"{o_val:.2f}", width=10, fg="#dc3545", font=("Segoe UI", 10, "bold"))
                lbl_o.grid(row=r, column=3, padx=5)
                
                entries.append((e_a, e_h))
                oran_labels.append(lbl_o)
                
                # Bind to update oran when values change
                idx = i
                e_a.bind("<KeyRelease>", lambda e, idx=idx: update_oran(idx))
                e_h.bind("<KeyRelease>", lambda e, idx=idx: update_oran(idx))
                
            def save():
                for i, (ea, eh) in enumerate(entries):
                    arac = ea.get()
                    hata = eh.get()
                    try:
                        a = int(arac)
                        h = int(hata)
                        oran = h / a if a > 0 else 0
                    except:
                        oran = 0
                    
                    self.save_manual_data("combo", f"arac_{i}", context_key, arac)
                    self.save_manual_data("combo", f"hata_{i}", context_key, hata)
                    self.save_manual_data("combo", f"oran_{i}", context_key, f"{oran:.2f}")
                win.destroy()
                messagebox.showinfo("Başarılı", "Veriler kaydedildi. Lütfen sayfayı yenileyiniz.")

            tk.Button(win, text="KAYDET VE KAPAT", command=save, bg="black", fg="white", font=FONT_BOLD, pady=10).pack(pady=10)
            scroll_f.update_idletasks(); canvas_edit.config(scrollregion=canvas_edit.bbox("all"))

        # Manual overrides for drawing
        processed_stats = []
        actual_max_count = 0
        actual_max_rate = 0
        
        for i, m in enumerate(monthly_stats):
            db_a = self.get_manual_data("combo", f"arac_{i}", context_key)
            db_h = self.get_manual_data("combo", f"hata_{i}", context_key)
            db_o = self.get_manual_data("combo", f"oran_{i}", context_key)
            
            a_val = int(db_val) if (db_val := db_a) and str(db_val).strip() else m['arac_sayisi']
            h_val = int(db_val) if (db_val := db_h) and str(db_val).strip() else m['hata_sayisi']
            o_val = float(db_val) if (db_val := db_o) and str(db_val).strip() else m['hata_orani']
            
            processed_stats.append({'ay': m['ay'], 'arac_sayisi': a_val, 'hata_sayisi': h_val, 'hata_orani': o_val})
            actual_max_count = max(actual_max_count, a_val, h_val)
            actual_max_rate = max(actual_max_rate, o_val)
            
        canvas.update_idletasks()
        h = canvas.winfo_height()
        w = canvas.winfo_width()
        if h < 100: h = 500
        if w < 100: w = 900
        
        pad_x = 60; pad_y = 70
        num_months = len(processed_stats)
        
        # Fullscreen modda bar ve gap'i dinamik hesapla
        available_width = w - 2 * pad_x - 150  # Legend için alan bırak
        if fullscreen and num_months > 0:
            total_unit = num_months * 3  # her ay için 2 bar + 1 gap
            unit_width = available_width / total_unit
            bar_width = max(int(unit_width * 0.9), 20)
            gap = max(int(unit_width * 1.2), 30)
        else:
            bar_width = 22; gap = 45
        
        chart_width = pad_x + num_months * (bar_width * 2 + gap) + pad_x + 150
        if fullscreen:
            chart_width = max(chart_width, w)  # Fullscreen'de en az ekran genişliği kadar
        canvas.config(scrollregion=(0, 0, chart_width, h))
        
        max_count = actual_max_count if actual_max_count > 0 else 1
        max_rate = actual_max_rate if actual_max_rate > 0 else 1
        
        graph_h = h - 2 * pad_y
        rate_points = []
        
        # Edit Icon
        canvas.create_text(chart_width - 50, 30, text="📝", font=("Segoe UI", 12), fill="#666", tags="edit_btn")
        canvas.tag_bind("edit_btn", "<Button-1>", lambda e: edit_combo())

        # 1. Barları ve Eksenleri Çiz
        for i, m in enumerate(processed_stats):
            x = pad_x + i * (bar_width * 2 + gap)
            
            # Araç sayısı bar
            arac_h = (m['arac_sayisi'] / max_count) * graph_h
            canvas.create_rectangle(x, h - pad_y - arac_h, x + bar_width, h - pad_y, fill=color_bar1, outline="")
            
            # Hata sayısı bar
            hata_h = (m['hata_sayisi'] / max_count) * graph_h
            canvas.create_rectangle(x + bar_width + 5, h - pad_y - hata_h, x + bar_width * 2 + 5, h - pad_y, fill=color_bar2, outline="")
            
            # Ay etiketi
            canvas.create_text(x + bar_width, h - pad_y + 20, text=m['ay'], font=("Segoe UI", 9), fill="#666")
            
            # Rate çizgisi koordinatlarını topla
            rate_y = h - pad_y - (m['hata_orani'] / max_rate) * graph_h
            rate_points.append((x + bar_width, rate_y))

        # 2. Çizgiyi Çiz (Barların üstünde, yazıların altında olmalı) - Kalın çizgi
        if len(rate_points) > 1:
            canvas.create_line(rate_points, fill=color_line, width=4)

        # 3. Yazıları ve Noktaları Çiz (En üstte)
        for i, m in enumerate(processed_stats):
            x = pad_x + i * (bar_width * 2 + gap)
            arac_h = (m['arac_sayisi'] / max_count) * graph_h
            hata_h = (m['hata_sayisi'] / max_count) * graph_h
            rate_y = rate_points[i][1]
            
            # Araç Sayısı Yazısı
            self._draw_canvas_text(canvas, x + bar_width/2, h - pad_y - arac_h - 12, str(m['arac_sayisi']), font=("Segoe UI", 8), fill="#666")
            
            # Hata Sayısı Yazısı
            self._draw_canvas_text(canvas, x + bar_width * 1.5 + 5, h - pad_y - hata_h - 12, str(m['hata_sayisi']), font=("Segoe UI", 8, "bold"), fill="#333")
            
            # Rate Noktası ve Yazısı
            canvas.create_oval(x + bar_width - 5, rate_y - 5, x + bar_width + 5, rate_y + 5, fill=color_line, outline="white", width=2)
            # Eğer rate yazısı bar yazısı ile çakışıyorsa biraz daha yukarı alalım
            y_offset = -28
            if abs(rate_y + y_offset - (h - pad_y - hata_h - 12)) < 15: y_offset -= 15
            
            self._draw_canvas_text(canvas, x + bar_width, rate_y + y_offset, f"{m['hata_orani']:.2f}", font=("Segoe UI", 10, "bold"), fill=color_line)

        # Legend
        self._draw_legend(canvas, chart_width, [color_bar1, color_bar2, color_line], ["Araç Sayısı", "Hata Sayısı", "Hata Oranı"])

    def _draw_canvas_text(self, canvas, x, y, text, **kwargs):
        """Metnin arkasına küçük bir beyaz alan ekleyerek okunabilirliği artırır."""
        t = canvas.create_text(x, y, text=text, **kwargs)
        bbox = canvas.bbox(t)
        if bbox:
            # Biraz padding ekleyelim
            bg = canvas.create_rectangle(bbox[0]-2, bbox[1]-1, bbox[2]+2, bbox[3]+1, fill="white", outline="")
            canvas.tag_lower(bg, t)
        return t

    def _draw_legend(self, canvas, chart_width, colors, labels):
        y_start = 15
        for col, lab in zip(colors, labels):
            if col.startswith("#"):
                canvas.create_rectangle(chart_width - 200, y_start, chart_width - 180, y_start + 12, fill=col, outline="")
            canvas.create_text(chart_width - 170, y_start + 6, text=lab, font=("Segoe UI", 8), anchor="w")
            y_start += 20
            
    def save_manual_data(self, r_type, c_key, d_key, val):
        conn = sqlite3.connect(DB_NAME)
        conn.execute("INSERT OR REPLACE INTO report_manual_data (report_type, context_key, data_key, data_value) VALUES (?,?,?,?)", 
                     (r_type, c_key, d_key, str(val)))
        conn.commit()
        conn.close()

    def get_manual_data(self, r_type, c_key, d_key, default=None):
        conn = sqlite3.connect(DB_NAME)
        res = conn.execute("SELECT data_value FROM report_manual_data WHERE report_type=? AND context_key=? AND data_key=?", 
                           (r_type, c_key, d_key)).fetchone()
        conn.close()
        return res[0] if res else default
    
    # GRAFİK: Donut bölümü
    def draw_donut_section(self, parent, data):
        parent.pack(fill="x", pady=10)
        
        donuts = [
            (f"{datetime.now().year - 1} AVG", data['avg_last_year'], data['last_year_vehicles'], data['last_year_errors'], "#888888"),
            ("12 Ay AVG", data['avg_12_month'], data['total_12_vehicles'], data['total_12_errors'], "#4A90A4"),
            (f"{datetime.now().year} AVG", data['avg_current_year'], data['current_year_vehicles'], data['current_year_errors'], "#2E5A6B")
        ]
        
        for i, (label, rate, vehicles, errors, color) in enumerate(donuts):
            frame = tk.Frame(parent, bg=COLOR_WHITE, padx=30)
            frame.pack(side="left", expand=True)
            
            canvas = tk.Canvas(frame, width=220, height=180, bg=COLOR_WHITE, highlightthickness=0)
            canvas.pack()
            
            # Donut chart - halka şeklinde
            cx, cy = 110, 90  # Merkez koordinatları
            outer_r = 70  # Dış yarıçap
            inner_r = 45  # İç yarıçap (boşluk için)
            
            # Araç sayısı oranı (gri kısım) ve Hata sayısı oranı (koyu kısım)
            # Oranı hesapla - hata/araç oranından açı hesapla
            if vehicles > 0:
                # Hata oranına göre açı hesapla (max 360 derece)
                ratio = min(errors / vehicles, 2.0) / 2.0  # 0-1 arası normalize et
                hata_extent = -ratio * 360
                arac_extent = -(360 + hata_extent)  # Kalan kısım
            else:
                hata_extent = -180
                arac_extent = -180
            
            # Dış halka - Hata kısmı (koyu renk)
            canvas.create_arc(cx-outer_r, cy-outer_r, cx+outer_r, cy+outer_r, 
                            start=90, extent=hata_extent, fill="#555555", outline="", width=0)
            # Dış halka - Araç kısmı (gri renk)
            canvas.create_arc(cx-outer_r, cy-outer_r, cx+outer_r, cy+outer_r, 
                            start=90+hata_extent, extent=arac_extent, fill="#999999", outline="", width=0)
            
            # İç beyaz daire (ortayı boşaltmak için)
            canvas.create_oval(cx-inner_r, cy-inner_r, cx+inner_r, cy+inner_r, fill=COLOR_WHITE, outline="")
            
            # Merkez metin
            canvas.create_text(cx, cy-15, text=label, font=("Segoe UI", 10, "bold"), fill="#333")
            canvas.create_text(cx, cy+2, text="Hata Oranı;", font=("Segoe UI", 9), fill="#666")
            canvas.create_text(cx, cy+22, text=f"{rate:.2f}".replace(".", ","), font=("Segoe UI", 16, "bold"), fill="#333")
            
            # Sol taraf etiketi - Araç Sayısı
            canvas.create_text(25, cy-10, text="Araç", font=("Segoe UI", 9), fill="#666", anchor="w")
            canvas.create_text(25, cy+5, text="Sayısı", font=("Segoe UI", 9), fill="#666", anchor="w")
            canvas.create_text(25, cy+22, text=str(vehicles), font=("Segoe UI", 11, "bold"), fill="#333", anchor="w")
            
            # Sağ taraf etiketi - Hata Sayısı
            canvas.create_text(195, cy-10, text="Hata", font=("Segoe UI", 9), fill="#666", anchor="e")
            canvas.create_text(195, cy+5, text="Sayısı", font=("Segoe UI", 9), fill="#666", anchor="e")
            canvas.create_text(195, cy+22, text=str(errors), font=("Segoe UI", 11, "bold"), fill="#333", anchor="e")
    
    # GRAFİK: Dual line chart
    def draw_dual_line_chart(self, canvas, monthly_data, context_key="general"):
        h = 220; w = 960; pad_x = 60; pad_y = 40  # Yükseklik azaltıldı
        canvas.config(width=w, height=h)
        
        max_rate = 0
        for label in ['TRV', 'TOU']:
            if monthly_data[label]:
                max_rate = max(max_rate, max(m['oran'] for m in monthly_data[label]))
        max_rate = max_rate if max_rate > 0 else 1
        
        graph_h = h - 2 * pad_y
        num_months = len(monthly_data['TRV'])
        step_x = (w - 2 * pad_x) / max(num_months - 1, 1) if num_months > 1 else 0
        
        colors = {'TRV': '#0066cc', 'TOU': '#ff8c00'}
        
        # 1. Önce Çizgileri Çiz
        for label in ['TRV', 'TOU']:
            points = []
            for i, m in enumerate(monthly_data[label]):
                x = pad_x + i * step_x
                y = h - pad_y - (m['oran'] / max_rate) * graph_h
                points.append((x, y))
            if len(points) > 1:
                canvas.create_line(points, fill=colors[label], width=2)

        # 2. Sonra Noktaları ve Yazıları Çiz (Çakışma kontrolü ile)
        for i in range(num_months):
            x = pad_x + i * step_x
            
            # Alt grup TRV
            m_trv = monthly_data['TRV'][i]
            y_trv = h - pad_y - (m_trv['oran'] / max_rate) * graph_h
            
            # Alt grup TOU
            m_tou = monthly_data['TOU'][i]
            y_tou = h - pad_y - (m_tou['oran'] / max_rate) * graph_h
            
            # Ay Etiketi (Sadece bir kez) - %0 yazıları ile çakışmasın diye daha aşağıda
            canvas.create_text(x, h - 5, text=m_trv['ay'], font=("Segoe UI", 8), fill="#666")
            
            # Çakışma kontrolü: Eğer y değerleri birbirine çok yakınsa etiketleri kaydır
            # Ayrıca her ikisi de 0 ise sadece birini göster
            show_trv_label = True
            show_tou_label = True
            
            offset_trv = -22
            offset_tou = -22
            
            if abs(y_trv - y_tou) < 25:
                if m_trv['oran'] == 0 and m_tou['oran'] == 0:
                    # İkisi de sıfır, sadece birini göster
                    show_tou_label = False
                    offset_trv = -18
                elif y_trv <= y_tou:  # TRV daha yukarada
                    offset_trv = -28
                    offset_tou = 18
                else:
                    offset_trv = 18
                    offset_tou = -28
            
            # Noktalar
            canvas.create_oval(x-4, y_trv-4, x+4, y_trv+4, fill=colors['TRV'], outline="white", width=2)
            canvas.create_oval(x-4, y_tou-4, x+4, y_tou+4, fill=colors['TOU'], outline="white", width=2)
            
            # Yazılar - sadece gösterilmesi gerekenler
            if show_trv_label:
                self._draw_canvas_text(canvas, x, y_trv + offset_trv, f"{m_trv['oran']:.0%}", font=("Segoe UI", 8, "bold"), fill=colors['TRV'])
            if show_tou_label:
                self._draw_canvas_text(canvas, x, y_tou + offset_tou, f"{m_tou['oran']:.0%}", font=("Segoe UI", 8, "bold"), fill=colors['TOU'])
        
        # Legend canvas dışında çizilecek (caller tarafından)
    
    # TABLO: Detay tablosu
    def draw_detail_table(self, parent, monthly_data):
        # Düzenleme Modu için frame
        table_container = tk.Frame(parent, bg=COLOR_WHITE)
        table_container.pack(fill="both", expand=True)
        
        def render_table(data_to_show):
            for w in table_container.winfo_children(): w.destroy()
            
            header_frame = tk.Frame(table_container, bg="#4a4a4a")
            header_frame.pack(fill="x")
            lbl_title = tk.Label(header_frame, text="Son 12 Ay", font=FONT_BOLD, bg="#4a4a4a", fg="white", width=12)
            lbl_title.pack(side="left", padx=2, pady=5)
            
            # Edit Icon in Header
            tk.Button(header_frame, text="📝", command=open_edit_dialog, bg="#4a4a4a", fg="white", bd=0, font=FONT_BOLD).pack(side="right", padx=5)
            
            for m in data_to_show['TRV']:
                tk.Label(header_frame, text=m['ay'], font=FONT_BOLD, bg="#4a4a4a", fg="white", width=7).pack(side="left", padx=2, pady=5)
            
            # TRV rows
            for metric in ['arac', 'hata', 'oran']:
                row_frame = tk.Frame(table_container, bg="#e8f4f8" if metric != 'oran' else "#d4e8ed")
                row_frame.pack(fill="x")
                label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
                tk.Label(row_frame, text=f"TRV {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
                for i, m in enumerate(data_to_show['TRV']):
                    db_val = self.get_manual_data("top_hata_detail", f"TRV_{metric}_{i}", m['ay'])
                    val = db_val if db_val is not None else (m[metric] if metric != 'oran' else (f"{m['oran']:.0%}" if isinstance(m['oran'], (int, float)) else m['oran']))
                    tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)
            
            # TOU rows
            for metric in ['arac', 'hata', 'oran']:
                row_frame = tk.Frame(table_container, bg="#fff5e6" if metric != 'oran' else "#ffe8cc")
                row_frame.pack(fill="x")
                label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
                tk.Label(row_frame, text=f"TOU {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
                for i, m in enumerate(data_to_show['TOU']):
                    db_val = self.get_manual_data("top_hata_detail", f"TOU_{metric}_{i}", m['ay'])
                    val = db_val if db_val is not None else (m[metric] if metric != 'oran' else (f"{m['oran']:.0%}" if isinstance(m['oran'], (int, float)) else m['oran']))
                    tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)

        def open_edit_dialog():
            edit_win = tk.Toplevel(self.root)
            edit_win.title("Tablo Verilerini Düzenle")
            edit_win.geometry("800x600")
            
            main_scroll = tk.Canvas(edit_win)
            main_scroll.pack(side="left", fill="both", expand=True)
            vsb = ttk.Scrollbar(edit_win, orient="vertical", command=main_scroll.yview)
            vsb.pack(side="right", fill="y")
            main_scroll.configure(yscrollcommand=vsb.set)
            
            scroll_frame = tk.Frame(main_scroll)
            main_scroll.create_window((0,0), window=scroll_frame, anchor="nw")
            
            entries = {'TRV': [], 'TOU': []}
            
            for tip in ['TRV', 'TOU']:
                tk.Label(scroll_frame, text=f"{tip} VERİLERİ", font=FONT_BOLD).pack(pady=10)
                f = tk.Frame(scroll_frame)
                f.pack()
                
                # Header
                tk.Label(f, text="Ay", width=10).grid(row=0, column=0)
                tk.Label(f, text="Araç Sayısı", width=12).grid(row=0, column=1)
                tk.Label(f, text="Hata Sayısı", width=12).grid(row=0, column=2)
                tk.Label(f, text="Oran (%)", width=12).grid(row=0, column=3)
                
                for i, m in enumerate(monthly_data[tip]):
                    r = i + 1
                    tk.Label(f, text=m['ay']).grid(row=r, column=0)
                    
                    db_arac = self.get_manual_data("top_hata_detail", f"{tip}_arac_{i}", m['ay'])
                    e_arac = ttk.Entry(f, width=10); e_arac.insert(0, str(db_arac if db_arac is not None else m['arac'])); e_arac.grid(row=r, column=1)
                    
                    db_hata = self.get_manual_data("top_hata_detail", f"{tip}_hata_{i}", m['ay'])
                    e_hata = ttk.Entry(f, width=10); e_hata.insert(0, str(db_hata if db_hata is not None else m['hata'])); e_hata.grid(row=r, column=2)
                    
                    db_oran = self.get_manual_data("top_hata_detail", f"{tip}_oran_{i}", m['ay'])
                    e_oran = ttk.Entry(f, width=10); e_oran.insert(0, str(db_oran if db_oran is not None else (m['oran'] if not isinstance(m['oran'], (int, float)) else f"{m['oran']*100:.1f}"))); e_oran.grid(row=r, column=3)
                    entries[tip].append((e_arac, e_hata, e_oran))

            def save_changes():
                for tip in ['TRV', 'TOU']:
                    for i, (ea, eh, eo) in enumerate(entries[tip]):
                        try:
                            o_val = eo.get()
                            if '%' not in o_val: o_val += "%"
                            
                            self.save_manual_data("top_hata_detail", f"{tip}_arac_{i}", monthly_data[tip][i]['ay'], ea.get())
                            self.save_manual_data("top_hata_detail", f"{tip}_hata_{i}", monthly_data[tip][i]['ay'], eh.get())
                            self.save_manual_data("top_hata_detail", f"{tip}_oran_{i}", monthly_data[tip][i]['ay'], o_val)
                            
                        except Exception as e: print(f"Save error: {e}")
                
                # Re-fetch or at least inform the data structure
                # For simplicity, we just rebuild the display data
                self.show_top_hata_detail(hata_adi, arac_tipleri) # This might reload entire window, maybe just render_table?
                edit_win.destroy()

            tk.Button(scroll_frame, text="KAYDET VE GÜNCELLE", command=save_changes, bg=COLOR_GREEN, fg="white", font=FONT_BOLD, pady=10).pack(pady=20)
            
            scroll_frame.update_idletasks()
            main_scroll.config(scrollregion=main_scroll.bbox("all"))

        render_table(monthly_data)

    def draw_detail_table_old(self, parent, monthly_data):
        # Header
        header_frame = tk.Frame(parent, bg="#4a4a4a")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="Son 12 Ay", font=FONT_BOLD, bg="#4a4a4a", fg="white", width=12).pack(side="left", padx=2, pady=5)
        for m in monthly_data['TRV']:
            tk.Label(header_frame, text=m['ay'], font=FONT_BOLD, bg="#4a4a4a", fg="white", width=7).pack(side="left", padx=2, pady=5)
        
        # TRV rows
        for metric in ['arac', 'hata', 'oran']:
            row_frame = tk.Frame(parent, bg="#e8f4f8" if metric != 'oran' else "#d4e8ed")
            row_frame.pack(fill="x")
            label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
            tk.Label(row_frame, text=f"TRV {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
            for m in monthly_data['TRV']:
                val = m[metric] if metric != 'oran' else f"{m['oran']:.0%}"
                tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)
        
        # TOU rows
        for metric in ['arac', 'hata', 'oran']:
            row_frame = tk.Frame(parent, bg="#fff5e6" if metric != 'oran' else "#ffe8cc")
            row_frame.pack(fill="x")
            label = "Araç Sayısı" if metric == 'arac' else "Hata Sayısı" if metric == 'hata' else "Hata Oranı"
            tk.Label(row_frame, text=f"TOU {label}", font=FONT_NORMAL, bg=row_frame['bg'], width=12, anchor="w").pack(side="left", padx=2, pady=3)
            for m in monthly_data['TOU']:
                val = m[metric] if metric != 'oran' else f"{m['oran']:.0%}"
                tk.Label(row_frame, text=str(val), font=FONT_NORMAL, bg=row_frame['bg'], width=7).pack(side="left", padx=2, pady=3)

    def export_dashboard_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Ozet_Rapor.xlsx")
        if not file_path: return
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet oluştur ve genişlikleri ayarla
                writer.book.create_sheet('Özet Tablolar')
                ws = writer.book['Özet Tablolar']
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15

                # 1. Alt Grup
                if not self.df_alt.empty:
                    ws.cell(row=1, column=2, value="ALT GRUP ANALİZİ").font = ws.cell(row=1, column=2).font.copy(bold=True)
                    self.df_alt.to_excel(writer, sheet_name='Özet Tablolar', startrow=2, startcol=1, index=False)
                
                # 2. Top 5
                current_row = len(self.df_alt) + 5 if not self.df_alt.empty else 1
                if self.top_5_errors:
                    ws.cell(row=current_row, column=2, value="EN ÇOK GÖRÜLEN HATALAR (TOP 5)").font = ws.cell(row=current_row, column=2).font.copy(bold=True)
                    df_err = pd.DataFrame(self.top_5_errors, columns=["Hata Konumu", "Adet"])
                    df_err.to_excel(writer, sheet_name='Özet Tablolar', startrow=current_row+1, startcol=1, index=False)
                    current_row += len(df_err) + 4
                
                # 3. Aylık Trend
                if self.monthly_data:
                    ws.cell(row=current_row, column=2, value="AYLIK HATA ORANI TRENDİ").font = ws.cell(row=current_row, column=2).font.copy(bold=True)
                    df_trend = pd.DataFrame(self.monthly_data, columns=["Ay", "Araç Sayısı", "Hata Sayısı", "Oran"])
                    df_trend.to_excel(writer, sheet_name='Özet Tablolar', startrow=current_row+1, startcol=1, index=False)
            
            messagebox.showinfo("Başarılı", "Özet tablolar Excel'e aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulamadı: {str(e)}")

    def create_info_card(self, parent, title, value, icon, col):
        card = tk.Frame(parent, bg=COLOR_WHITE, padx=20, pady=20); card.grid(row=0, column=col, sticky="ew", padx=10)
        header = tk.Frame(card, bg=COLOR_WHITE); header.pack(fill="x")
        tk.Label(header, text=title, font=("Segoe UI", 9, "bold"), fg="#999", bg=COLOR_WHITE).pack(side="left")
        if icon: tk.Label(header, text=icon, font=("Segoe UI", 14), fg="#ccc", bg=COLOR_WHITE).pack(side="right")
        tk.Label(card, text=value, font=FONT_CARD_NUM, bg=COLOR_WHITE, fg="black").pack(anchor="w", pady=(10,0))

    def draw_simple_bar_chart(self, canvas, df):
        if df.empty: return
        w = 400; h = 250; max_val = df['sayi'].max() if not df.empty else 1; max_val = max_val if max_val > 0 else 1
        bar_width = 40; gap = 30; start_x = 50; base_y = h - 30
        canvas.create_line(40, 20, 40, base_y, fill="#eee", width=1); canvas.create_line(40, base_y, w, base_y, fill="#eee", width=1)
        for i, row in df.iterrows():
            val = row['sayi']; label = row['alt_grup']; bar_h = (val / max_val) * (h - 60)
            x0 = start_x + (i * (bar_width + gap)); y0 = base_y - bar_h; x1 = x0 + bar_width; y1 = base_y
            canvas.create_rectangle(x0, y0, x1, y1, fill="#111", outline="")
            canvas.create_text((x0+x1)/2, y1+15, text=label, font=("Segoe UI", 8), fill="#666")
            canvas.create_text((x0+x1)/2, y0-10, text=str(val), font=("Segoe UI", 9, "bold"), fill="#000")

    def draw_dynamic_line_chart(self, canvas, top_5_data):
        if not top_5_data: canvas.create_text(200, 125, text="Yeterli Veri Yok", fill="#999"); return 
        w = 400; h = 250; max_val = top_5_data[0][1]; max_val = max_val if max_val > 0 else 1
        base_y = h - 40; start_x = 50; step_x = (w - 100) / max(len(top_5_data)-1, 1)
        canvas.create_line(40, 20, 40, base_y, fill="#eee", width=1); canvas.create_line(40, base_y, w, base_y, fill="#eee", width=1)
        points = []
        for i, (label, val) in enumerate(top_5_data):
            x = start_x + (i * step_x); y = base_y - ((val / max_val) * (h - 80)); points.append((x, y))
            canvas.create_oval(x-4, y-4, x+4, y+4, fill="black", outline="white", width=2)
            canvas.create_text(x, base_y + 15, text=label[:10], font=("Segoe UI", 8), fill="#666")
            canvas.create_text(x, y - 15, text=str(val), font=("Segoe UI", 9, "bold"), fill="#000")
        if len(points) > 1: canvas.create_line(points, fill="#333", width=2, smooth=False)

    def draw_monthly_rate_chart(self, canvas, data):
        canvas.delete("all")
        if not data:
            canvas.create_text(400, 125, text="Görüntülenecek Veri Yok", fill="#999", font=FONT_NORMAL)
            return
        h = 250; pad_x = 60; pad_y = 40; point_spacing = 150 
        num_points = len(data); required_width = pad_x + (max(num_points - 1, 1) * point_spacing) + pad_x
        visible_width = 900; final_width = max(required_width, visible_width)
        canvas.config(scrollregion=(0, 0, final_width, h))
        max_rate = max([x[3] for x in data]) if data else 1; max_rate = max_rate if max_rate > 0 else 1
        graph_h = h - 2*pad_y
        canvas.create_line(pad_x, h-pad_y, final_width - pad_x, h-pad_y, fill="#ddd", width=2) 
        canvas.create_line(pad_x, pad_y, pad_x, h-pad_y, fill="#ddd", width=2)
        points = []
        for i, row in enumerate(data):
            month_label = row[0]; rate = row[3]; x = pad_x + (i * point_spacing); y = (h - pad_y) - ((rate / max_rate) * graph_h); points.append((x, y))
            canvas.create_line(x, h-pad_y, x, y, fill="#f0f0f0", dash=(2, 4))
            canvas.create_oval(x-6, y-6, x+6, y+6, fill=COLOR_CHART_LINE, outline="white", width=2)
            canvas.create_text(x, h-pad_y+20, text=month_label, font=("Segoe UI", 9), fill="#666")
            canvas.create_text(x, y-20, text=f"{rate:.2f}", font=("Segoe UI", 10, "bold"), fill=COLOR_CHART_LINE)
        if len(points) > 1: canvas.create_line(points, fill=COLOR_CHART_LINE, width=3, smooth=False)

    # ------------------------------------------------------
    # LİSTE EKRANI
    # ------------------------------------------------------
    def show_list_view(self):
        self.clear_content()
        filter_frame = tk.Frame(self.content_area, bg=COLOR_WHITE, padx=15, pady=15); filter_frame.pack(fill="x", pady=(0, 10))
        
        # Grid layout for filters
        tk.Label(filter_frame, text="Şasi No:", bg=COLOR_WHITE).grid(row=0, column=0, sticky="w", padx=5)
        self.f_sasi = ttk.Entry(filter_frame); self.f_sasi.grid(row=1, column=0, padx=5)
        
        tk.Label(filter_frame, text="Araç Tipi:", bg=COLOR_WHITE).grid(row=0, column=1, sticky="w", padx=5)
        self.f_tip = ttk.Combobox(filter_frame, values=["Tümü"] + ARAC_TIPI, state="readonly"); self.f_tip.current(0); self.f_tip.grid(row=1, column=1, padx=5)
        
        tk.Label(filter_frame, text="Alt Grup:", bg=COLOR_WHITE).grid(row=0, column=2, sticky="w", padx=5)
        self.f_alt = ttk.Combobox(filter_frame, values=["Tümü"] + ALT_GRUP, state="readonly"); self.f_alt.current(0); self.f_alt.grid(row=1, column=2, padx=5)
        
        tk.Label(filter_frame, text="Dönem:", bg=COLOR_WHITE).grid(row=0, column=3, sticky="w", padx=5)
        # Dashboard'dakine benzer ay filtresi
        months = ["TÜMÜ"]
        now = datetime.now()
        for i in range(24):
            m_idx = (now.month - i - 1) % 12 + 1
            y_offset = (now.month - i - 1) // 12
            y = now.year + y_offset
            months.append(f"{TURKISH_MONTHS[m_idx]} {y}")
            
        self.f_period = ttk.Combobox(filter_frame, values=months, state="readonly"); self.f_period.current(0); self.f_period.grid(row=1, column=3, padx=5)

        btn_f = tk.Button(filter_frame, text="FİLTRELE", command=self.load_list_data, bg="black", fg="white", relief="flat", padx=10)
        btn_f.grid(row=1, column=4, padx=15)
        
        tk.Button(filter_frame, text="EXCEL İNDİR", command=self.export_excel_with_images, bg="#007bff", fg="white", relief="flat", padx=10).grid(row=1, column=5, padx=5)
        
        if self.current_role == "admin":
            tk.Button(filter_frame, text="DÖNEMİ SİL", command=self.delete_filtered_records, bg="#dc3545", fg="white", relief="flat", padx=10).grid(row=1, column=6, padx=5)
        table_frame = tk.Frame(self.content_area, bg=COLOR_WHITE); table_frame.pack(fill="both", expand=True)
        cols = ("ID", "BB No", "Şasi No", "Araç", "Tarih", "Alt Grup", "Hata", "Top Hata", "Ekleyen")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=vsb.set); vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        
        # Sıralama durumunu tutmak için
        self.tree_sort_state = {}
        
        def sort_treeview(col):
            # Mevcut sıralama durumunu al (varsayılan False = Artan)
            reverse = self.tree_sort_state.get(col, False)
            
            # Tüm verileri al
            data = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
            
            # Tarih sıralaması için özel işlem (DD-MM-YYYY -> YYYY-MM-DD)
            def get_sort_key(item):
                val = item[0]
                # Tarih formatı kontrolü (DD-MM-YYYY)
                if len(val) == 10 and val[2] == '-' and val[5] == '-':
                    try:
                        parts = val.split('-')
                        # YYYY-MM-DD formatına çevir (doğru sıralama için)
                        return f"{parts[2]}-{parts[1]}-{parts[0]}"
                    except:
                        pass
                # Sayısal sıralama
                try:
                    return float(val)
                except:
                    return val
            
            try:
                data.sort(key=get_sort_key, reverse=reverse)
            except:
                data.sort(key=lambda x: x[0], reverse=reverse)
            
            # Yeniden sırala
            for idx, (val, child) in enumerate(data):
                self.tree.move(child, '', idx)
            
            # Sıralama durumunu ters çevir
            self.tree_sort_state[col] = not reverse
            
            # Başlıkları güncelle (ok işareti ekle)
            for c in cols:
                arrow = ""
                if c == col:
                    arrow = " ▼" if not reverse else " ▲"
                self.tree.heading(c, text=c + arrow, command=lambda c=c: sort_treeview(c))
        
        for c in cols: 
            self.tree.heading(c, text=c, command=lambda c=c: sort_treeview(c))
            width = 120 if c in ["Hata", "Top Hata"] else 100
            self.tree.column(c, width=width)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        # Middle Mouse Scroll
        self.tree.bind("<Button-2>", lambda e: self.tree.scan_mark(e.x, e.y))
        self.tree.bind("<B2-Motion>", lambda e: self.tree.scan_dragto(e.x, e.y, gain=1))
        bottom_container = tk.Frame(self.content_area, bg=COLOR_BG, pady=10); bottom_container.pack(fill="x")
        self.preview_frame = tk.Frame(bottom_container, bg=COLOR_WHITE, height=150, bd=1, relief="solid"); self.preview_frame.pack(side="left", fill="x", expand=True, padx=(0, 20)); self.preview_frame.pack_propagate(False)
        self.lbl_preview_img = tk.Label(self.preview_frame, text="Kayıt seçiniz...", bg=COLOR_WHITE, fg="#999"); self.lbl_preview_img.pack(side="left", padx=10, fill="y")
        self.btn_open_big = tk.Button(self.preview_frame, text="BÜYÜK GÖRÜNTÜLE", state="disabled", bg="#eee", relief="flat"); self.btn_open_big.pack(side="right", padx=10, pady=10); self.current_preview_path = None
        action_bar = tk.Frame(bottom_container, bg=COLOR_BG); action_bar.pack(side="right", fill="y")
        tk.Button(action_bar, text="DÜZENLE", command=self.open_edit_window, bg="#ffc107", relief="flat", width=15, height=2).pack(side="top", pady=2)
        if self.current_role == "admin": tk.Button(action_bar, text="SİL", command=self.delete_record, bg="#dc3545", fg="white", relief="flat", width=15, height=2).pack(side="top", pady=2)
        self.load_list_data()

    def on_tree_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        rec_id = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); path = conn.execute("SELECT fotograf_yolu FROM pdi_kayitlari WHERE id=?", (rec_id,)).fetchone()[0]; conn.close(); self.current_preview_path = path
        if path and os.path.exists(path):
            try:
                img = PilImage.open(path); base_height = 130; w_percent = (base_height / float(img.size[1])); w_size = int((float(img.size[0]) * float(w_percent)))
                img = img.resize((w_size, base_height), PilImage.Resampling.LANCZOS); photo = ImageTk.PhotoImage(img); self.lbl_preview_img.config(image=photo, text=""); self.lbl_preview_img.image = photo
                self.btn_open_big.config(state="normal", bg=COLOR_BLUE, fg="white", command=lambda: os.startfile(path))
            except: self.lbl_preview_img.config(image="", text="⚠️ Resim Önizlenemiyor"); self.btn_open_big.config(state="disabled", bg="#eee")
        else: self.lbl_preview_img.config(image="", text="📷 Dosya Yok"); self.btn_open_big.config(state="disabled", bg="#eee")

    def load_list_data(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        sasi = self.f_sasi.get().strip()
        tip = self.f_tip.get()
        alt = self.f_alt.get()
        period = self.f_period.get()
        
        conn = self.get_db_connection()
        query = "SELECT id, bb_no, sasi_no, arac_tipi, tarih_saat, alt_grup, hata_konumu, top_hata, kullanici FROM pdi_kayitlari"
        df = pd.read_sql_query(query, conn)
        
        # Filtreleme (PANDAS)
        if sasi:
            df = df[df['sasi_no'].str.contains(sasi, case=False, na=False) | df['bb_no'].str.contains(sasi, case=False, na=False)]
        
        if tip != "Tümü":
            df = df[df['arac_tipi'] == tip]
            
        if alt != "Tümü":
            df = df[df['alt_grup'] == alt]
            
        if period != "TÜMÜ":
            try:
                parts = period.split()
                m_name = parts[0]
                year_f = int(parts[1])
                m_idx = next((k for k, v in TURKISH_MONTHS.items() if v == m_name), 0)
                
                df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
                df = df.dropna(subset=['dt'])
                df = df[(df['dt'].dt.month == m_idx) & (df['dt'].dt.year == year_f)]
            except: pass
            
        # Sonuçları ekle (Ters sırada, en yeni en üstte)
        df = df.sort_values(by='id', ascending=False)
        
        for _, row in df.iterrows():
            # Tarih kısmını sadece gün olarak göster (eski mantıkta split(" ")[0] vardı ama format GG-AA-YYYY zaten)
            date_val = str(row['tarih_saat']).split(" ")[0] if row['tarih_saat'] else ""
            
            vals = [
                row['id'], row['bb_no'], row['sasi_no'], row['arac_tipi'],
                date_val, row['alt_grup'], row['hata_konumu'], 
                row['top_hata'], row['kullanici']
            ]
            vals = [str(v) if v is not None else "" for v in vals]
            self.tree.insert("", "end", values=vals)

    def delete_filtered_records(self):
        sasi = self.f_sasi.get().strip()
        tip = self.f_tip.get()
        alt = self.f_alt.get()
        period = self.f_period.get()
        
        filter_desc = []
        if period != "TÜMÜ": filter_desc.append(f"Dönem: {period}")
        if sasi: filter_desc.append(f"Şasi: {sasi}")
        if tip != "Tümü": filter_desc.append(f"Tip: {tip}")
        if alt != "Tümü": filter_desc.append(f"Alt Grup: {alt}")
        
        msg = "Filtrelenmiş TÜM kayıtları siliyorsunuz." if filter_desc else "Sistemdeki TÜM kayıtları siliyorsunuz."
        if filter_desc:
            msg += "\nKriterler: " + ", ".join(filter_desc)
            
        ans = messagebox.askyesno("TOPLU SİLME ONAYI", f"{msg}\n\nSeçili kriterlere uyan TÜM veriler kalıcı olarak silinecektir.\n\nBu işlem geri alınamaz! Onaylıyor musunuz?")
        if not ans: return
        
        # OBTİMİZE EDİLMİŞ VE GÜVENLİ SİLME (Pandas ile ID'leri bulup siliyoruz)
        conn = self.get_db_connection()
        query = "SELECT id, bb_no, sasi_no, arac_tipi, tarih_saat, alt_grup FROM pdi_kayitlari"
        df = pd.read_sql_query(query, conn)
        
        # Filtreleme (load_list_data ile aynı mantık)
        if sasi:
            df = df[df['sasi_no'].str.contains(sasi, case=False, na=False) | df['bb_no'].str.contains(sasi, case=False, na=False)]
        if tip != "Tümü":
            df = df[df['arac_tipi'] == tip]
        if alt != "Tümü":
            df = df[df['alt_grup'] == alt]
        if period != "TÜMÜ":
            try:
                parts = period.split()
                m_name = parts[0]
                year_f = int(parts[1])
                m_idx = next((k for k, v in TURKISH_MONTHS.items() if v == m_name), 0)
                df['dt'] = pd.to_datetime(df['tarih_saat'], format='%d-%m-%Y', errors='coerce')
                df = df.dropna(subset=['dt'])
                df = df[(df['dt'].dt.month == m_idx) & (df['dt'].dt.year == year_f)]
            except: pass
        
        ids_to_delete = df['id'].tolist()
        
        if not ids_to_delete:
            messagebox.showinfo("BİLGİ", "Silinecek kayıt bulunamadı.")
            return

        try:
            c = conn.cursor()
            # SQLite batch size limitine takılmamak için parçalı silebiliriz
            batch_size = 500
            for i in range(0, len(ids_to_delete), batch_size):
                batch = ids_to_delete[i:i+batch_size]
                placeholders = ",".join(["?" for _ in batch])
                c.execute(f"DELETE FROM pdi_kayitlari WHERE id IN ({placeholders})", batch)
            conn.commit()
            messagebox.showinfo("BAŞARILI", f"{len(ids_to_delete)} kayıt başarıyla silindi.")
            self.load_list_data()
        except Exception as e:
            messagebox.showerror("HATA", f"Silme işlemi sırasında hata oluştu:\n{e}")

    def clear_content(self):
        for widget in self.content_area.winfo_children(): widget.destroy()

    def export_excel_with_images(self):
        conn = sqlite3.connect(DB_NAME); query = "SELECT id, bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, tarih_saat, kullanici, duzenleyen, fotograf_yolu FROM pdi_kayitlari"; df = pd.read_sql_query(query, conn); conn.close()
        if df.empty: messagebox.showwarning("Uyarı", "Veri yok."); return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not file_path: return
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.drop(columns=['fotograf_yolu']).to_excel(writer, sheet_name='Kayıtlar', index=False)
                wb = writer.book; ws_img = wb.create_sheet("FOTOĞRAFLAR"); ws_img.column_dimensions['A'].width = 10; ws_img.column_dimensions['B'].width = 50
                ws_img.cell(row=1, column=1, value="ID"); ws_img.cell(row=1, column=2, value="FOTOĞRAF"); row_idx = 2
                from openpyxl.drawing.image import Image as XLImage
                for index, row in df.iterrows():
                    rid = row['id']; path = row['fotograf_yolu']; ws_img.cell(row=row_idx, column=1, value=rid); ws_img.row_dimensions[row_idx].height = 100
                    if path and os.path.exists(path):
                        try:
                            img = PilImage.open(path); img.thumbnail((300, 130)); xl_img = XLImage(path); xl_img.height = 130; xl_img.width = 130 * (img.width / img.height); ws_img.add_image(xl_img, f"B{row_idx}")
                        except Exception as e: ws_img.cell(row=row_idx, column=2, value=f"Hata: {str(e)}")
                    else: ws_img.cell(row=row_idx, column=2, value="Resim Yok")
                    row_idx += 1
            messagebox.showinfo("Başarılı", "Excel başarıyla oluşturuldu.")
        except Exception as e: messagebox.showerror("Hata", f"Excel oluşturulurken hata: {e}")

    def delete_record(self):
        sel = self.tree.selection(); 
        if not sel: return
        if messagebox.askyesno("Sil", "Kayıt silinsin mi?"): rid = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); conn.execute("DELETE FROM pdi_kayitlari WHERE id=?", (rid,)); conn.commit(); conn.close(); self.load_list_data()

    def open_add_window(self): self.popup_window("Yeni Kayıt Ekle")
    def open_edit_window(self):
        sel = self.tree.selection(); 
        if not sel: return
        rid = self.tree.item(sel[0])['values'][0]; conn = sqlite3.connect(DB_NAME); row = conn.execute("SELECT * FROM pdi_kayitlari WHERE id=?", (rid,)).fetchone(); conn.close(); self.popup_window("Kayıt Düzenle", row)

    def open_user_management(self):
        win = tk.Toplevel(self.root); win.title("Kullanıcı Yönetimi"); win.geometry("800x500"); win.configure(bg=COLOR_BG)
        tk.Label(win, text="KULLANICI YÖNETİMİ", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=10)
        main_frame = tk.Frame(win, bg=COLOR_BG); main_frame.pack(fill="both", expand=True, padx=20, pady=10)
        left_frame = tk.Frame(main_frame, bg=COLOR_WHITE, padx=10, pady=10); left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))
        tk.Label(left_frame, text="Kayıtlı Kullanıcılar", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w", pady=(0,5))
        lb_users = tk.Listbox(left_frame, font=("Segoe UI", 11), borderwidth=0, bg="#f9f9f9", selectbackground="#ddd", selectforeground="black"); lb_users.pack(fill="both", expand=True)
        right_frame = tk.Frame(main_frame, bg=COLOR_WHITE, padx=10, pady=10); right_frame.pack(side="right", fill="both", expand=True)
        tk.Label(right_frame, text="Yeni Kullanıcı Ekle / Düzenle", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w", pady=(0,15))
        tk.Label(right_frame, text="Kullanıcı Adı:", bg=COLOR_WHITE).pack(anchor="w"); e_u = ttk.Entry(right_frame); e_u.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Şifre:", bg=COLOR_WHITE).pack(anchor="w"); e_p = ttk.Entry(right_frame); e_p.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Rol:", bg=COLOR_WHITE).pack(anchor="w"); cb_role = ttk.Combobox(right_frame, values=["Admin", "Kullanıcı"], state="readonly"); cb_role.current(1); cb_role.pack(fill="x", pady=(0,10))
        tk.Label(right_frame, text="Açıklama:", bg=COLOR_WHITE).pack(anchor="w"); e_desc = ttk.Entry(right_frame); e_desc.pack(fill="x", pady=(0,20))
        def load_users():
            lb_users.delete(0, "end"); conn = sqlite3.connect(DB_NAME); rows = conn.execute("SELECT username, role FROM users").fetchall(); conn.close()
            for r in rows: lb_users.insert("end", r[0] + (" (Yönetici)" if r[1] == 1 else " (Kullanıcı)"))
        
        def on_user_select(event):
            selection = lb_users.curselection()
            if not selection: return
            text = lb_users.get(selection[0])
            username = text.split(" (")[0]
            conn = sqlite3.connect(DB_NAME)
            user_data = conn.execute("SELECT username, password, role, aciklama FROM users WHERE username=?", (username,)).fetchone()
            conn.close()
            if user_data:
                e_u.delete(0, "end"); e_p.delete(0, "end"); e_desc.delete(0, "end")
                e_u.insert(0, user_data[0]); e_p.insert(0, user_data[1])
                cb_role.set("Admin" if user_data[2] == 1 else "Kullanıcı")
                if user_data[3]: e_desc.insert(0, user_data[3])

        lb_users.bind("<<ListboxSelect>>", on_user_select)

        def save_user():
            u = e_u.get().strip(); p = e_p.get().strip(); role = 1 if cb_role.get() == "Admin" else 0; desc = e_desc.get().strip()
            if not u or not p: messagebox.showwarning("Hata", "Kullanıcı adı ve şifre zorunludur."); return
            conn = sqlite3.connect(DB_NAME)
            try: conn.execute("INSERT OR REPLACE INTO users (username, password, role, aciklama) VALUES (?,?,?,?)", (u, p, role, desc)); conn.commit(); messagebox.showinfo("Başarılı", "Kullanıcı kaydedildi."); e_u.delete(0, "end"); e_p.delete(0, "end"); e_desc.delete(0, "end"); load_users()
            except Exception as e: messagebox.showerror("Hata", str(e));
            finally: conn.close()
        def delete_user():
            sel = lb_users.curselection(); 
            if not sel: return
            username = lb_users.get(sel[0]).split(" (")[0]
            if username == "admin": messagebox.showerror("Hata", "Ana 'admin' silinemez."); return
            if messagebox.askyesno("Sil", f"{username} silinsin mi?"): conn = sqlite3.connect(DB_NAME); conn.execute("DELETE FROM users WHERE username=?", (username,)); conn.commit(); conn.close(); load_users(); messagebox.showinfo("Silindi", "Kullanıcı silindi.")
        tk.Button(right_frame, text="KAYDET", command=save_user, bg="black", fg="white", relief="flat", height=2).pack(fill="x", pady=(0,5))
        tk.Button(right_frame, text="SEÇİLİ KULLANICIYI SİL", command=delete_user, bg=COLOR_RED, fg="white", relief="flat").pack(fill="x"); load_users()

    def popup_window(self, title, data=None):
        top = tk.Toplevel(self.root); top.title(title); top.geometry("950x750"); top.configure(bg=COLOR_BG)
        header_frame = tk.Frame(top, bg="black", pady=10, padx=20); header_frame.pack(fill="x")
        tk.Label(header_frame, text=title.upper(), font=("Segoe UI", 14, "bold"), bg="black", fg="white").pack(side="left")
        main_frame = tk.Frame(top, bg=COLOR_WHITE, padx=20, pady=20); main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        info_lbl = tk.Label(main_frame, text="ARAÇ BİLGİLERİ", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888"); info_lbl.pack(anchor="w", pady=(0,10))
        grid_frame = tk.Frame(main_frame, bg=COLOR_WHITE); grid_frame.pack(fill="x", pady=(0, 20))
        
        # --- TARİH SEÇİCİ İÇİN ÖZEL ALAN ---
        def open_calendar(entry_widget):
            current_val = entry_widget.get()
            dialog = AdvancedCalendarDialog(top, current_val)
            if dialog.result:
                entry_widget.delete(0, "end")
                entry_widget.insert(0, dialog.result)

        def add_field(lbl, r, c, val=None, combo=None, is_date=False):
            tk.Label(grid_frame, text=lbl, bg=COLOR_WHITE, font=FONT_NORMAL, fg="#333").grid(row=r, column=c, sticky="w", padx=10)
            
            if is_date:
                f = tk.Frame(grid_frame, bg=COLOR_WHITE)
                f.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
                w = ttk.Entry(f, font=FONT_NORMAL)
                w.pack(side="left", fill="x", expand=True)
                if val: w.insert(0, val)
                btn_cal = tk.Button(f, text="📅", command=lambda: open_calendar(w), 
                                    relief="flat", bg="#eee", width=3)
                btn_cal.pack(side="right", padx=(5,0))
            elif combo: 
                w = ttk.Combobox(grid_frame, values=combo, state="readonly", font=FONT_NORMAL)
                if not val: w.current(0) 
                else: w.set(val)
                w.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
            else: 
                w = ttk.Entry(grid_frame, font=FONT_NORMAL)
                if val: w.insert(0, val)
                w.grid(row=r+1, column=c, sticky="ew", padx=10, pady=(2, 10))
            return w
            
        grid_frame.columnconfigure(0, weight=1); grid_frame.columnconfigure(1, weight=1); grid_frame.columnconfigure(2, weight=1)
        
        e_bb = add_field("BB No", 0, 0, data[1] if data else "")
        e_sasi = add_field("Şasi No", 0, 1, data[2] if data else "")
        e_arac = add_field("Araç Tipi", 0, 2, data[3] if data else "", ARAC_TIPI)
        e_isemri = add_field("İş Emri No", 2, 0, data[4] if data else "")
        
        today_str = datetime.now().strftime("%d-%m-%Y")
        e_tarih = add_field("Tarih (G-A-Y)", 2, 1, data[9].split(" ")[0] if data and data[9] else today_str, is_date=True)
        
        e_alt = add_field("Alt Grup", 2, 2, data[5] if data else "", ALT_GRUP)
        
        split_frame = tk.Frame(main_frame, bg=COLOR_WHITE); split_frame.pack(fill="both", expand=True)
        left_col = tk.Frame(split_frame, bg=COLOR_WHITE, width=300); left_col.pack(side="left", fill="both", expand=False, padx=(0, 20))
        
        # HATA KONUMU - Artık serbest metin alanı
        tk.Label(left_col, text="HATA KONUMU", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w")
        tk.Label(left_col, text="(Virgülle ayırarak yazın)", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#aaa").pack(anchor="w")
        e_hata_konumu = tk.Text(left_col, height=4, font=FONT_NORMAL, bg="#f9f9f9", bd=1, relief="solid")
        e_hata_konumu.pack(fill="x", pady=5)
        if data and data[7]: e_hata_konumu.insert("1.0", data[7])
        
        # TOP HATA - Sadece düzenleme modunda görünür (kayıt sonrası eklenir)
        e_top_hata = None
        if data:
            tk.Label(left_col, text="TOP HATA", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w", pady=(15, 0))
            tk.Label(left_col, text="(Kayıt sonrası seçilir)", font=("Segoe UI", 8), bg=COLOR_WHITE, fg="#aaa").pack(anchor="w")
            
            # Top Hata listesini veritabanından al
            conn = sqlite3.connect(DB_NAME)
            top_hatalar = conn.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()
            conn.close()
            top_hata_list = [""] + [h[0] for h in top_hatalar]
            
            e_top_hata = ttk.Combobox(left_col, values=top_hata_list, state="readonly", font=FONT_NORMAL)
            e_top_hata.pack(fill="x", pady=5)
            
            # Mevcut top_hata değerini bul ve seç
            try:
                current_top_hata = conn.execute("SELECT top_hata FROM pdi_kayitlari WHERE id=?", (data[0],)).fetchone()
                if current_top_hata and current_top_hata[0]:
                    conn = sqlite3.connect(DB_NAME)
                    current_top_hata = conn.execute("SELECT top_hata FROM pdi_kayitlari WHERE id=?", (data[0],)).fetchone()
                    conn.close()
                    if current_top_hata and current_top_hata[0] in top_hata_list:
                        e_top_hata.set(current_top_hata[0])
            except: pass
        
        right_col = tk.Frame(split_frame, bg=COLOR_WHITE); right_col.pack(side="right", fill="both", expand=True)
        tk.Label(right_col, text="TESPİT DETAYLARI", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w")
        txt_tespit = tk.Text(right_col, height=6, font=FONT_NORMAL, bg="#f9f9f9", bd=1, relief="solid"); txt_tespit.pack(fill="x", pady=5)
        if data and data[6]: txt_tespit.insert("1.0", data[6])
        
        # FOTOĞRAF KANITI - Sadece düzenleme modunda görünür
        # FOTOĞRAF KANITI - Hem yeni hem düzenleme için
        self.selected_photo_path = data[8] if data and data[8] else ""
        tk.Label(right_col, text="FOTOĞRAF KANITI", font=("Segoe UI", 11, "bold"), bg=COLOR_WHITE, fg="#888").pack(anchor="w", pady=(10,0))
        photo_frame = tk.Frame(right_col, bg="#f9f9f9", bd=1, relief="solid", pady=10, padx=10); photo_frame.pack(fill="x", pady=5)
        lbl_path = tk.Label(photo_frame, text=os.path.basename(self.selected_photo_path) if self.selected_photo_path else "Dosya seçilmedi...", bg="#f9f9f9", fg="#555", anchor="w"); lbl_path.pack(side="left", fill="x", expand=True)
        def sel_photo():
            f = filedialog.askopenfilename(filetypes=[("Resimler", "*.jpg *.png *.jpeg")]); 
            if f: self.selected_photo_path = f; lbl_path.config(text=os.path.basename(f))
        tk.Button(photo_frame, text="GÖZAT...", command=sel_photo, bg="#ddd", relief="flat").pack(side="right")
        
        def save():
            bb = e_bb.get(); sasi = e_sasi.get()
            hata_str = e_hata_konumu.get("1.0", "end-1c").strip()
            top_hata_val = e_top_hata.get() if e_top_hata else ""
            
            if not sasi: messagebox.showwarning("Eksik", "Şasi No giriniz."); return
            
            # Fotoğrafı ortak klasöre kopyala (Eğer seçilmişse ve henüz orada değilse)
            final_photo_path = self.selected_photo_path
            if self.selected_photo_path and os.path.exists(self.selected_photo_path):
                # Eğer dosya zaten PHOTO_DIR içinde değilse kopyala
                if not self.selected_photo_path.startswith(PHOTO_DIR):
                    try:
                        ext = os.path.splitext(self.selected_photo_path)[1]
                        # Benzersiz isim: SASI_ZAMAN damgası
                        new_filename = f"{sasi}_{datetime.now().strftime('%Y%H%M%S')}{ext}"
                        new_dest = os.path.join(PHOTO_DIR, new_filename)
                        shutil.copy2(self.selected_photo_path, new_dest)
                        final_photo_path = new_dest
                    except Exception as e:
                        print(f"Fotoğraf kopyalama hatası: {e}")

            conn = sqlite3.connect(DB_NAME); c = conn.cursor()
            if data: 
                c.execute("""UPDATE pdi_kayitlari SET bb_no=?, sasi_no=?, arac_tipi=?, is_emri_no=?, alt_grup=?, tespitler=?, hata_konumu=?, fotograf_yolu=?, tarih_saat=?, duzenleyen=?, top_hata=? WHERE id=?""", 
                         (bb, sasi, e_arac.get(), e_isemri.get(), e_alt.get(), txt_tespit.get("1.0", "end-1c"), hata_str, final_photo_path, e_tarih.get(), self.current_user, top_hata_val, data[0]))
            else: 
                c.execute("""INSERT INTO pdi_kayitlari (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, fotograf_yolu, tarih_saat, kullanici) VALUES (?,?,?,?,?,?,?,?,?,?)""", 
                         (bb, sasi, e_arac.get(), e_isemri.get(), e_alt.get(), txt_tespit.get("1.0", "end-1c"), hata_str, final_photo_path, e_tarih.get(), self.current_user))
            conn.commit(); conn.close(); top.destroy(); self.show_dashboard(); messagebox.showinfo("Başarılı", "Kayıt işlemi tamamlandı.")
        tk.Button(main_frame, text="KAYDET VE KAPAT", command=save, bg="black", fg="white", font=("Segoe UI", 12, "bold"), height=2, relief="flat").pack(fill="x", pady=20)

    # ------------------------------------------------------
    # TOP HATA YÖNETİMİ (ADMIN)
    # ------------------------------------------------------
    def open_top_hata_management(self):
        win = tk.Toplevel(self.root)
        win.title("Top Hata Yönetimi")
        win.geometry("600x500")
        win.configure(bg=COLOR_BG)
        
        tk.Label(win, text="TOP HATA YÖNETİMİ", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=15)
        
        main_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Liste
        tk.Label(main_frame, text="Kayıtlı Top Hatalar", font=FONT_BOLD, bg=COLOR_WHITE).pack(anchor="w")
        
        list_frame = tk.Frame(main_frame, bg="#f9f9f9", bd=1, relief="solid")
        list_frame.pack(fill="both", expand=True, pady=10)
        
        lb_hatalar = tk.Listbox(list_frame, font=FONT_NORMAL, height=12, bd=0, bg="#f9f9f9", selectbackground="#ddd", selectforeground="black")
        lb_hatalar.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        sb = tk.Scrollbar(list_frame, orient="vertical", command=lb_hatalar.yview)
        sb.pack(side="right", fill="y")
        lb_hatalar.config(yscrollcommand=sb.set)
        
        def load_hatalar():
            lb_hatalar.delete(0, "end")
            conn = sqlite3.connect(DB_NAME)
            rows = conn.execute("SELECT hata_adi, aktif FROM top_hatalar ORDER BY hata_adi").fetchall()
            conn.close()
            for r in rows:
                status = " ✓" if r[1] == 1 else " (pasif)"
                lb_hatalar.insert("end", r[0] + status)
        
        # Ekleme alanı
        add_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        add_frame.pack(fill="x", pady=10)
        
        tk.Label(add_frame, text="Yeni Top Hata:", bg=COLOR_WHITE, font=FONT_NORMAL).pack(side="left")
        e_new = ttk.Entry(add_frame, font=FONT_NORMAL, width=40)
        e_new.pack(side="left", padx=10)
        
        def add_hata():
            hata = e_new.get().strip()
            if not hata:
                messagebox.showwarning("Uyarı", "Hata adı boş olamaz.")
                return
            try:
                conn = sqlite3.connect(DB_NAME)
                conn.execute("INSERT INTO top_hatalar (hata_adi, aktif, olusturma_tarihi) VALUES (?, 1, ?)", 
                           (hata, datetime.now().strftime("%d-%m-%Y")))
                conn.commit()
                conn.close()
                e_new.delete(0, "end")
                load_hatalar()
                messagebox.showinfo("Başarılı", f"'{hata}' eklendi.")
            except sqlite3.IntegrityError:
                messagebox.showerror("Hata", "Bu hata zaten mevcut.")
        
        tk.Button(add_frame, text="EKLE", command=add_hata, bg=COLOR_GREEN, fg="white", relief="flat", padx=15).pack(side="left")
        
        # Silme butonu
        def delete_hata():
            sel = lb_hatalar.curselection()
            if not sel:
                messagebox.showwarning("Uyarı", "Silmek için bir hata seçin.")
                return
            hata = lb_hatalar.get(sel[0]).replace(" ✓", "").replace(" (pasif)", "")
            if messagebox.askyesno("Sil", f"'{hata}' silinsin mi?"):
                conn = sqlite3.connect(DB_NAME)
                conn.execute("DELETE FROM top_hatalar WHERE hata_adi=?", (hata,))
                conn.commit()
                conn.close()
                load_hatalar()
                messagebox.showinfo("Silindi", "Hata silindi.")
        
        btn_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="SEÇİLİ HATAYI SİL", command=delete_hata, bg=COLOR_RED, fg="white", relief="flat", padx=15).pack(side="right")
        
        load_hatalar()

    # ------------------------------------------------------
    # EXCEL IMPORT (ADMIN)
    # ------------------------------------------------------
    def open_excel_import(self):
        win = tk.Toplevel(self.root)
        win.title("Geçmiş Veri Yükle")
        win.geometry("700x550")
        win.configure(bg=COLOR_BG)
        
        tk.Label(win, text="GEÇMİŞ VERİ YÜKLE (EXCEL)", font=FONT_HEADER, bg=COLOR_BG, fg="#333").pack(pady=15)
        
        main_frame = tk.Frame(win, bg=COLOR_WHITE, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Açıklama
        info_text = """Excel dosyanız aşağıdaki sütunları içermelidir (Otomatik tanınır):
        
Sıralama Örneği:
BB No | Şasi No | Araç Tipi | İş Emri No | PDI Tarihi | Tespitler | Hata Konumu | Alt Grup

⚠️ "Kullanıcı" sütunu gerekmez (otomatik atanır).
⚠️ Araç Tipi Kısaltmaları Desteklenir:
   - TOU -> Tourismo
   - TRV -> Travego
   - CON -> Connecto
   - CONNECTO -> Connecto
"""
        
        tk.Label(main_frame, text=info_text, font=("Segoe UI", 9), bg=COLOR_WHITE, fg="#555", justify="left").pack(anchor="w", pady=(0, 20))
        
        # Dosya seçici
        file_frame = tk.Frame(main_frame, bg=COLOR_WHITE)
        file_frame.pack(fill="x", pady=10)
        
        self.import_file_path = tk.StringVar()
        tk.Label(file_frame, text="Excel Dosyası:", bg=COLOR_WHITE, font=FONT_BOLD).pack(side="left")
        tk.Entry(file_frame, textvariable=self.import_file_path, font=FONT_NORMAL, width=50, state="readonly").pack(side="left", padx=10)
        
        def browse_file():
            # Parent parametresi pencerenin arkaya düşmesini engeller
            f = filedialog.askopenfilename(parent=win, filetypes=[("Excel", "*.xlsx *.xls")])
            if f:
                self.import_file_path.set(f)
            win.lift() # Pencereyi tekrar en üste getir
            win.focus_force() # Odağı geri al
        
        tk.Button(file_frame, text="GÖZAT", command=browse_file, bg="#ddd", relief="flat", padx=10).pack(side="left")
        
        # Kapatıldığında ana ekranı yenile (Opsiyonel ama kullanışlı)
        win.protocol("WM_DELETE_WINDOW", lambda: (self.show_dashboard(), win.destroy()))
        
        # Sonuç alanı (En altta ama tanımlama için önce oluşturuyoruz)
        result_frame = tk.Frame(main_frame, bg="#f9f9f9", bd=1, relief="solid")
        result_text = tk.Text(result_frame, font=("Consolas", 9), bg="#f9f9f9", height=10, bd=0)

        def import_data():
            file_path = self.import_file_path.get()
            if not file_path:
                messagebox.showwarning("Uyarı", "Lütfen bir Excel dosyası seçin.")
                return
            
            try:
                df = pd.read_excel(file_path, dtype=str)
                result_text.delete("1.0", "end")
                df.columns = [str(c).strip() for c in df.columns]
                
                success_count = 0; error_count = 0
                conn = sqlite3.connect(DB_NAME); c = conn.cursor()
                
                type_map = {
                    "TOU": "Tourismo", "TRV": "Travego", "CON": "Connecto",
                    "TOURISMO": "Tourismo", "TRAVEGO": "Travego", "CONNECTO": "Connecto",
                    "TOURİSMO": "Tourismo"
                }

                # Sütun isimlerinden indexleri bul (Süper esnek ve Locale bağımsız)
                def str_norm(s):
                    if not s: return ""
                    s = str(s).lower()
                    # Manuel karakter eşleme
                    mapping = {'ç':'c','ğ':'g','ı':'i','i':'i','ö':'o','ş':'s','ü':'u','â':'a','î':'i','û':'u'}
                    for k, v in mapping.items():
                        s = s.replace(k, v)
                    # Sadece a-z ve 0-9 kalsın
                    return "".join(c for c in s if 'a' <= c <= 'z' or '0' <= c <= '9').upper()

                cols_map = {
                    "bb_no": ["BB No", "BB NO", "BBNO"],
                    "sasi_no": ["Şasi No", "ŞASİ NO", "SASİ NO", "Sasi No", "SASI NO", "SASI", "CHASSIS"],
                    "arac_tipi": ["Araç Tipi", "ARAÇ TİPİ", "ARAC TIPI", "TİP", "ARAC", "TIP", "MODEL"],
                    "is_emri": ["İş Emri No", "İŞ EMRİ NO", "IS EMRI NO", "İŞ EMRİ", "IS EMRI", "IS EMRE", "ISEMRI"],
                    "tarih": ["PDI Tarihi", "PDI TARİHİ", "PDI Yapılış Tarihi", "Tarih", "TARİH", "YAPILIŞ TARİHİ", "PDI DATE", "DATE"],
                    "tespitler": ["Tespitler", "TESPİTLER", "TESPITLER", "HATA", "TESPİT", "FINDINGS", "DESCRIPTION"],
                    "konum": ["Hata Konumu", "HATA KONUMU", "KONUM", "HATA YERİ", "LOCATION"],
                    "grup": ["Alt Grup", "ALT GRUP", "GRUP", "SUBGROUP"]
                }
                
                found_indices = {}
                # Debug: Gelen başlıkları göster
                raw_cols = [str(c) for c in df.columns if not str(c).startswith("Unnamed")]
                result_text.insert("end", f"📄 Algılanan Başlıklar: {', '.join(raw_cols[:8])}...\n")
                result_text.insert("end", "🔍 Sütun Tanımlama:\n")
                
                # Normalize edilmiş başlıkları bir kez hazırla
                df_cols_norm = [str_norm(c) for c in df.columns]
                
                for key, variations in cols_map.items():
                    norm_vars = [str_norm(v) for v in variations]
                    for idx, norm_col_name in enumerate(df_cols_norm):
                        if norm_col_name in norm_vars:
                            found_indices[key] = idx
                            result_text.insert("end", f"   ✅ {key.replace('_',' ').upper()}: '{df.columns[idx]}' eşleşti.\n")
                            break
                    if key not in found_indices:
                        result_text.insert("end", f"   ❌ {key.replace('_',' ').upper()}: Eşleşme Bulunamadı!\n")

                # Kritik Sütun Kontrolü
                critical_missing = []
                if "sasi_no" not in found_indices: critical_missing.append("Şasi No")
                if "arac_tipi" not in found_indices: critical_missing.append("Araç Tipi")
                if "tarih" not in found_indices: critical_missing.append("Tarih")
                
                if critical_missing:
                    messagebox.showerror("Hata", f"Excel'de şu kritik sütunlar bulunamadı:\n\n{', '.join(critical_missing)}\n\nLütfen başlıkları kontrol edin.")
                    return

                # Kritik Sütun Kontrolü
                if "sasi_no" not in found_indices:
                    messagebox.showerror("Hata", "Excel dosyasında 'Şasi No' sütunu bulunamadı!")
                    return

                def get_val(row_vals, key, default=""):
                    idx = found_indices.get(key)
                    if idx is not None and idx < len(row_vals):
                        v = row_vals[idx]
                        return str(v).strip() if pd.notna(v) and str(v).lower() != 'nan' else default
                    return default

                # İlk Geçiş: Kayıtları hazırla ve mükerrer kontrolü yap
                to_import = []
                duplicates_found = 0
                
                for idx, row in df.iterrows():
                    row_vals = row.values
                    sasi_no = get_val(row_vals, "sasi_no")
                    if not sasi_no: continue
                    
                    bb_no = get_val(row_vals, "bb_no")
                    raw_type = get_val(row_vals, "arac_tipi")
                    is_emri_no = get_val(row_vals, "is_emri")
                    raw_date = get_val(row_vals, "tarih")
                    tespitler = get_val(row_vals, "tespitler")
                    hata_konumu = get_val(row_vals, "konum")
                    alt_grup = get_val(row_vals, "grup")
                    
                    # Araç tipi eşleştirme (Kısaltmalar için)
                    arac_tipi = type_map.get(raw_type.upper(), raw_type)
                    if arac_tipi not in ["Tourismo", "Travego", "Connecto"]:
                        arac_tipi = arac_tipi.capitalize()
                        if arac_tipi not in ["Tourismo", "Travego", "Connecto"]:
                            # Eğer tip bulunamazsa boş kalsın mı? Hayır, Tourismo'ya default edelim veya uyarı verelim
                            # Kullanıcının gönderdiği resimde "Şehir İçi" vs olabilir diye daha esnek yapalım.
                            pass

                    # Tarih Normalizasyonu
                    tarih = str(raw_date).split(" ")[0].replace("/", "-").replace(".", "-")
                    try:
                        parts = tarih.split("-")
                        if len(parts) == 3:
                            if len(parts[0]) == 4: tarih = f"{parts[2].zfill(2)}-{parts[1].zfill(2)}-{parts[0]}"
                            else: tarih = f"{parts[0].zfill(2)}-{parts[1].zfill(2)}-{parts[2]}"
                    except: pass
                    
                    # Mevcut mu kontrolü
                    q = "SELECT id FROM pdi_kayitlari WHERE sasi_no=? AND is_emri_no=? AND tespitler=? AND tarih_saat=?"
                    exists = c.execute(q, (sasi_no, is_emri_no, tespitler, tarih)).fetchone()
                    if exists: duplicates_found += 1
                    
                    to_import.append({
                        "bb_no": bb_no, "sasi_no": sasi_no, "arac_tipi": arac_tipi, 
                        "is_emri_no": is_emri_no, "alt_grup": alt_grup, "tespitler": tespitler, 
                        "hata_konumu": hata_konumu, "tarih": tarih, "idx": idx+2
                    })

                if not to_import:
                    messagebox.showwarning("Uyarı", "Excel'de işlenebilecek geçerli kayıt bulunamadı.")
                    conn.close()
                    return

                if duplicates_found > 0:
                    ans = messagebox.askyesno("Mükerrer Kayıt", 
                        f"Seçtiğiniz dosyada daha önce yüklenmiş {duplicates_found} adet kayıt tespit edildi.\n\n"
                        "Bu kayıtları silip yenilerini üzerine yazmak istiyor musunuz?\n"
                        "(Hayır derseniz işlem iptal edilecektir.)")
                    if not ans:
                        conn.close()
                        return

                # İkinci Geçiş: Gerçek Yazma İşlemi
                success_count = 0; error_count = 0
                for item in to_import:
                    try:
                        # Eğer onay verildiyse ve mükerrerse önce eskiyi siliyor
                        if duplicates_found > 0:
                            c.execute("DELETE FROM pdi_kayitlari WHERE sasi_no=? AND is_emri_no=? AND tespitler=? AND tarih_saat=?",
                                     (item["sasi_no"], item["is_emri_no"], item["tespitler"], item["tarih"]))
                        
                        c.execute("INSERT INTO pdi_kayitlari (bb_no, sasi_no, arac_tipi, is_emri_no, alt_grup, tespitler, hata_konumu, top_hata, tarih_saat, kullanici) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                 (item["bb_no"], item["sasi_no"], item["arac_tipi"], item["is_emri_no"], item["alt_grup"], item["tespitler"], item["hata_konumu"], "", item["tarih"], self.current_user or "admin"))
                        success_count += 1
                    except Exception as ex:
                        result_text.insert("end", f"❌ Satır {item['idx']}: {str(ex)}\n")
                        error_count += 1
                
                conn.commit(); conn.close()
                result_text.insert("end", f"\n✅ İşlem Tamamlandı!\n✅ Başarılı: {success_count}\n❌ Hatalı: {error_count}\n")
                if duplicates_found > 0:
                    result_text.insert("end", f"ℹ️ {duplicates_found} kayıt güncellendi (üstüne yazıldı).\n")
                messagebox.showinfo("Bitti", f"{success_count} kayıt sisteme işlendi.")
            except Exception as e:
                messagebox.showerror("Hata", f"Excel hatası: {str(e)}")

        # Buton
        tk.Button(main_frame, text="YÜKLEMEYİ BAŞLAT", command=import_data, bg="black", fg="white", font=("Segoe UI", 12, "bold"), height=2).pack(fill="x", pady=10)
        
        # Sonuç alanını şimdi pack ediyoruz
        result_frame.pack(fill="both", expand=True, pady=10)
        result_text.pack(fill="both", expand=True, padx=5, pady=5)


    # ------------------------------------------------------
    # İMALATA GÖNDERİLEN ARAÇLAR (YENİ MODÜL)
    # ------------------------------------------------------
    
    def show_imalat_form(self):
        self.clear_content()
        self.create_report_header("İMALATA ARAÇ GÖNDERİM FORMU", self.show_imalat_form, show_filter=False)
        
        main_frame = tk.Frame(self.content_area, bg=COLOR_BG)
        main_frame.pack(fill="both", expand=True, padx=30, pady=10)
        
        card = tk.Frame(main_frame, bg=COLOR_WHITE, padx=40, pady=30, relief="flat", bd=0)
        card.place(relx=0.5, rely=0.45, anchor="center", width=700)
        
        # Gölgeli efekt için basit bir yöntem (opsiyonel)
        tk.Label(card, text="İMALAT GÖNDERİM DETAYLARI", font=("Segoe UI", 16, "bold"), bg=COLOR_WHITE, fg=COLOR_ACCENT).grid(row=0, column=0, columnspan=2, pady=(0, 30), sticky="w")
        
        def add_form_row(label, row_idx, widget_class=ttk.Entry, use_calendar=False, **kwargs):
            tk.Label(card, text=label, font=FONT_BOLD, bg=COLOR_WHITE, fg="#555").grid(row=row_idx, column=0, sticky="w", pady=12, padx=(0, 20))
            frame = tk.Frame(card, bg=COLOR_WHITE)
            frame.grid(row=row_idx, column=1, sticky="ew", pady=12)
            
            w = widget_class(frame, font=FONT_NORMAL, **kwargs)
            w.pack(side="left", fill="x", expand=True)
            
            if use_calendar:
                def open_cal():
                    msg = AdvancedCalendarDialog(self.root, w.get())
                    self.root.wait_window(msg)
                    if msg.result: w.delete(0, tk.END); w.insert(0, msg.result)
                tk.Button(frame, text="📅", command=open_cal, bg="#eee", relief="flat", padx=10).pack(side="right", padx=(5, 0))
            return w

        e_tarih = add_form_row("İmalata Gitme Tarihi:", 1, use_calendar=True)
        e_tarih.insert(0, datetime.now().strftime("%d-%m-%Y"))
        
        e_arac = add_form_row("Araç No / Şasi:", 2)
        
        tk.Label(card, text="Top Hata Seçimi:", font=FONT_BOLD, bg=COLOR_WHITE, fg="#555").grid(row=4, column=0, sticky="w", pady=12, padx=(0, 20))
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        top_hatalar = [r[0] for r in c.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()]
        conn.close()
        e_top_hata = ttk.Combobox(card, values=top_hatalar, font=FONT_NORMAL)
        e_top_hata.grid(row=4, column=1, sticky="ew", pady=12)
        
        tk.Label(card, text="Hata Detayları:\n(Her satıra bir hata)", font=FONT_BOLD, bg=COLOR_WHITE, fg="#555", justify="left").grid(row=5, column=0, sticky="nw", pady=12, padx=(0, 20))
        txt_hatalar = tk.Text(card, font=FONT_NORMAL, height=6, bd=1, relief="solid", padx=10, pady=10)
        txt_hatalar.grid(row=5, column=1, sticky="ew", pady=12)
        
        card.columnconfigure(1, weight=1)

        def save():
            an = e_arac.get().strip()
            th = e_top_hata.get()
            hm = txt_hatalar.get("1.0", "end-1c").strip()
            t = e_tarih.get()
            
            if not an: messagebox.showwarning("Eksik Bilgi", "Lütfen Araç No giriniz."); return
            
            conn = sqlite3.connect(DB_NAME); c = conn.cursor()
            c.execute("INSERT INTO imalat_kayitlari (arac_no, tarih, adet, top_hata, hata_metni, kullanici, olusturma_tarihi) VALUES (?,?,?,?,?,?,?)",
                     (an, t, 1, th, hm, self.current_user, datetime.now().strftime("%d-%m-%Y %H:%M")))
            conn.commit(); conn.close()
            messagebox.showinfo("Başarılı", "İmalata gönderim kaydı oluşturuldu."); self.show_imalat_list()
            
        tk.Button(card, text="KAYDI SİSTEME İŞLE", command=save, bg="black", fg="white", font=("Segoe UI", 12, "bold"), height=2, relief="flat").grid(row=6, column=0, columnspan=2, pady=30, sticky="ew")

    def show_imalat_list(self):
        self.clear_content()
        self.create_report_header("İMALATA GÖNDERİLEN ARAÇ KAYITLARI", self.show_imalat_list, show_filter=False)
        
        main_container = tk.Frame(self.content_area, bg=COLOR_BG, padx=20, pady=10)
        main_container.pack(fill="both", expand=True)
        
        # ÜST: Liste Bölümü
        list_card = tk.Frame(main_container, bg=COLOR_WHITE, padx=10, pady=10, relief="flat")
        list_card.pack(fill="both", expand=True, pady=(0, 10))
        
        cols = ("ID", "Araç No", "Tarih", "Top Hata")
        self.tree_imalat = ttk.Treeview(list_card, columns=cols, show="headings", height=15)
        
        # Sıralama durumunu tutmak için
        imalat_sort_state = {}
        
        def sort_imalat_tree(col):
            reverse = imalat_sort_state.get(col, False)
            data = [(self.tree_imalat.set(child, col), child) for child in self.tree_imalat.get_children('')]
            try:
                data.sort(key=lambda x: float(x[0]) if x[0].replace('.','').replace('-','').isdigit() else x[0], reverse=reverse)
            except:
                data.sort(key=lambda x: x[0], reverse=reverse)
            for idx, (val, child) in enumerate(data):
                self.tree_imalat.move(child, '', idx)
            imalat_sort_state[col] = not reverse
            for c in cols:
                arrow = ""
                if c == col:
                    arrow = " ▼" if not reverse else " ▲"
                self.tree_imalat.heading(c, text=c.upper() + arrow, command=lambda c=c: sort_imalat_tree(c))
        
        for col in cols: 
            self.tree_imalat.heading(col, text=col.upper(), command=lambda c=col: sort_imalat_tree(c))
            self.tree_imalat.column(col, width=150, anchor="center")
        self.tree_imalat.column("Top Hata", width=350, anchor="w")
        
        # Middle Mouse Scroll
        self.tree_imalat.bind("<Button-2>", lambda e: self.tree_imalat.scan_mark(e.x, e.y))
        self.tree_imalat.bind("<B2-Motion>", lambda e: self.tree_imalat.scan_dragto(e.x, e.y, gain=1))
        
        self.tree_imalat.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(list_card, orient="vertical", command=self.tree_imalat.yview)
        self.tree_imalat.configure(yscroll=sb.set); sb.pack(side="right", fill="y")
        
        # Action Sidebar for Imalat List
        action_bar = tk.Frame(list_card, bg=COLOR_WHITE, padx=10)
        action_bar.pack(side="right", fill="y")
        
        tk.Button(action_bar, text="DÜZENLE", command=self.open_edit_imalat_window, bg="#ffc107", relief="flat", width=12, pady=5).pack(pady=5)
        if self.current_role == "admin":
            tk.Button(action_bar, text="SİL", command=self.delete_imalat_record, bg="#dc3545", fg="white", relief="flat", width=12, pady=5).pack(pady=5)
        
        # ALT: Detay Bölümü
        detail_card = tk.Frame(main_container, bg=COLOR_WHITE, height=200, relief="flat", padx=20, pady=15)
        detail_card.pack(fill="x")
        detail_card.pack_propagate(False)
        
        info_frame = tk.Frame(detail_card, bg=COLOR_WHITE)
        info_frame.pack(side="left", fill="both", expand=True)
        
        lbl_sel = tk.Label(info_frame, text="Lütfen listeden bir araç seçiniz...", font=FONT_BOLD, bg=COLOR_WHITE, fg="#666")
        lbl_sel.pack(anchor="w")
        
        detail_text = tk.Text(info_frame, font=("Segoe UI", 10), bg="#f9f9f9", height=6, bd=0, padx=10, pady=10)
        detail_text.pack(fill="both", expand=True, pady=10)
        detail_text.config(state="disabled")

        def on_tree_select(e):
            item = self.tree_imalat.selection()
            if not item: return
            rid = self.tree_imalat.item(item[0])['values'][0]
            
            conn = sqlite3.connect(DB_NAME); c = conn.cursor()
            row = c.execute("SELECT arac_no, tarih, top_hata, hata_metni FROM imalat_kayitlari WHERE id=?", (rid,)).fetchone()
            conn.close()
            
            if row:
                lbl_sel.config(text=f"Araç: {row[0]} | Tarih: {row[1]} | Top Hata: {row[2]}", fg="black")
                detail_text.config(state="normal")
                detail_text.delete("1.0", "end")
                detail_text.insert("end", f"HATA DETAYLARI:\n----------------\n{row[3]}")
                detail_text.config(state="disabled")
            
        self.tree_imalat.bind("<<TreeviewSelect>>", on_tree_select)
        
        # Verileri Yükle
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        for r in c.execute("SELECT id, arac_no, tarih, top_hata FROM imalat_kayitlari ORDER BY id DESC").fetchall():
            self.tree_imalat.insert("", "end", values=r)
        conn.close()
        
        # SIRALAMA FONKSİYONU
        self.tree_imalat_sort_state = {}
        cols = ["ID", "Araç No", "Tarih", "Top Hata"]
        
        def sort_imalat_tree(col):
            reverse = self.tree_imalat_sort_state.get(col, False)
            data = [(self.tree_imalat.set(child, col), child) for child in self.tree_imalat.get_children('')]
            
            def get_sort_key(item):
                val = item[0]
                if len(val) == 10 and val[2] == '-' and val[5] == '-':
                    try: parts = val.split('-'); return f"{parts[2]}-{parts[1]}-{parts[0]}"
                    except: pass
                try: return float(val)
                except: return val
            
            data.sort(key=get_sort_key, reverse=reverse)
            for idx, (val, child) in enumerate(data): self.tree_imalat.move(child, '', idx)
            self.tree_imalat_sort_state[col] = not reverse
            
            for c in cols:
                arrow = " ▼" if (c == col and not reverse) else (" ▲" if (c == col and reverse) else "")
                self.tree_imalat.heading(c, text=c + arrow, command=lambda c=c: sort_imalat_tree(c))
                
        for c in cols: self.tree_imalat.heading(c, text=c, command=lambda c=c: sort_imalat_tree(c))

    def delete_imalat_record(self):
        sel = self.tree_imalat.selection()
        if not sel: return
        if messagebox.askyesno("Sil", "İmalat kaydı silinsin mi?"):
            rid = self.tree_imalat.item(sel[0])['values'][0]
            conn = sqlite3.connect(DB_NAME); conn.execute("DELETE FROM imalat_kayitlari WHERE id=?", (rid,)); conn.commit(); conn.close(); self.show_imalat_list()

    def open_edit_imalat_window(self):
        sel = self.tree_imalat.selection()
        if not sel: return
        rid = self.tree_imalat.item(sel[0])['values'][0]
        conn = sqlite3.connect(DB_NAME); row = conn.execute("SELECT * FROM imalat_kayitlari WHERE id=?", (rid,)).fetchone(); conn.close()
        if row: self.imalat_popup_window("İmalat Kaydı Düzenle", row)

    def imalat_popup_window(self, title, data=None):
        win = tk.Toplevel(self.root); win.title(title); win.geometry("600x600"); win.configure(bg=COLOR_WHITE)
        tk.Label(win, text=title.upper(), font=FONT_HEADER, bg="black", fg="white", pady=10).pack(fill="x")
        
        main_f = tk.Frame(win, bg=COLOR_WHITE, padx=20, pady=20); main_f.pack(fill="both", expand=True)
        
        tk.Label(main_f, text="Tarih:", bg=COLOR_WHITE).pack(anchor="w")
        e_t = ttk.Entry(main_f); e_t.insert(0, data[2] if data else ""); e_t.pack(fill="x", pady=5)
        
        tk.Label(main_f, text="Araç No:", bg=COLOR_WHITE).pack(anchor="w")
        e_a = ttk.Entry(main_f); e_a.insert(0, data[1] if data else ""); e_a.pack(fill="x", pady=5)
        
        tk.Label(main_f, text="Top Hata:", bg=COLOR_WHITE).pack(anchor="w")
        conn = sqlite3.connect(DB_NAME); top_hatalar = [r[0] for r in conn.execute("SELECT hata_adi FROM top_hatalar WHERE aktif=1").fetchall()]; conn.close()
        cb_th = ttk.Combobox(main_f, values=top_hatalar); cb_th.set(data[4] if data else ""); cb_th.pack(fill="x", pady=5)
        
        tk.Label(main_f, text="Hata Detayları:", bg=COLOR_WHITE).pack(anchor="w")
        txt_h = tk.Text(main_f, height=8, font=FONT_NORMAL, bd=1, relief="solid"); txt_h.insert("1.0", data[5] if data else ""); txt_h.pack(fill="x", pady=5)
        
        def save():
            conn = sqlite3.connect(DB_NAME)
            if data:
                conn.execute("UPDATE imalat_kayitlari SET tarih=?, arac_no=?, top_hata=?, hata_metni=? WHERE id=?",
                            (e_t.get(), e_a.get(), cb_th.get(), txt_h.get("1.0", "end-1c"), data[0]))
            else:
                conn.execute("INSERT INTO imalat_kayitlari (tarih, arac_no, top_hata, hata_metni, adet) VALUES (?,?,?,?,?)",
                            (e_t.get(), e_a.get(), cb_th.get(), txt_h.get("1.0", "end-1c"), 1))
            conn.commit(); conn.close(); win.destroy(); self.show_imalat_list()
            
        tk.Button(win, text="KAYDET", command=save, bg="black", fg="white", font=FONT_BOLD, height=2, relief="flat").pack(fill="x", padx=20, pady=20)

    def show_imalat_reports(self):
        self.clear_content()
        self.create_report_header("İMALAT VERİ ANALİZİ VE RAPORLARI", self.show_imalat_reports)
        
        month, year = self.get_selected_month_year()
        if month is None: 
            messagebox.showwarning("Filtre Gerekli", "Bu raporları görüntülemek için lütfen sağ üstten bir AY VE YIL seçiniz.")
            return

        notebook = ttk.Notebook(self.content_area)
        notebook.pack(fill="both", expand=True, padx=20, pady=10)
        
        # --- TAB 1: AYLIK ÖZET TABLO ---
        tab1 = tk.Frame(notebook, bg=COLOR_BG); notebook.add(tab1, text="  AYLIK ÖZET TABLO  ")
        self.render_imalat_summary_tab(tab1, month, year)
        
        # --- TAB 2: GİDİŞ ORANLARI GRAFİĞİ ---
        tab2 = tk.Frame(notebook, bg=COLOR_BG); notebook.add(tab2, text="  İMALAT GİDİŞ ORANLARI  ")
        self.render_imalat_trend_tab(tab2, month, year)
        
        # --- TAB 3: TOP HATA ANALİZİ ---
        tab3 = tk.Frame(notebook, bg=COLOR_BG); notebook.add(tab3, text="  TOP HATA ANALİZİ  ")
        self.render_imalat_top_errors_tab(tab3, month, year)
        
        # --- TAB 4: ÇÖZÜLEN HATALAR ---
        tab4 = tk.Frame(notebook, bg=COLOR_BG); notebook.add(tab4, text="  ÇÖZÜLEN/İYİLEŞENLER  ")
        self.render_imalat_solved_tab(tab4, month, year)

    def render_imalat_summary_tab(self, parent, month, year):
        # Üst Metin Alanı
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        # Araç Adedi (Distinct arac_no)
        sql_arac = "SELECT COUNT(DISTINCT arac_no) FROM imalat_kayitlari WHERE CAST(substr(tarih,4,2) AS INTEGER)=? AND CAST(substr(tarih,7,4) AS INTEGER)=?"
        arac_count = c.execute(sql_arac, (month, year)).fetchone()[0]
        # Hata Adedi (Her kayıttaki hata_metni'nin satır sayısı + 1 (eğer doluysa) )
        recs = c.execute("SELECT hata_metni FROM imalat_kayitlari WHERE CAST(substr(tarih,4,2) AS INTEGER)=? AND CAST(substr(tarih,7,4) AS INTEGER)=?", (month, year)).fetchall()
        total_hata = 0
        for r in recs:
            if r[0]: lines = r[0].strip().split('\n'); total_hata += len(lines)
            else: total_hata += 1 # Eğer metin boşsa ama kayıt varsa en az 1 hata sayalım (veya 0)
            
        summary_text = f"{year} yılı {TURKISH_MONTHS[month]} ayında imalata {arac_count} adet araç gönderilmiş olup, {total_hata} adet hata tespit edilmiştir."
        
        text_frame = tk.Frame(parent, bg=COLOR_WHITE, padx=15, pady=10, relief="solid", bd=1)
        text_frame.pack(fill="x", padx=20, pady=20)
        
        summary_entry = tk.Entry(text_frame, font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE, bd=0, readonlybackground=COLOR_WHITE, fg="black")
        summary_entry.insert(0, summary_text)
        summary_entry.config(state="readonly")
        summary_entry.pack(side="left", fill="x", expand=True)
        
        # Tablo
        table_card = tk.Frame(parent, bg=COLOR_WHITE, padx=10, pady=10)
        table_card.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        cols = ("SIra", "Araç No", "İmalata Gitme Tarihi", "Hata")
        tree = ttk.Treeview(table_card, columns=cols, show="headings")
        tree.heading("SIra", text="Sıra")
        tree.heading("Araç No", text="Araç No")
        tree.heading("İmalata Gitme Tarihi", text="İmalata Gitme Tarihi")
        tree.heading("Hata", text="Hata")
        tree.column("SIra", width=50, anchor="center")
        tree.column("Hata", width=500)
        tree.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(table_card, orient="vertical", command=tree.yview)
        tree.configure(yscroll=sb.set); sb.pack(side="right", fill="y")
        
        rows = c.execute("SELECT arac_no, tarih, top_hata FROM imalat_kayitlari WHERE CAST(substr(tarih,4,2) AS INTEGER)=? AND CAST(substr(tarih,7,4) AS INTEGER)=? ORDER BY tarih ASC", (month, year)).fetchall()
        for idx, r in enumerate(rows):
            tree.insert("", "end", values=(idx+1, r[0], r[1], r[2]))
        conn.close()

    def render_imalat_trend_tab(self, parent, month, year):
        # Yıl Grubu Seçimi
        filter_frame = tk.Frame(parent, bg=COLOR_BG)
        filter_frame.pack(fill="x", padx=20, pady=(10, 0))
        
        tk.Label(filter_frame, text="Kıyaslama Yıl Grubu: ", bg=COLOR_BG, font=FONT_BOLD).pack(side="left")
        
        year_groups = []
        current_y = datetime.now().year
        for y in range(current_y, 2022, -1):
            year_groups.append(f"{y-1} - {y}")
            
        selected_group = tk.StringVar(value=f"{year-1} - {year}")
        cb_group = ttk.Combobox(filter_frame, textvariable=selected_group, values=year_groups, state="readonly", width=15)
        cb_group.pack(side="left", padx=10)
        
        # Metin Alanı (Dinamik olacak)
        text_frame = tk.Frame(parent, bg=COLOR_WHITE, padx=15, pady=10, relief="solid", bd=1)
        text_frame.pack(fill="x", padx=20, pady=20)
        
        summary_entry_trend = tk.Entry(text_frame, font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE, bd=0, readonlybackground=COLOR_WHITE, fg="black")
        summary_entry_trend.pack(side="left", fill="x", expand=True)
        
        # Grafik Alanı
        chart_frame = tk.Frame(parent, bg=COLOR_WHITE, highlightbackground="#ddd", highlightthickness=1)
        chart_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        canvas = tk.Canvas(chart_frame, bg=COLOR_WHITE)
        canvas.pack(fill="both", expand=True)
        
        def update_view(event=None):
            # Define all variables at start of function to avoid NameError via late assignment
            x_rate, y_rate = 0, 0
            x_val, y_val = 0, 0
            
            grp_data = selected_group.get().split(" - ")
            if len(grp_data) < 2: return
            try:
                y_base = int(grp_data[0])
                y_target = int(grp_data[1])
            except ValueError: return
            
            # Veri Hesabı - OBTİMİZE EDİLMİŞ (PANDAS)
            conn = self.get_db_connection()
            
            # PDI ve İmalat kayıtlarını çek (Sadece gerekli yılları)
            pdi_query = "SELECT bb_no, tarih_saat FROM pdi_kayitlari"
            im_query = "SELECT arac_no, tarih FROM imalat_kayitlari"
            
            df_pdi = pd.read_sql_query(pdi_query, conn)
            df_im = pd.read_sql_query(im_query, conn)
            
            # Tarih parse ve normalize
            df_pdi['dt'] = pd.to_datetime(df_pdi['tarih_saat'], format='%d-%m-%Y', errors='coerce')
            df_im['dt'] = pd.to_datetime(df_im['tarih'], format='%d-%m-%Y', errors='coerce')
            
            df_pdi = df_pdi.dropna(subset=['dt'])
            df_im = df_im.dropna(subset=['dt'])
            
            def get_stats_for_year(y):
                stats = []
                for m in range(1, 13):
                    pdi_count = df_pdi[(df_pdi['dt'].dt.year == y) & (df_pdi['dt'].dt.month == m)]['bb_no'].nunique()
                    im_count = df_im[(df_im['dt'].dt.year == y) & (df_im['dt'].dt.month == m)]['arac_no'].nunique()
                    rate = (im_count / pdi_count * 100.0) if pdi_count > 0 else 0
                    stats.append(rate)
                return stats
            
            data_curr = get_stats_for_year(y_target)
            data_prev = get_stats_for_year(y_base)
            
            # Sadece veri olan ayların ortalamasını al (0 olmayan değerler)
            non_zero_curr = [v for v in data_curr if v > 0]
            x_rate = sum(non_zero_curr) / len(non_zero_curr) if non_zero_curr else 0
            y_rate = data_curr[month-1] if (month and 0 < month <= 12 and month <= len(data_curr)) else 0

            # Manual Override for Summary
            db_x = self.get_manual_data("imalat_trend_summary", "yearly_avg", str(y_target))
            db_y = self.get_manual_data("imalat_trend_summary", "monthly_rate", f"{month}-{y_target}")
            
            try:
                x_val = float(db_x) if db_x and str(db_x).strip() else x_rate
                y_val = float(db_y) if db_y and str(db_y).strip() else y_rate
            except ValueError:
                x_val, y_val = x_rate, y_rate
            
            trend_text = f"PDI yapılan araçların, {y_target} yılında ortalama %{x_val:.1f}'u, {y_target} {TURKISH_MONTHS.get(month, '')} ayı özelinde ise %{y_val:.1f}'si imalata geri gönderilmiştir."
            summary_entry_trend.config(state="normal")
            summary_entry_trend.delete(0, tk.END)
            summary_entry_trend.insert(0, trend_text)
            summary_entry_trend.config(state="readonly")
            
            def edit_summary():
                ewin = tk.Toplevel(self.root); ewin.title("Özet Metni Düzenle"); ewin.geometry("400x300")
                tk.Label(ewin, text=f"{y_target} Yıllık Ortalama %:").pack(pady=5)
                e_x = ttk.Entry(ewin); e_x.insert(0, f"{x_val:.1f}"); e_x.pack()
                tk.Label(ewin, text=f"{y_target} {TURKISH_MONTHS.get(month, '')} Ayı Oranı %:").pack(pady=5)
                e_y = ttk.Entry(ewin); e_y.insert(0, f"{y_val:.1f}"); e_y.pack()
                
                def save_s():
                    self.save_manual_data("imalat_trend_summary", "yearly_avg", str(y_target), e_x.get())
                    self.save_manual_data("imalat_trend_summary", "monthly_rate", f"{month}-{y_target}", e_y.get())
                    ewin.destroy(); update_view()
                tk.Button(ewin, text="KAYDET", command=save_s).pack(pady=20)
            
            # Use local reference to avoid hasattr check re-creation issues
            for widget in text_frame.winfo_children():
                if isinstance(widget, tk.Button) and widget.cget("text") == "📝":
                    widget.config(command=edit_summary)
                    break
            else:
                tk.Button(text_frame, text="📝", command=edit_summary, bg="#eee", relief="flat").pack(side="right", padx=5)
            
            
            # Manual Data Override
            final_data_curr = []
            final_data_prev = []
            for m in range(1, 13):
                db_c = self.get_manual_data("imalat_trend", f"curr_{m}", f"{y_base}-{y_target}")
                db_p = self.get_manual_data("imalat_trend", f"prev_{m}", f"{y_base}-{y_target}")
                
                try:
                    f_curr = float(db_val) if (db_val := db_c) and str(db_val).strip() else data_curr[m-1]
                    f_prev = float(db_val) if (db_val := db_p) and str(db_val).strip() else data_prev[m-1]
                except ValueError:
                    f_curr, f_prev = data_curr[m-1], data_prev[m-1]
                    
                final_data_curr.append(f_curr)
                final_data_prev.append(f_prev)

            # Grafiği Çiz
            canvas.delete("all")
            canvas.update_idletasks()
            w = canvas.winfo_width(); h = canvas.winfo_height()
            if w < 100: w = 800
            if h < 100: h = 400
            
            padding = 60; plot_w = w - 2*padding; plot_h = h - 2*padding
            max_val = max(max(final_data_curr), max(final_data_prev), 10)
            
            # Edit Icon
            canvas.create_text(w - 50, 30, text="📝", font=("Segoe UI", 12), fill="#666", tags="edit_imalat_trend")
            def edit_imalat_trend_dialog():
                ewin = tk.Toplevel(self.root)
                ewin.title(f"{y_base} - {y_target} Verilerini Düzenle")
                ewin.geometry("400x600")
                
                # Scrollable
                c_e = tk.Canvas(ewin); c_e.pack(side="left", fill="both", expand=True)
                vs = ttk.Scrollbar(ewin, orient="vertical", command=c_e.yview); vs.pack(side="right", fill="y"); sf = tk.Frame(c_e); c_e.create_window((0,0), window=sf, anchor="nw")
                
                ents = []
                tk.Label(sf, text="Ay", font=FONT_BOLD).grid(row=0, column=0)
                tk.Label(sf, text=f"{y_base} %", font=FONT_BOLD).grid(row=0, column=1)
                tk.Label(sf, text=f"{y_target} %", font=FONT_BOLD).grid(row=0, column=2)
                
                for m in range(1, 13):
                    r = m
                    tk.Label(sf, text=TURKISH_MONTHS[m][:3]).grid(row=r, column=0)
                    e_p = ttk.Entry(sf, width=10); e_p.insert(0, f"{final_data_prev[m-1]:.2f}"); e_p.grid(row=r, column=1)
                    e_c = ttk.Entry(sf, width=10); e_c.insert(0, f"{final_data_curr[m-1]:.2f}"); e_c.grid(row=r, column=2)
                    ents.append((e_p, e_c))
                    
                def save_e():
                    for i, (ep, ec) in enumerate(ents):
                        m_idx = i + 1
                        self.save_manual_data("imalat_trend", f"prev_{m_idx}", f"{y_base}-{y_target}", ep.get())
                        self.save_manual_data("imalat_trend", f"curr_{m_idx}", f"{y_base}-{y_target}", ec.get())
                    ewin.destroy()
                    update_view()
                
                tk.Button(ewin, text="KAYDET", command=save_e).pack(pady=10)
                sf.update_idletasks(); c_e.config(scrollregion=c_e.bbox("all"))

            canvas.tag_bind("edit_imalat_trend", "<Button-1>", lambda e: edit_imalat_trend_dialog())

            # Izgara
            for i in range(13):
                x = padding + i * (plot_w / 12)
                canvas.create_line(x, padding, x, h-padding, fill="#f0f0f0")
                if i < 12: canvas.create_text(x + plot_w/24, h-padding+20, text=TURKISH_MONTHS[i+1][:3], font=("Segoe UI", 8))
            
            for i in range(6):
                y = h - padding - i * (plot_h / 5)
                canvas.create_line(padding, y, w-padding, y, fill="#f0f0f0")
                canvas.create_text(padding-25, y, text=f"%{i*(max_val/5):.0f}", font=("Segoe UI", 8))
                
            def draw_line(data, color, label):
                points = []
                for i, v in enumerate(data):
                    x = padding + i * (plot_w / 12) + plot_w/24
                    y = h - padding - (v / max_val * plot_h)
                    points.extend([x, y])
                    canvas.create_oval(x-4, y-4, x+4, y+4, fill=color, outline=color)
                    canvas.create_text(x, y-15, text=f"%{v:.1f}", font=("Segoe UI", 8, "bold"), fill=color)
                if len(points) >= 4:
                    canvas.create_line(points, fill=color, width=2, smooth=False)
                
            draw_line(final_data_prev, "#aaaaaa", f"{y_base} Oranı")
            draw_line(final_data_curr, "#ff7f0e", f"{y_target} Oranı")
            
            # Legend
            canvas.create_rectangle(w-200, 30, w-185, 45, fill="#ff7f0e", outline="")
            canvas.create_text(w-130, 37, text=f"{y_target} Yılı Oranı", font=FONT_BOLD)
            canvas.create_rectangle(w-200, 55, w-185, 70, fill="#aaaaaa", outline="")
            canvas.create_text(w-130, 62, text=f"{y_base} Yılı Oranı", font=FONT_BOLD)

        cb_group.bind("<<ComboboxSelected>>", update_view)
        canvas.bind("<Configure>", update_view)
        self.root.after(200, update_view)

    def render_imalat_top_errors_tab(self, parent, month, year):
        # 3.ssteki Top Hata tablosu
        # Metin: "a Kasım ayı itibarıyla; imalata gönderilen x araçta y adet hata tespit edilmiştir."
        conn = self.get_db_connection()
        
        # OBTİMİZE EDİLMİŞ (PANDAS)
        im_query = "SELECT arac_no, tarih, top_hata, hata_metni FROM imalat_kayitlari"
        df_im = pd.read_sql_query(im_query, conn)
        df_im['dt'] = pd.to_datetime(df_im['tarih'], format='%d-%m-%Y', errors='coerce')
        df_im = df_im.dropna(subset=['dt'])
        
        # YTD ve Ay verisi filtrele
        mask_ytd = (df_im['dt'].dt.year == year) & (df_im['dt'].dt.month <= month)
        df_ytd = df_im[mask_ytd]
        
        arac_count = df_ytd['arac_no'].nunique()
        
        total_hata = 0
        for _, row in df_ytd.iterrows():
            h_metni = row['hata_metni']
            if h_metni: 
                total_hata += len(str(h_metni).strip().split('\n'))
            else: 
                total_hata += 1
            
        top_text = f"{year} {TURKISH_MONTHS[month]} ayı itibarıyla; imalata gönderilen {arac_count} araçta {total_hata} adet hata tespit edilmiştir."
        
        text_frame = tk.Frame(parent, bg=COLOR_WHITE, padx=15, pady=10, relief="solid", bd=1)
        text_frame.pack(fill="x", padx=20, pady=20)
        
        top_entry = tk.Entry(text_frame, font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE, bd=0, readonlybackground=COLOR_WHITE, fg="black")
        top_entry.insert(0, top_text)
        top_entry.config(state="readonly")
        top_entry.pack(side="left", fill="x", expand=True)
        
        # Tablo
        grid_container = tk.Frame(parent, bg=COLOR_WHITE)
        grid_container.pack(fill="x", padx=100, pady=20)
        
        tk.Label(grid_container, text=f"{year} İmalata Gönderilen Araçlarda Top Hata Analizi", font=("Segoe UI", 12, "bold"), bg=COLOR_WHITE).pack(pady=10)
        
        headers = ["İMALATA GÖNDERİLEN ARAÇLARDA TOP HATA", "HATA SAYISI", "ORAN", "İŞLEM"]
        h_frame = tk.Frame(grid_container, bg="#4f5b66")
        h_frame.pack(fill="x")
        for i, h in enumerate(headers):
             tk.Label(h_frame, text=h, bg="#4f5b66", fg="white", font=FONT_BOLD, pady=10).grid(row=0, column=i, sticky="nsew")
        h_frame.columnconfigure(0, weight=3); h_frame.columnconfigure(1, weight=1); h_frame.columnconfigure(2, weight=1); h_frame.columnconfigure(3, weight=1)
        
        def edit_imalat_top():
            win = tk.Toplevel(self.root)
            win.title("İmalat Top Hata Düzenle")
            win.geometry("600x500")
            
            main_f = tk.Frame(win, padx=20, pady=20)
            main_f.pack(fill="both", expand=True)
            
            tk.Label(main_f, text="No", font=FONT_BOLD).grid(row=0, column=0, padx=5)
            tk.Label(main_f, text="Hata Tanımı", font=FONT_BOLD).grid(row=0, column=1, padx=5)
            tk.Label(main_f, text="Sayı", font=FONT_BOLD).grid(row=0, column=2, padx=5)
            
            entries = []
            for i in range(10):
                tk.Label(main_f, text=f"{i+1}").grid(row=i+1, column=0, pady=2)
                e1 = ttk.Entry(main_f, width=40)
                e2 = ttk.Entry(main_f, width=10)
                
                db_h = self.get_manual_data("imalat_top", f"hata_{i}", f"{month}-{year}")
                db_c = self.get_manual_data("imalat_top", f"count_{i}", f"{month}-{year}")
                
                if db_h: e1.insert(0, db_h)
                if db_c: e2.insert(0, db_c)
                
                e1.grid(row=i+1, column=1, padx=5)
                e2.grid(row=i+1, column=2, padx=5)
                entries.append((e1, e2))
            
            def save_imalat_top_local():
                for i, (e1, e2) in enumerate(entries):
                    self.save_manual_data("imalat_top", f"hata_{i}", f"{month}-{year}", e1.get())
                    self.save_manual_data("imalat_top", f"count_{i}", f"{month}-{year}", e2.get())
                self.show_imalat_reports()
                win.destroy()
            
            tk.Button(main_f, text="GÜNCELLE", command=save_imalat_top_local, bg="black", fg="white", font=FONT_BOLD, width=15, pady=10).grid(row=11, column=0, columnspan=3, pady=20)

        # Add edit button to a row or just a floating button
        tk.Button(grid_container, text="📝 Tabloyu Düzenle", command=edit_imalat_top, bg="#eee", relief="flat").pack(pady=5)
        
        # Data OBTİMİZE EDİLMİŞ
        df_top = df_ytd[df_ytd['top_hata'].notna() & (df_ytd['top_hata'] != '')]
        data_rows = df_top.groupby('top_hata').size().sort_values(ascending=False).head(10).reset_index().values.tolist()
        
        for idx in range(10):
            db_h = self.get_manual_data("imalat_top", f"hata_{idx}", f"{month}-{year}")
            db_c = self.get_manual_data("imalat_top", f"count_{idx}", f"{month}-{year}")
            
            if db_h or (idx < len(data_rows)):
                hata = db_h if db_h and db_h.strip() else (data_rows[idx][0] if idx < len(data_rows) else "")
                cnt = int(db_cnt) if (db_cnt := db_c) and str(db_cnt).strip() else (data_rows[idx][1] if idx < len(data_rows) else 0)
                
                bg = "white" if idx % 2 == 0 else "#f5f5f5"
                r_frame = tk.Frame(grid_container, bg=bg)
                r_frame.pack(fill="x")
                
                oran = (cnt / arac_count * 100) if arac_count > 0 else 0
                
                tk.Label(r_frame, text=hata, bg=bg, font=FONT_NORMAL, pady=8, anchor="w").grid(row=0, column=0, sticky="nsew", padx=10)
                tk.Label(r_frame, text=str(cnt), bg=bg, font=FONT_NORMAL).grid(row=0, column=1, sticky="nsew")
                tk.Label(r_frame, text=f"%{oran:.1f}", bg=bg, font=FONT_NORMAL).grid(row=0, column=2, sticky="nsew")
                
                def del_row(i):
                    self.save_manual_data("imalat_top", f"hata_{i}", f"{month}-{year}", "")
                    self.save_manual_data("imalat_top", f"count_{i}", f"{month}-{year}", "")
                    self.show_imalat_reports()
                
                if db_h:
                    tk.Button(r_frame, text="🗑", command=lambda i=idx: del_row(i), bg=bg, fg="red", bd=0).grid(row=0, column=3)
                else:
                    tk.Label(r_frame, text="-", bg=bg).grid(row=0, column=3)
                
                r_frame.columnconfigure(0, weight=3); r_frame.columnconfigure(1, weight=1); r_frame.columnconfigure(2, weight=1); r_frame.columnconfigure(3, weight=1)
            

    def render_imalat_solved_tab(self, parent, month, year):
        # Şık ve modern tasarım - Sunum için hazır
        
        # Üst başlık
        header_frame = tk.Frame(parent, bg=COLOR_BG)
        header_frame.pack(fill="x", pady=(20, 10))
        
        title_text = f"{year} {TURKISH_MONTHS[month]} sonu itibarı ile iyileşen/çözülen hatalar:"
        tk.Label(header_frame, text=title_text, font=("Segoe UI", 14), bg=COLOR_BG, fg="#333").pack()
        
        main_frame = tk.Frame(parent, bg=COLOR_BG)
        main_frame.pack(fill="both", expand=True, padx=50)
        
        # Giriş Alanı - Üstte minimal
        input_frame = tk.Frame(main_frame, bg=COLOR_WHITE, padx=15, pady=12)
        input_frame.pack(fill="x", pady=(0, 15))
        
        tk.Label(input_frame, text="➕ Yeni Hata Ekle:", font=("Segoe UI", 10, "bold"), bg=COLOR_WHITE, fg="#555").pack(side="left", padx=(0, 10))
        e_hata = ttk.Entry(input_frame, width=50, font=("Segoe UI", 10))
        e_hata.pack(side="left", padx=5)
        
        grid_container = tk.Frame(main_frame, bg=COLOR_WHITE)
        grid_container.pack(fill="both", expand=True)
        
        # Seçili öğe ID'sini tutmak için
        selected_id = tk.IntVar(value=-1)
        
        def clear_selection(event=None):
            selected_id.set(-1)
            # Tüm satırların arka plan rengini sıfırla
            for child in grid_container.winfo_children():
                if hasattr(child, 'original_bg'):
                    child.configure(bg=child.original_bg)
                    for w in child.winfo_children():
                        if isinstance(w, tk.Label) and hasattr(w, 'cget') and "Solved" not in str(w.cget("text")):
                            try:
                                w.configure(bg=child.original_bg)
                            except: pass
        
        # Ana frame'e ve boş alanlara tıklayınca seçimi temizle
        main_frame.bind("<Button-1>", clear_selection)
        grid_container.bind("<Button-1>", clear_selection)
        
        def refresh_list():
            for w in grid_container.winfo_children(): w.destroy()
            selected_id.set(-1)
            
            # Şık başlık satırı
            h_frame = tk.Frame(grid_container, bg="#2c3e50")
            h_frame.pack(fill="x")
            
            tk.Label(h_frame, text=f"{year-1} → {year} YTD{month} İyileşen/Çözülen Hatalar", 
                    bg="#2c3e50", fg="white", font=("Segoe UI", 11, "bold"), pady=12, anchor="w").pack(side="left", fill="x", expand=True, padx=15)
            tk.Label(h_frame, text="Güncel Durum", bg="#2c3e50", fg="white", font=("Segoe UI", 11, "bold"), pady=12, width=14, anchor="center").pack(side="right", padx=20)
            
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("CREATE TABLE IF NOT EXISTS manuel_cozulenler (id INTEGER PRIMARY KEY, hata TEXT, donem TEXT)")
            donem_str = f"{month}-{year}"
            rows = c.execute("SELECT id, hata FROM manuel_cozulenler WHERE donem=?", (donem_str,)).fetchall()
            
            if not rows:
                empty_frame = tk.Frame(grid_container, bg="#f8f9fa", pady=40)
                empty_frame.pack(fill="x")
                tk.Label(empty_frame, text="📋 Henüz çözülen hata kaydı eklenmemiş.", font=("Segoe UI", 11), bg="#f8f9fa", fg="#888").pack()
            
            for idx, (rid, hata) in enumerate(rows):
                bg = "#ffffff" if idx % 2 == 0 else "#f8f9fa"
                r_frame = tk.Frame(grid_container, bg=bg, cursor="hand2")
                r_frame.pack(fill="x")
                
                # Seçim için tıklama
                def make_select_handler(record_id, frame):
                    def on_select(event=None):
                        # Önceki seçimi temizle
                        for child in grid_container.winfo_children():
                            if hasattr(child, 'original_bg'):
                                child.configure(bg=child.original_bg)
                                for w in child.winfo_children():
                                    if isinstance(w, tk.Label):
                                        w.configure(bg=child.original_bg)
                        # Yeni seçimi uygula
                        selected_id.set(record_id)
                        frame.configure(bg="#d4edda")
                        for w in frame.winfo_children():
                            if isinstance(w, tk.Label) and "Solved" not in w.cget("text"):
                                w.configure(bg="#d4edda")
                    return on_select
                
                r_frame.original_bg = bg
                r_frame.bind("<Button-1>", make_select_handler(rid, r_frame))
                
                # Sağda - Solved badge (Güncel Durum ile aynı hizada)
                status_label = tk.Label(r_frame, text="✓ Solved", bg="#27ae60", fg="white", 
                        font=("Segoe UI", 10, "bold"), width=10, pady=4)
                status_label.pack(side="right", pady=10, padx=20)
                
                # Hata metni - sol tarafta
                hata_label = tk.Label(r_frame, text=f"• {hata}", bg=bg, font=("Segoe UI", 11), pady=14, anchor="w", fg="#333")
                hata_label.pack(side="left", fill="x", expand=True, padx=15)
                hata_label.bind("<Button-1>", make_select_handler(rid, r_frame))
                
            conn.close()

        def add_hata():
            h = e_hata.get().strip(); donem = f"{month}-{year}"
            if not h: return
            conn = sqlite3.connect(DB_NAME); conn.execute("INSERT INTO manuel_cozulenler (hata, donem) VALUES (?,?)", (h, donem)); conn.commit(); conn.close()
            e_hata.delete(0, tk.END)
            refresh_list()
        
        def delete_selected():
            sel_id = selected_id.get()
            if sel_id <= 0:
                messagebox.showwarning("Uyarı", "Lütfen önce listeden silmek istediğiniz öğeyi seçin.")
                return
            conn = sqlite3.connect(DB_NAME)
            conn.execute("DELETE FROM manuel_cozulenler WHERE id=?", (sel_id,))
            conn.commit()
            conn.close()
            refresh_list()

        tk.Button(input_frame, text="EKLE", command=add_hata, bg="#3498db", fg="white", 
                 font=("Segoe UI", 10, "bold"), relief="flat", padx=15, cursor="hand2").pack(side="left", padx=10)
        tk.Button(input_frame, text="SİL", command=delete_selected, bg="#e74c3c", fg="white", 
                 font=("Segoe UI", 10, "bold"), relief="flat", padx=15, cursor="hand2").pack(side="left", padx=5)
        
        refresh_list()
if __name__ == "__main__":
    db_folder = os.path.dirname(os.path.abspath(DB_NAME))
    if not os.access(db_folder, os.W_OK):
         root = tk.Tk()
         root.withdraw()
         messagebox.showerror("Bağlantı Hatası", 
                              "Veritabanı konumuna erişilemiyor!\n\n"
                              "Lütfen şunları kontrol edin:\n"
                              "1. VPN veya Şirket Ağına bağlı mısınız?\n"
                              "2. Ortak klasöre erişim izniniz var mı?")
         root.destroy()
    else:
        init_db()  # Veritabanı tablolarını oluştur
        root = tk.Tk()
        try: from ctypes import windll; windll.shcore.SetProcessDpiAwareness(1)
        except: pass
        app = PDIApp(root)
        
        # Uygulama kapanırken temizlik yap
        def on_closing():
            try:
                app.close_db_connection()
                app.clear_cache()
            except: pass
            root.destroy()
        
        root.protocol("WM_DELETE_WINDOW", on_closing)
        root.mainloop()